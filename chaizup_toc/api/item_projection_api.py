# =============================================================================
# CONTEXT: Item Projection View — Page-facing whitelisted API.
#   Single surface: Frappe Page at
#   `chaizup_toc/chaizup_toc/page/item_projection_view/`. Compute layer is
#   `chaizup_toc.api.item_projection_compute`. This file is a routing shim
#   so the page never touches the compute path directly (clean perm boundary
#   + identical pattern to item_shortage_api).
#
# MEMORY: app_chaizup_toc.md § Item Projection View
#
# INSTRUCTIONS:
#   - `get_dashboard_data(filters)` returns flat dict consumed by Tabulator.
#   - `get_breakdown(column, item_code, warehouse=None)` is the drill-down.
#     Column is one of: shortage_physical / shortage_projected /
#     wo_remaining_production / po_remaining / will_consume_open_wo /
#     will_dispatch_pending_so / net_available / days_of_cover.
#   - `get_filter_options()` bootstraps the filter dropdowns.
#   - `export_xlsx(filters)` streams the branded multi-sheet workbook.
#
# DANGER ZONE:
#   - frappe.parse_json('') raises → guard with `or {}`.
#   - The drill-down SQL re-uses the same `_build_status_clause` semantics
#     as the compute path; if you fork it here, the breakdown will not match
#     the cell total and users will lose trust.
#
# RESTRICT:
#   - Keep this file thin. If logic grows, factor into the compute module.
# =============================================================================

import json

import frappe
from frappe import _
from frappe.utils import flt


def _parse_filters(filters):
    if not filters:
        return {}
    if isinstance(filters, dict):
        return filters
    if isinstance(filters, str):
        try:
            parsed = json.loads(filters)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}


# =============================================================================
# 1. PAGE BOOTSTRAP + REFRESH
# =============================================================================
@frappe.whitelist()
def get_dashboard_data(filters=None):
    """
    Single backend call per page refresh. Returns:
        {
          "columns":      [...],     # column metadata, includes drilldown + numeric flags
          "rows":         [...],     # group headers + leaves OR flat leaves
          "banner":       "<HTML…>", # pending-status chip strip
          "summary":      [...],     # 6 KPI cards
          "filters_used": {...},     # echoed back so client can render chips
          "row_count":    int,
        }
    """
    from chaizup_toc.api import item_projection_compute as ipc
    f = _parse_filters(filters)
    columns, rows, banner, chart, summary = ipc.execute(f)
    # 2026-05-18 — Backend always returns flat leaf rows now (grouping
    # moved client-side to Tabulator's native groupBy). row_count is
    # simply len(rows).
    leaf_count = len(rows)
    return {
        "columns":       columns,
        "rows":          rows or [],
        "banner":        banner or "",
        "summary":       summary or [],
        "chart":         chart,
        "filters_used":  f,
        "row_count":     leaf_count,
    }


# =============================================================================
# 2. FILTER OPTIONS
# =============================================================================
@frappe.whitelist()
def get_filter_options():
    """Bootstrap dropdown payload for the page filter bar."""
    companies   = frappe.get_all("Company", pluck="name", order_by="name")
    warehouses  = frappe.get_all(
        "Warehouse",
        filters={"disabled": 0, "is_group": 0},
        pluck="name", order_by="name",
    )
    item_groups = frappe.get_all("Item Group", pluck="name", order_by="name")
    default_company = frappe.defaults.get_user_default("Company") or (
        companies[0] if companies else "")
    try:
        from chaizup_toc.api.wo_kitting_api import get_toc_pending_filters
        pending = get_toc_pending_filters()
    except Exception:
        pending = {"wo": [], "so": [], "po": [], "edit_route": "/app/toc-settings"}
    return {
        "companies":       companies,
        "warehouses":      warehouses,
        "item_groups":     item_groups,
        "default_company": default_company,
        "pending":         pending,
    }


# =============================================================================
# 3. DRILL-DOWNS — one helper per drillable column.
# =============================================================================
def _toc_pending_lists():
    from chaizup_toc.api.item_projection_compute import _toc_pending_lists as _f
    return _f()


def _wh_clause(prefix_col: str, warehouse: str | None):
    if not warehouse:
        return "", {}
    return f" AND {prefix_col} = %(_wh)s", {"_wh": warehouse}


def _bd_wo_production(item_code, warehouse):
    """Per-WO breakdown of pending PRODUCTION qty (this item is FG)."""
    from chaizup_toc.api.item_shortage_compute import _build_status_clause
    wo_plain, wo_wf, *_ = _toc_pending_lists()
    clause, sp = _build_status_clause(wo_plain, wo_wf, "Work Order",
                                      "tabWork Order", "wo")
    wh_sql, wh_p = _wh_clause("wo.fg_warehouse", warehouse)
    p = {"item": item_code, **sp, **wh_p}
    return frappe.db.sql(f"""
        SELECT wo.name        AS voucher_name,
               'Work Order'   AS voucher_type,
               wo.status,
               wo.fg_warehouse AS warehouse,
               wo.qty                                 AS planned_qty,
               IFNULL(wo.produced_qty, 0)             AS produced_qty,
               GREATEST(wo.qty - IFNULL(wo.produced_qty, 0), 0) AS remaining_qty,
               wo.planned_start_date,
               wo.expected_delivery_date
        FROM `tabWork Order` wo
        WHERE wo.production_item = %(item)s
          AND {clause}
          {wh_sql}
        ORDER BY wo.planned_start_date ASC, wo.name ASC
    """, p, as_dict=True)


def _bd_wo_consumption(item_code, warehouse):
    """Per-WO breakdown of pending CONSUMPTION (this item is component)."""
    from chaizup_toc.api.item_shortage_compute import _build_status_clause
    wo_plain, wo_wf, *_ = _toc_pending_lists()
    clause, sp = _build_status_clause(wo_plain, wo_wf, "Work Order",
                                      "tabWork Order", "wo")
    wh_sql, wh_p = _wh_clause("woi.source_warehouse", warehouse)
    p = {"item": item_code, **sp, **wh_p}
    return frappe.db.sql(f"""
        SELECT wo.name                                   AS voucher_name,
               'Work Order'                              AS voucher_type,
               wo.status,
               woi.source_warehouse                      AS warehouse,
               wo.production_item                        AS produces_item,
               woi.required_qty,
               IFNULL(woi.transferred_qty, 0)            AS transferred_qty,
               GREATEST(woi.required_qty
                        - IFNULL(woi.transferred_qty, 0), 0) AS remaining_qty,
               wo.planned_start_date
        FROM `tabWork Order Item` woi
        JOIN `tabWork Order` wo ON wo.name = woi.parent
        WHERE woi.item_code = %(item)s
          AND {clause}
          {wh_sql}
        ORDER BY wo.planned_start_date ASC, wo.name ASC
    """, p, as_dict=True)


def _bd_po(item_code, warehouse):
    """Per-PO Item row with remaining qty."""
    _, _, po_plain, po_wf, *_ = _toc_pending_lists()
    from chaizup_toc.api.item_shortage_compute import _build_status_clause
    clause, sp = _build_status_clause(po_plain, po_wf, "Purchase Order",
                                      "tabPurchase Order", "po")
    wh_sql, wh_p = _wh_clause("poi.warehouse", warehouse)
    p = {"item": item_code, **sp, **wh_p}
    return frappe.db.sql(f"""
        SELECT po.name                                   AS voucher_name,
               'Purchase Order'                          AS voucher_type,
               po.status,
               po.supplier,
               poi.warehouse,
               poi.qty,
               IFNULL(poi.received_qty, 0)               AS received_qty,
               GREATEST(poi.qty - IFNULL(poi.received_qty, 0), 0) AS remaining_qty,
               IFNULL(poi.conversion_factor, 1)          AS cf,
               GREATEST((poi.qty - IFNULL(poi.received_qty,0))
                        * IFNULL(poi.conversion_factor,1), 0) AS remaining_stock_qty,
               poi.schedule_date,
               po.transaction_date
        FROM `tabPurchase Order Item` poi
        JOIN `tabPurchase Order` po ON po.name = poi.parent
        WHERE poi.item_code = %(item)s
          AND {clause}
          {wh_sql}
        ORDER BY poi.schedule_date ASC, po.name ASC
    """, p, as_dict=True)


def _bd_so(item_code, warehouse):
    """Per-SO Item row with remaining dispatch qty."""
    _, _, _, _, so_plain, so_wf = _toc_pending_lists()
    from chaizup_toc.api.item_shortage_compute import _build_status_clause
    clause, sp = _build_status_clause(so_plain, so_wf, "Sales Order",
                                      "tabSales Order", "so")
    wh_sql, wh_p = _wh_clause(
        "COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse)", warehouse)
    p = {"item": item_code, **sp, **wh_p}
    return frappe.db.sql(f"""
        SELECT so.name                                                 AS voucher_name,
               'Sales Order'                                           AS voucher_type,
               so.status,
               so.customer,
               COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse)    AS warehouse,
               soi.qty,
               soi.stock_qty,
               IFNULL(soi.delivered_qty, 0)                            AS delivered_qty,
               GREATEST(soi.stock_qty - IFNULL(soi.delivered_qty,0)
                        * IFNULL(soi.conversion_factor,1), 0)          AS remaining_stock_qty,
               IFNULL(soi.conversion_factor, 1)                        AS cf,
               soi.delivery_date,
               so.transaction_date
        FROM `tabSales Order Item` soi
        JOIN `tabSales Order` so ON so.name = soi.parent
        WHERE soi.item_code = %(item)s
          AND {clause}
          {wh_sql}
        ORDER BY soi.delivery_date ASC, so.name ASC
    """, p, as_dict=True)


def _bd_stock(item_code, warehouse):
    """Per-Bin row contributing to current stock."""
    where = ["item_code = %(item)s"]
    p = {"item": item_code}
    if warehouse:
        where.append("warehouse = %(wh)s")
        p["wh"] = warehouse
    return frappe.db.sql(f"""
        SELECT 'Bin'         AS voucher_type,
               name          AS voucher_name,
               warehouse,
               actual_qty,
               ordered_qty,
               reserved_qty,
               projected_qty,
               valuation_rate,
               stock_uom
        FROM `tabBin`
        WHERE {' AND '.join(where)}
        ORDER BY warehouse
    """, p, as_dict=True)


def _shape_breakdown(column, item_code, warehouse, ladder_cf):
    """
    Return the breakdown payload for a given column. Each item carries:
      - voucher_type, voucher_name (hyperlink),
      - qty fields,
      - higher_uom_qty if a CF is known,
      - tooltip lines for the cell-level math.
    """
    higher_meta = {}
    try:
        from chaizup_toc.api.item_projection_compute import _pick_higher_uoms
        higher_meta = (_pick_higher_uoms([item_code]) or {}).get(item_code) or {}
    except Exception:
        higher_meta = {}
    cf = flt(higher_meta.get("conversion_factor") or 1.0) or 1.0
    h_uom = higher_meta.get("higher_uom") or higher_meta.get("stock_uom") or ""

    if column == "wo_remaining_production":
        rows = _bd_wo_production(item_code, warehouse)
        for r in rows:
            r["higher_uom"] = h_uom
            r["higher_uom_qty"] = round(flt(r.get("remaining_qty") or 0) / cf, 3)
            r["voucher_url"] = f"/app/work-order/{r['voucher_name']}"
        return {
            "title": f"WO Remaining Production — {item_code}",
            "header_cols": ["WO", "Status", "FG Warehouse", "Planned", "Produced",
                            "Remaining (Stock UOM)", f"Remaining ({h_uom})",
                            "Planned Start"],
            "rows":   rows,
        }
    if column == "will_consume_open_wo":
        rows = _bd_wo_consumption(item_code, warehouse)
        for r in rows:
            r["higher_uom"] = h_uom
            r["higher_uom_qty"] = round(flt(r.get("remaining_qty") or 0) / cf, 3)
            r["voucher_url"] = f"/app/work-order/{r['voucher_name']}"
        return {
            "title": f"Will Consume (Open WOs) — {item_code}",
            "header_cols": ["WO", "Status", "Source WH", "Produces (FG)",
                            "Required", "Transferred",
                            "Remaining (Stock UOM)", f"Remaining ({h_uom})",
                            "Planned Start"],
            "rows":   rows,
        }
    if column == "po_remaining":
        rows = _bd_po(item_code, warehouse)
        for r in rows:
            r["higher_uom"] = h_uom
            r["higher_uom_qty"] = round(flt(r.get("remaining_stock_qty") or 0) / cf, 3)
            r["voucher_url"] = f"/app/purchase-order/{r['voucher_name']}"
        return {
            "title": f"PO Remaining — {item_code}",
            "header_cols": ["PO", "Status", "Supplier", "Warehouse",
                            "Ordered Qty", "Received", "CF",
                            "Remaining (Stock UOM)", f"Remaining ({h_uom})",
                            "Schedule"],
            "rows":   rows,
        }
    if column == "will_dispatch_pending_so":
        rows = _bd_so(item_code, warehouse)
        for r in rows:
            r["higher_uom"] = h_uom
            r["higher_uom_qty"] = round(flt(r.get("remaining_stock_qty") or 0) / cf, 3)
            r["voucher_url"] = f"/app/sales-order/{r['voucher_name']}"
        return {
            "title": f"Will Dispatch (Pending SOs) — {item_code}",
            "header_cols": ["SO", "Status", "Customer", "Warehouse",
                            "Stock Qty", "Delivered", "CF",
                            "Remaining (Stock UOM)", f"Remaining ({h_uom})",
                            "Delivery Date"],
            "rows":   rows,
        }

    # Composite views — shortage / net / DoC — show the math on top of
    # the component breakdowns.
    if column in ("shortage_physical", "shortage_projected", "net_available"):
        wo_p = _bd_wo_production(item_code, warehouse)
        wo_c = _bd_wo_consumption(item_code, warehouse)
        po   = _bd_po(item_code, warehouse)
        so   = _bd_so(item_code, warehouse)
        stk  = _bd_stock(item_code, warehouse)
        for r in wo_p: r["higher_uom_qty"] = round(flt(r.get("remaining_qty") or 0) / cf, 3); r["voucher_url"] = f"/app/work-order/{r['voucher_name']}"
        for r in wo_c: r["higher_uom_qty"] = round(flt(r.get("remaining_qty") or 0) / cf, 3); r["voucher_url"] = f"/app/work-order/{r['voucher_name']}"
        for r in po:   r["higher_uom_qty"] = round(flt(r.get("remaining_stock_qty") or 0) / cf, 3); r["voucher_url"] = f"/app/purchase-order/{r['voucher_name']}"
        for r in so:   r["higher_uom_qty"] = round(flt(r.get("remaining_stock_qty") or 0) / cf, 3); r["voucher_url"] = f"/app/sales-order/{r['voucher_name']}"
        for r in stk:  r["higher_uom_qty"] = round(flt(r.get("actual_qty") or 0) / cf, 3)
        title_map = {
            "shortage_physical":  "Shortage — Physical",
            "shortage_projected": "Shortage — Projected",
            "net_available":      "Net Available",
        }
        return {
            "title":          f"{title_map[column]} — {item_code}",
            "composite":      True,
            "higher_uom":     h_uom,
            "conversion_factor": cf,
            "sections": [
                {"key": "stock",        "title": "Current Stock (per Bin)",     "rows": stk},
                {"key": "wo_production","title": "Will Receive — Work Orders (FG produced)", "rows": wo_p},
                {"key": "po",           "title": "Will Receive — Purchase Orders", "rows": po},
                {"key": "wo_consumption","title": "Will Consume — Open Work Orders (component required)", "rows": wo_c},
                {"key": "so",           "title": "Will Dispatch — Pending Sales Orders", "rows": so},
            ],
        }

    if column == "days_of_cover":
        adu_val = frappe.db.get_value(
            "Item", item_code,
            ["item_name", "custom_toc_adu_value",
             "custom_toc_adu_period_days"],
            as_dict=True) or {}
        stock = sum(flt(r.get("actual_qty") or 0) for r in _bd_stock(item_code, warehouse))
        adu   = flt(adu_val.get("custom_toc_adu_value") or 0)
        doc   = (stock / adu) if adu > 0 else None
        return {
            "title": f"Days of Cover — {item_code}",
            "scalar": True,
            "lines": [
                f"Item: {item_code}  ({adu_val.get('item_name') or ''})",
                f"Current Stock: {stock:.3f}",
                f"ADU (Item.custom_toc_adu_value): {adu:.3f}",
                f"ADU Period: {adu_val.get('custom_toc_adu_period_days') or '—'} days",
                (f"Days of Cover = Stock ÷ ADU = {doc:.2f} days"
                 if doc is not None else
                 "Days of Cover unavailable — ADU is 0 or unset."),
            ],
        }

    return {"title": column, "rows": []}


@frappe.whitelist()
def get_breakdown(column, item_code, warehouse=None):
    """Drill-down dispatcher — see `_shape_breakdown` for shape."""
    return _shape_breakdown(column, item_code, warehouse, None)


# =============================================================================
# 4. EXPORT — branded multi-sheet XLSX.
# =============================================================================
@frappe.whitelist()
def export_xlsx(filters=None):
    """
    Multi-sheet workbook: Cover / Items / Shortage Drivers / Pending WO /
    Pending PO / Pending SO / UOM Ladder. Stream as the HTTP response.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side,
    )
    from openpyxl.formatting.rule import DataBarRule
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_dashboard_data(filters)
    raw_rows = data.get("rows") or []
    # Flatten — strip group headers, keep leaves.
    rows = []
    for r in raw_rows:
        if r.get("_group_header") and r.get("_children"):
            rows.extend(r["_children"])
        elif not r.get("_group_header"):
            rows.append(r)
    cols    = [c for c in (data.get("columns") or []) if not c.get("hidden")]
    summary = data.get("summary") or []
    fopts   = get_filter_options() or {}
    pending_status = (fopts.get("pending") or {})
    filters_used = (data.get("filters_used") or {})

    wb = Workbook()
    # ── Style primitives ────────────────────────────────────────────────
    fill_header_brand = PatternFill("solid", fgColor="0F172A")
    fill_header_accent = PatternFill("solid", fgColor="1F2937")
    fill_subheader   = PatternFill("solid", fgColor="334155")
    fill_zebra       = PatternFill("solid", fgColor="F8FAFC")
    fill_shortage    = PatternFill("solid", fgColor="FEE2E2")
    fill_warn        = PatternFill("solid", fgColor="FEF3C7")
    fill_neg         = PatternFill("solid", fgColor="FCA5A5")
    fill_cover_title = PatternFill("solid", fgColor="111827")
    font_h1     = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    font_h2     = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    font_label  = Font(name="Calibri", size=10, bold=True, color="475569")
    font_value  = Font(name="Calibri", size=11, color="0F172A")
    font_th     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_normal = Font(name="Calibri", size=10, color="0F172A")
    thin = Side(border_style="thin", color="E2E8F0")
    border_all   = Border(top=thin, bottom=thin, left=thin, right=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    align_right  = Alignment(horizontal="right",  vertical="center")
    align_wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1 — COVER
    # ═══════════════════════════════════════════════════════════════════════
    cover = wb.active
    cover.title = "Cover"
    cover.sheet_view.showGridLines = False
    cover.column_dimensions["A"].width = 4
    for col, w in zip("BCDEFGHIJ", [22, 22, 22, 22, 22, 22, 22, 22, 4]):
        cover.column_dimensions[col].width = w

    cover.merge_cells("B2:I3")
    cover["B2"] = "Item Projection View"
    cover["B2"].font = font_h1
    cover["B2"].fill = fill_cover_title
    cover["B2"].alignment = align_center

    cover.merge_cells("B4:I4")
    cover["B4"] = (
        f"Snapshot generated {frappe.utils.now_datetime():%Y-%m-%d %H:%M} "
        f"by {frappe.session.user}"
    )
    cover["B4"].font = Font(name="Calibri", size=10, italic=True, color="64748B")
    cover["B4"].alignment = align_center

    cover.merge_cells("B6:I6")
    cover["B6"] = "At-a-glance KPIs"
    cover["B6"].font = font_h2
    cover["B6"].fill = fill_subheader
    cover["B6"].alignment = align_left

    kpi_cols = ["B", "D", "F", "H"]
    kpi_row = 7
    for i, s in enumerate(summary[:4]):
        col1 = kpi_cols[i]
        col2 = chr(ord(col1) + 1)
        cover.merge_cells(f"{col1}{kpi_row}:{col2}{kpi_row}")
        cell = cover[f"{col1}{kpi_row}"]
        cell.value = str(s.get("label", ""))
        cell.font = font_label
        cell.fill = fill_zebra
        cell.alignment = align_center
        cell.border = border_all
        cover.merge_cells(f"{col1}{kpi_row+1}:{col2}{kpi_row+1}")
        cv = cover[f"{col1}{kpi_row+1}"]
        cv.value = s.get("value", 0)
        cv.font = Font(
            name="Calibri", size=18, bold=True,
            color={"red":"DC2626","orange":"D97706","green":"059669",
                   "blue":"1D4ED8"}.get((s.get("indicator") or "").lower(),
                                         "0F172A"))
        cv.alignment = align_center
        cv.border = border_all
    cover.row_dimensions[kpi_row].height = 28
    cover.row_dimensions[kpi_row + 1].height = 40

    cover.merge_cells("B10:I10")
    cover["B10"] = "Pending statuses (single source of truth — TOC Settings)"
    cover["B10"].font = font_h2
    cover["B10"].fill = fill_subheader
    cover["B10"].alignment = align_left
    banner_rows = [
        ("Work Order",     ", ".join(pending_status.get("wo") or []) or "—"),
        ("Sales Order",    ", ".join(pending_status.get("so") or []) or "—"),
        ("Purchase Order", ", ".join(pending_status.get("po") or []) or "—"),
    ]
    r = 11
    for lbl, val in banner_rows:
        cover.merge_cells(f"B{r}:C{r}")
        cover[f"B{r}"] = lbl
        cover[f"B{r}"].font = font_label
        cover[f"B{r}"].alignment = align_left
        cover[f"B{r}"].fill = fill_zebra
        cover.merge_cells(f"D{r}:I{r}")
        cover[f"D{r}"] = val
        cover[f"D{r}"].font = font_value
        cover[f"D{r}"].alignment = align_wrap
        cover.row_dimensions[r].height = 22
        r += 1

    cover.merge_cells(f"B{r+1}:I{r+1}")
    cover[f"B{r+1}"] = "Filters applied"
    cover[f"B{r+1}"].font = font_h2
    cover[f"B{r+1}"].fill = fill_subheader
    cover[f"B{r+1}"].alignment = align_left
    r += 2
    filter_pairs = [
        ("Company",     filters_used.get("company") or "(all)"),
        ("Item",        ", ".join(filters_used.get("item") or []) or "(all)"),
        ("Item Group",  ", ".join(filters_used.get("item_group") or []) or "(all)"),
        ("Warehouse",   ", ".join(filters_used.get("warehouse") or []) or "(all)"),
        ("Group by Item Group",  "yes" if filters_used.get("group_by_item_group") else "no"),
        ("Only Shortage",        "yes" if filters_used.get("only_shortage") else "no"),
    ]
    for lbl, val in filter_pairs:
        cover.merge_cells(f"B{r}:C{r}")
        cover[f"B{r}"] = lbl
        cover[f"B{r}"].font = font_label
        cover[f"B{r}"].fill = fill_zebra
        cover[f"B{r}"].alignment = align_left
        cover.merge_cells(f"D{r}:I{r}")
        cover[f"D{r}"] = val
        cover[f"D{r}"].font = font_value
        cover[f"D{r}"].alignment = align_wrap
        r += 1

    r += 1
    cover.merge_cells(f"B{r}:I{r}")
    cover[f"B{r}"] = "Sheet guide"
    cover[f"B{r}"].font = font_h2
    cover[f"B{r}"].fill = fill_subheader
    cover[f"B{r}"].alignment = align_left
    r += 1
    sheet_guide = [
        ("Cover", "KPIs + filters + pending-status snapshot + sheet guide."),
        ("Items", "Full grid — every column, frozen header, auto-filter, data bars on shortage columns."),
        ("Shortage Drivers", "Items where Shortage — Projected > 0. Sorted desc."),
        ("Pending WO Production", "Per-WO list of pending production qty (this item is FG)."),
        ("Pending WO Consumption", "Per-WO list of pending consumption (this item is component)."),
        ("Pending PO", "Per-PO list of pending receipts."),
        ("Pending SO", "Per-SO list of pending dispatches."),
        ("UOM Ladder", "Item × alt-UOM × CF table."),
    ]
    for lbl, val in sheet_guide:
        cover.merge_cells(f"B{r}:C{r}")
        cover[f"B{r}"] = lbl
        cover[f"B{r}"].font = font_label
        cover[f"B{r}"].fill = fill_zebra
        cover[f"B{r}"].alignment = align_left
        cover.merge_cells(f"D{r}:I{r}")
        cover[f"D{r}"] = val
        cover[f"D{r}"].font = font_value
        cover[f"D{r}"].alignment = align_wrap
        cover.row_dimensions[r].height = 32
        r += 1

    # ── helper: standard data sheet writer ──────────────────────────────
    def _write_data_sheet(ws, rows_in, title_text):
        ws.sheet_view.showGridLines = False
        if cols:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(row=1, column=1, value=title_text)
        ws.cell(row=1, column=1).font = font_h1
        ws.cell(row=1, column=1).fill = fill_header_brand
        ws.cell(row=1, column=1).alignment = align_center
        ws.row_dimensions[1].height = 36
        for idx, c in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=idx,
                           value=c.get("label") or c.get("fieldname"))
            cell.font = font_th
            cell.fill = fill_header_accent
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            ftype = c.get("fieldtype")
            width = 14
            if ftype in ("Float", "Currency"): width = 18
            if ftype == "Int":                 width = 12
            if c.get("fieldname") == "item_name": width = 36
            if c.get("fieldname") in ("item_code", "warehouse", "item_group"): width = 22
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.row_dimensions[2].height = 38
        numeric_fmt = '#,##0.000'
        for ridx, row in enumerate(rows_in, start=3):
            zebra = (ridx % 2 == 0)
            try:
                flags = json.loads(row.get("_flags") or "{}")
            except Exception:
                flags = {}
            is_neg_stock = bool(flags.get("negative_stock"))
            is_short_phy = bool(flags.get("shortage_physical"))
            is_short_proj = bool(flags.get("shortage_projected"))
            for cidx, c in enumerate(cols, start=1):
                v = row.get(c.get("fieldname"))
                cell = ws.cell(row=ridx, column=cidx, value=v)
                cell.font = font_normal
                cell.border = border_all
                ftype = c.get("fieldtype")
                if ftype in ("Float", "Currency"):
                    cell.alignment = align_right
                    cell.number_format = numeric_fmt
                elif ftype == "Int":
                    cell.alignment = align_right
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = align_left
                if is_neg_stock:
                    cell.fill = fill_neg
                elif is_short_phy and c.get("fieldname") in (
                        "shortage_physical", "current_stock_higher_uom"):
                    cell.fill = fill_shortage
                elif is_short_proj and c.get("fieldname") in (
                        "shortage_projected", "net_available"):
                    cell.fill = fill_warn
                elif zebra:
                    cell.fill = fill_zebra
        ws.freeze_panes = "B3"
        if cols:
            last_col = get_column_letter(len(cols))
            last_row = max(2, len(rows_in) + 2)
            ws.auto_filter.ref = f"A2:{last_col}{last_row}"
        target_fields = {
            "shortage_projected":  "FCA5A5",
            "shortage_physical":   "F87171",
            "net_available":       "BFDBFE",
        }
        last_row = max(3, len(rows_in) + 2)
        for cidx, c in enumerate(cols, start=1):
            colour = target_fields.get(c.get("fieldname"))
            if not colour:
                continue
            col_letter = get_column_letter(cidx)
            rng = f"{col_letter}3:{col_letter}{last_row}"
            try:
                ws.conditional_formatting.add(rng, DataBarRule(
                    start_type="min", end_type="max", color=colour, showValue=True,
                ))
            except Exception:
                pass

    ws_items = wb.create_sheet("Items")
    _write_data_sheet(ws_items, rows, "Items — Full Grid")

    short_rows = [r for r in rows if (r.get("shortage_projected") or 0) > 0]
    short_rows.sort(key=lambda x: -(x.get("shortage_projected") or 0))
    ws_short = wb.create_sheet("Shortage Drivers")
    _write_data_sheet(ws_short, short_rows, "Shortage Drivers (Projected > 0)")

    # ── Pending vouchers per row (one sheet each) ───────────────────────
    def _voucher_sheet(name, title, header, fetcher, qty_field):
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(header))
        ws.cell(row=1, column=1, value=title)
        ws.cell(row=1, column=1).font = font_h1
        ws.cell(row=1, column=1).fill = fill_header_brand
        ws.cell(row=1, column=1).alignment = align_center
        ws.row_dimensions[1].height = 32
        for i, h in enumerate(header, start=1):
            cell = ws.cell(row=2, column=i, value=h)
            cell.font = font_th
            cell.fill = fill_header_accent
            cell.alignment = align_center
            cell.border = border_all
        ws.row_dimensions[2].height = 28
        for col_letter, w in zip("ABCDEFGHIJKLMN", [22] * len(header)):
            ws.column_dimensions[col_letter].width = w
        row_idx = 3
        for src in rows:
            ic = src.get("item_code")
            for v in fetcher(ic, None) or []:
                values = [ic, src.get("item_name"), v.get("voucher_name"),
                          v.get("status"),
                          v.get("warehouse") or "",
                          v.get(qty_field) or 0]
                # tack on extra context fields if available
                if "supplier" in v:        values.append(v.get("supplier"))
                if "customer" in v:        values.append(v.get("customer"))
                if "schedule_date" in v:   values.append(v.get("schedule_date"))
                if "delivery_date" in v:   values.append(v.get("delivery_date"))
                if "planned_start_date" in v: values.append(v.get("planned_start_date"))
                for ci, val in enumerate(values[:len(header)], start=1):
                    cell = ws.cell(row=row_idx, column=ci, value=val)
                    cell.font = font_normal
                    cell.border = border_all
                    cell.alignment = (align_right
                                      if isinstance(val, (int, float))
                                      else align_left)
                    if isinstance(val, (int, float)):
                        cell.number_format = '#,##0.000'
                    if row_idx % 2 == 0:
                        cell.fill = fill_zebra
                row_idx += 1
        ws.freeze_panes = "A3"

    _voucher_sheet(
        "Pending WO Production",
        "Pending Work Orders (this item is the FG)",
        ["Item Code", "Item Name", "WO", "Status", "FG Warehouse",
         "Remaining Qty", "Planned Start"],
        _bd_wo_production, "remaining_qty",
    )
    _voucher_sheet(
        "Pending WO Consumption",
        "Pending Work Orders (this item is a component)",
        ["Item Code", "Item Name", "WO", "Status", "Source WH",
         "Remaining Qty", "Planned Start"],
        _bd_wo_consumption, "remaining_qty",
    )
    _voucher_sheet(
        "Pending PO",
        "Pending Purchase Orders",
        ["Item Code", "Item Name", "PO", "Status", "Warehouse",
         "Remaining Qty (Stock UOM)", "Supplier", "Schedule"],
        _bd_po, "remaining_stock_qty",
    )
    _voucher_sheet(
        "Pending SO",
        "Pending Sales Orders",
        ["Item Code", "Item Name", "SO", "Status", "Warehouse",
         "Remaining Qty (Stock UOM)", "Customer", "Delivery"],
        _bd_so, "remaining_stock_qty",
    )

    # ── UOM Ladder ──────────────────────────────────────────────────────
    ws_uom = wb.create_sheet("UOM Ladder")
    ws_uom.sheet_view.showGridLines = False
    headers = ["Item Code", "Item Name", "Stock UOM", "Alt UOM",
               "Conversion Factor", "Current Stock (Stock UOM)",
               "Current Stock (Alt UOM)"]
    ws_uom.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws_uom.cell(row=1, column=1, value="Item × UOM Ladder")
    ws_uom.cell(row=1, column=1).font = font_h1
    ws_uom.cell(row=1, column=1).fill = fill_header_brand
    ws_uom.cell(row=1, column=1).alignment = align_center
    ws_uom.row_dimensions[1].height = 32
    for i, h in enumerate(headers, start=1):
        cell = ws_uom.cell(row=2, column=i, value=h)
        cell.font = font_th
        cell.fill = fill_header_accent
        cell.alignment = align_center
        cell.border = border_all
    for col_letter, w in zip("ABCDEFG", [20, 36, 14, 14, 18, 22, 22]):
        ws_uom.column_dimensions[col_letter].width = w
    ws_uom.row_dimensions[2].height = 28
    r = 3
    for row in rows:
        try:
            ladder = json.loads(row.get("_ladder") or "[]")
        except Exception:
            ladder = []
        stock_qty = row.get("current_stock_stock_uom") or 0
        for c in ladder:
            ws_uom.cell(row=r, column=1, value=row.get("item_code"))
            ws_uom.cell(row=r, column=2, value=row.get("item_name"))
            ws_uom.cell(row=r, column=3, value=row.get("stock_uom"))
            ws_uom.cell(row=r, column=4, value=c.get("uom"))
            ws_uom.cell(row=r, column=5, value=c.get("factor"))
            ws_uom.cell(row=r, column=6, value=stock_qty)
            cf = c.get("factor") or 1
            ws_uom.cell(row=r, column=7,
                        value=(stock_qty / cf) if cf else 0)
            for col in range(1, len(headers) + 1):
                cell = ws_uom.cell(row=r, column=col)
                cell.font = font_normal
                cell.border = border_all
                if col in (5, 6, 7):
                    cell.alignment = align_right
                    cell.number_format = '#,##0.0000'
                else:
                    cell.alignment = align_left
                if r % 2 == 0:
                    cell.fill = fill_zebra
            r += 1
    ws_uom.freeze_panes = "A3"
    ws_uom.auto_filter.ref = f"A2:G{max(2, r - 1)}"

    buf = BytesIO()
    wb.save(buf)
    filename = f"Item-Projection-View-{frappe.utils.now()[:10]}.xlsx"
    frappe.local.response.update({
        "type":        "binary",
        "filename":    filename,
        "filecontent": buf.getvalue(),
    })
