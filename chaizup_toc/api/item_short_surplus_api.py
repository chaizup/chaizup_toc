# =============================================================================
# CONTEXT: Item Short / Surplus report — backend API.
#   Custom Page at /app/item-short-surplus is a thin Tabulator + filter-bar
#   client; this module owns ALL data + math.
#
#   The report answers: "For each item, given the user-chosen warehouses
#   and pending status/workflow lists, is supply (current stock + pending
#   WOs + pending POs) enough to cover demand (pending SOs + remaining
#   WO consumption)?"
#
# MEMORY:  app_chaizup_toc.md § v0.0.22 — Item Short / Surplus report
# DOC:     ../../chaizup_toc/page/item_short_surplus/item_short_surplus.md
# RELATED: item_projection_compute.py (sibling report, similar UOM + xlsx
#          pattern), production_plan_engine.py (status/workflow parsers).
#
# 5 ENDPOINTS:
#   - get_filter_options()           → default warehouses + pending lists from TOC Settings
#                                       + all candidate SO/WO/PO statuses + workflow states
#                                       + Item Group / Company link options for chips
#   - get_report(filters)            → main rows + computed columns. The
#                                       LIVE filter accuracy principle holds:
#                                       no stored mirrors; every aggregate is
#                                       computed from the source tables at
#                                       query time.
#   - get_voucher_drilldown(item_code, source, filters) → modal payload
#                                       (per-voucher rows with planned vs
#                                       actual qty in both UOMs).
#   - export_xlsx(filters)           → multi-sheet branded XLSX. 4 sheets:
#                                       (1) main report, (2) filter snapshot
#                                       + run metadata, (3) shortage-only
#                                       sorted desc, (4) surplus-only sorted desc.
#   - _all_known_statuses()          → helper, used by get_filter_options to
#                                       populate the multi-select dropdowns
#                                       with every status/workflow_state that
#                                       has ever appeared in the relevant table.
#
# COMPUTATIONAL CONTRACT (the math, in plain English):
#   - Higher UOM picker: same as Item Projection View — largest non-stock
#     UOM in the item's UOM Conversion Detail (CF > 1). Falls back to
#     stock UOM (CF = 1) when no alternate UOM exists.
#   - to_higher(qty_in_stock) = qty_in_stock / cf
#   - current_stock_stock_uom = Σ Bin.actual_qty
#       scoped by (item, warehouses, company) — if warehouses empty,
#       sum across all warehouses; if company set, narrow to its WHs.
#   - pending_so_stock_uom = Σ (so_item.qty − so_item.delivered_qty)
#       scoped by (item, warehouses if so_item.warehouse in, company)
#       AND ((so.docstatus=1 AND so.status IN <user statuses>)
#            OR (so.docstatus=0 AND so.workflow_state IN <user wf states>)).
#       Standard ERPNext rule: cancelled (docstatus=2) AND status "Closed"
#       are always excluded.
#   - pending_wo_stock_uom = Σ (wo.qty − wo.produced_qty)
#       scoped by (production_item, fg_warehouse in <warehouses>, company)
#       AND ((wo.docstatus=1 AND wo.status IN <user statuses>)
#            OR (wo.docstatus=0 AND wo.workflow_state IN <user wf states>)).
#       "Closed", "Stopped", "Cancelled" excluded if not in user list.
#   - pending_po_stock_uom = Σ (po_item.qty − po_item.received_qty)
#       scoped by (po_item.item_code, po_item.warehouse in <warehouses>, company)
#       AND ((po.docstatus=1 AND po.status IN <user statuses>)
#            OR (po.docstatus=0 AND po.workflow_state IN <user wf states>)).
#   - remain_wo_consume_stock_uom = Σ (wo_item.required_qty − wo_item.transferred_qty)
#       scoped to: parent WO satisfies the user's pending WO filter
#       AND wo_item.item_code = X (the component being consumed)
#       AND wo.fg_warehouse in <warehouses> (component-level scoping is the
#         parent WO's FG warehouse — the WO's consumption pulls from its
#         source/wip warehouse but the demand it represents is for the FG
#         warehouse that owns the WO).
#   - total_demand = pending_so + remain_wo_consume
#   - supply = current_stock + pending_wo + pending_received_po
#   - net = supply − demand
#     → if net >= 0: status="Surplus", surplus=net, shortfall=0
#     → else:       status="Shortage", shortfall=|net|, surplus=0
#
# FILTER ACCURACY (locked, see Claude memory feedback_filter_accuracy_principle):
#   Every filter resolves LIVE at query time. The Item Group filter, for
#   example, runs `tabItem.item_group = X` at SELECT time — never reads
#   a fetched mirror. Item Group reclassifications propagate immediately.
#
# RESTRICTED:
#   - DO NOT cache per-item or per-warehouse aggregates between requests.
#     The user expects "current state at this moment". A 5-minute cache
#     would silently lie about Bin levels right after a Stock Entry.
#   - DO NOT use `frappe.db.count` for the aggregates — it returns row
#     counts, not summed quantities. Use Σ via raw SQL.
#   - DO NOT exclude items by item.disabled by default. The filter bar's
#     "Item" chip is the canonical scope; disabled items can still have
#     real stock + open vouchers that operators need to surface.
#   - The "active wo / active po / active so" boolean filters MUST be
#     post-aggregation (rows where the respective qty > 0). Pre-filtering
#     by EXISTS could miss items that should still appear with zero in
#     other columns. Same for "no so / no po / no wo" (qty == 0).
#   - DO NOT split SO/WO/PO eligibility into two queries (submitted vs draft).
#     The (status / workflow_state) OR-tuple is one logical predicate —
#     splitting risks double-counting if ERPNext ever stores both fields.
#   - The xlsx insights MUST stay at 4 sheets (per user spec). Don't add
#     a "summary chip" sheet — user explicitly said no extra chips.
# =============================================================================

from __future__ import annotations
from collections import defaultdict
from typing import Any

import frappe
from frappe import _
from frappe.utils import flt, cint, getdate, now_datetime, get_datetime


def _fmt_local_ts(dt) -> str:
    """v0.0.37 — Friendly local timestamp for every operator-visible
    surface (report footer payload, XLSX audit sheet, etc.).
    Output: "dd-MMM-yyyy hh:mm am/pm" — e.g. "28-May-2026 02:31 AM".

    DANGER: Do NOT use this for filenames or DB writes — those rely on
            the strict ISO-style %Y%m%d_%H%M%S layout. This helper is
            display-only.
    """
    try:
        if dt is None or dt == "":
            return ""
        # Coerce strings -> datetime (preserves time, unlike getdate
        # which strips it to a bare date).
        if isinstance(dt, str):
            dt = get_datetime(dt)
        return dt.strftime("%d-%b-%Y %I:%M %p") if hasattr(dt, "strftime") else str(dt)
    except Exception:
        return str(dt or "")

# Reuse status/workflow parsers + WO/PO eligibility helpers from the engine
# so the report's semantics match every other TOC consumer.
from chaizup_toc.chaizup_toc.toc_engine.production_plan_engine import (
    _parse_statuses,
    _parse_wo_statuses,
    _parse_wo_workflow_states,
    _parse_po_statuses,
    _parse_po_workflow_states,
    _wo_has_workflow_column,
    _po_has_workflow_column,
)


# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------

def _as_list(v: Any) -> list[str]:
    """Normalize a filter value to a list of non-empty strings.

    Filters can arrive from the JS as:
      - JSON-encoded list (Frappe converts dict -> dict, but lists in
        URL-style come as JSON strings)
      - native list (from frappe.call with `args` param)
      - comma-separated string (defensive)
      - single value (defensive)
    """
    if v is None:
        return []
    if isinstance(v, str):
        v = v.strip()
        if not v:
            return []
        if v.startswith("["):
            try:
                v = frappe.parse_json(v)
            except Exception:
                return [v]
        else:
            v = [x.strip() for x in v.split(",")]
    if isinstance(v, (list, tuple)):
        return [str(x).strip() for x in v if str(x).strip()]
    return [str(v).strip()]


def _wo_has_wf() -> bool:
    """Cached check for workflow_state column on tabWork Order."""
    try:
        return _wo_has_workflow_column()
    except Exception:
        return False


def _po_has_wf() -> bool:
    try:
        return _po_has_workflow_column()
    except Exception:
        return False


def _so_has_wf() -> bool:
    """Sales Order workflow_state column check — sibling of the WO/PO
    helpers, not yet abstracted into production_plan_engine because no
    automation reads SO workflow_state directly (only the projection
    engine does, via projection_confirmed_so_workflow_states)."""
    try:
        return bool(frappe.db.sql("""
            SELECT 1 FROM information_schema.COLUMNS
             WHERE TABLE_NAME = 'tabSales Order' AND COLUMN_NAME = 'workflow_state'
            LIMIT 1
        """))
    except Exception:
        return False


# -----------------------------------------------------------------------------
# Higher UOM resolver — single batched call (matches Item Projection View
# convention so two reports never disagree on the higher UOM for an item).
# -----------------------------------------------------------------------------

def _pick_higher_uoms(item_codes: list[str]) -> dict[str, dict]:
    """For each item, pick the largest non-stock UOM (CF > 1) from its
    UOM Conversion Detail. Returns {item_code: {higher_uom, cf, stock_uom}}.

    Falls back to (stock_uom, 1.0) when item has no alternate UOM.

    Idempotent + side-effect free — pure read of Item + UOM Conversion Detail.
    """
    if not item_codes:
        return {}
    rows = frappe.db.sql("""
        SELECT i.name AS item, i.stock_uom,
               ucd.uom, IFNULL(ucd.conversion_factor, 1) AS cf
          FROM `tabItem` i
          LEFT JOIN `tabUOM Conversion Detail` ucd
                 ON ucd.parent = i.name AND ucd.parenttype = 'Item'
                AND ucd.uom != i.stock_uom
                AND IFNULL(ucd.conversion_factor, 0) > 1
         WHERE i.name IN %(items)s
         ORDER BY i.name, ucd.conversion_factor DESC
    """, {"items": tuple(item_codes)}, as_dict=True)

    out: dict[str, dict] = {}
    seen: set[str] = set()
    for r in rows:
        if r.item in seen:
            continue
        seen.add(r.item)
        if r.uom:
            out[r.item] = {
                "higher_uom": r.uom,
                "cf": flt(r.cf) or 1.0,
                "stock_uom": r.stock_uom,
            }
        else:
            out[r.item] = {
                "higher_uom": r.stock_uom,
                "cf": 1.0,
                "stock_uom": r.stock_uom,
            }
    # Items missing from rows (no Item row at all — defensive)
    for code in item_codes:
        if code not in out:
            out[code] = {"higher_uom": "", "cf": 1.0, "stock_uom": ""}
    return out


def _to_higher(qty_stock: float, cf: float) -> float:
    """Stock UOM → higher UOM. cf > 0 enforced by _pick_higher_uoms."""
    return flt(qty_stock) / flt(cf) if cf else 0.0


# -----------------------------------------------------------------------------
# Endpoint 1 — get_filter_options
# -----------------------------------------------------------------------------

@frappe.whitelist()
def get_filter_options() -> dict:
    """Public wrapper that returns selectable options for the filter bar.

    v0.0.36 PERF — cached for 60 s per (site, user) — the underlying
    META queries (statuses, workflow states, companies, warehouses)
    NEVER change within a single user's session of any reasonable
    length. Cache key includes the user so per-user permission scopes
    on companies / warehouses don't leak across users.
    """
    key = f"iss:filter_options:{frappe.session.user}"
    cached = frappe.cache.get_value(key)
    if cached:
        return cached
    payload = _get_filter_options_uncached()
    # 60 s TTL — long enough to absorb a normal browsing burst, short
    # enough that operator-side TOC Settings edits surface within a
    # minute on the next page load.
    frappe.cache.set_value(key, payload, expires_in_sec=60)
    return payload


def _get_filter_options_uncached() -> dict:
    """Return selectable options for the filter bar.

    v0.0.26 — Status + Workflow State MERGED into a single combined
    option per voucher type. Each combined option is shipped as a dict:
        {"key": "<status>|<workflow_state>",
         "label": "<status> : <workflow_state>",
         "status": "...",
         "workflow_state": "..."}

    Why combined: operators think of "pending" as a single (status, workflow)
    tuple, not two independent multi-selects. v0.0.25's two-chip layout
    forced them to mentally cross-product the picks. v0.0.26 surfaces
    every (status, workflow_state) pair that actually exists in the data
    so the user picks the real combinations directly.

    Pairs come from a DISTINCT live scan of (status, workflow_state) on
    each voucher table — only pairs that actually exist are listed, no
    cartesian explosion. Always-excluded statuses (Closed, Cancelled) are
    omitted from the list entirely (no need to render disabled rows since
    the operator can't act on them anyway).

    The `|` delimiter in the key is safe because Frappe Status / Workflow
    State values never contain `|`.

    RESTRICT (v0.0.26):
      - Do NOT re-split into two chips. Operator mental model is a single
        pair per voucher; respect it.
      - Do NOT include Closed / Cancelled in the surfaced pairs. They're
        always excluded by the always_excluded list applied in get_report.
      - The `|` delimiter must NEVER appear in Frappe Status / Workflow
        State values. If a future ERPNext version allows it, switch to a
        different sentinel and update the splitter in get_report.
    """
    BLOCK = {"Closed", "Cancelled", "Stopped"}

    def _workflow_states_for(doctype: str) -> list[dict]:
        """Return all states DEFINED in the active Workflow attached to
        `doctype`, with their `doc_status` mapping (0=Draft, 1=Submitted,
        2=Cancelled). Without this, states like 'WO Approved' (defined
        but not yet observed) would be missing from the dropdown.

        Each entry: {"state": str, "doc_status": "0"|"1"|"2"}.
        """
        try:
            rows = frappe.db.sql("""
                SELECT DISTINCT wds.state, IFNULL(wds.doc_status, '0') AS doc_status
                  FROM `tabWorkflow Document State` wds
                  JOIN `tabWorkflow` w ON w.name = wds.parent
                 WHERE w.is_active = 1
                   AND w.document_type = %s
                   AND wds.state IS NOT NULL AND wds.state != ''
                 ORDER BY wds.doc_status, wds.state
            """, (doctype,), as_dict=True)
            return [{"state": r["state"], "doc_status": str(r["doc_status"])} for r in rows]
        except Exception:
            return []

    def _status_options_for(doctype: str) -> list[str]:
        """Return all defined `status` Select options for a doctype.
        Falls back to DISTINCT scan if meta lookup fails."""
        try:
            meta = frappe.get_meta(doctype)
            field = meta.get_field("status")
            if field and field.options:
                opts = [s.strip() for s in str(field.options).split("\n") if s.strip()]
                if opts:
                    return opts
        except Exception:
            pass
        # Fallback: DISTINCT scan of the table
        return _distinct(doctype, "status")

    # Frappe's status field on submittable doctypes is computed from
    # docstatus. The status value at docstatus=0 is always "Draft"; at
    # docstatus=2 always "Cancelled". Everything else is a docstatus=1
    # status. This mapping lets us pair-match workflow states (which
    # have their own doc_status) to the right status side.
    _DRAFT_STATUS    = "Draft"
    _CANCEL_STATUS   = "Cancelled"

    def _distinct_pairs(doctype: str, has_wf: bool) -> list[dict]:
        """Return (status, workflow_state) pairs for the dropdown.

        v0.0.29 — VALID-COMBINATIONS-ONLY. Both sides of every pair are
        filled (no `—` placeholders). Combinations are filtered by
        docstatus compatibility:

          - workflow_state.doc_status=0 (Draft workflow state)
              → pairs ONLY with status="Draft"  (the docstatus=0 status)
          - workflow_state.doc_status=1 (Submitted workflow state)
              → pairs with every docstatus=1 status (not Draft / Cancelled)
          - workflow_state.doc_status=2 (Cancelled workflow state)
              → always excluded (Closed/Cancelled is always blocked)

        Source — META ONLY:
          - statuses: `frappe.get_meta(doctype).get_field("status").options`
          - workflow states: `tabWorkflow Document State` + doc_status

        Why valid-only (per user spec 2026-05-27): the cartesian explosion
        produced unrealistic combos like "Completed : WO Rejected" (a
        completed WO can't be in a Draft workflow state). Docstatus
        compatibility eliminates them at source.
        """
        pairs_seen: set[str] = set()
        out: list[dict] = []

        def _add(st: str, wf: str):
            st, wf = (st or "").strip(), (wf or "").strip()
            if not st or not wf:
                # v0.0.29 — both sides MUST be filled. No "— : X" or
                # "X : —" entries. Operator must pick a complete pair.
                return
            if st in BLOCK:
                return
            key = f"{st}|{wf}"
            if key in pairs_seen:
                return
            pairs_seen.add(key)
            out.append({
                "key": key,
                "label": f"{st} : {wf}",
                "status": st,
                "workflow_state": wf,
            })

        statuses_defined = _status_options_for(doctype)
        wf_defined = _workflow_states_for(doctype) if has_wf else []

        if wf_defined:
            # Bucket workflow states by their doc_status
            wf_draft     = [w["state"] for w in wf_defined if w["doc_status"] == "0"]
            wf_submitted = [w["state"] for w in wf_defined if w["doc_status"] == "1"]
            # doc_status=2 workflow states are intentionally ignored
            # (those map to Cancelled, which is in BLOCK)

            # Submitted-side statuses = every defined status EXCEPT Draft + Cancelled
            submitted_statuses = [s for s in statuses_defined
                                  if s not in (_DRAFT_STATUS, _CANCEL_STATUS) and s not in BLOCK]

            # Draft workflow states pair only with Draft status
            for wf in wf_draft:
                _add(_DRAFT_STATUS, wf)

            # Submitted workflow states pair with every submitted-side status
            for wf in wf_submitted:
                for st in submitted_statuses:
                    _add(st, wf)
        else:
            # No workflow attached — no pair is possible (both sides required).
            # Operator can't filter by status alone in the pair model. Return
            # status-only fallback for back-compat (status as both halves).
            for st in statuses_defined:
                if st in (_DRAFT_STATUS, _CANCEL_STATUS) or st in BLOCK:
                    continue
                # Treat as status-only by writing it on both sides — the
                # parser splits and we get status="X", workflow_state="X".
                # In a no-workflow site, workflow_state is never queried, so
                # this is harmless. Better than no options at all.
                pairs_seen.add(f"{st}|")
                out.append({
                    "key": f"{st}|",
                    "label": st,
                    "status": st,
                    "workflow_state": "",
                })

        # Alphabetical for predictable dropdown order
        out.sort(key=lambda o: o["label"])
        return out

    so_pairs = _distinct_pairs("Sales Order",    _so_has_wf())
    wo_pairs = _distinct_pairs("Work Order",     _wo_has_wf())
    po_pairs = _distinct_pairs("Purchase Order", _po_has_wf())

    companies = [r["name"] for r in frappe.db.get_list(
        "Company", fields=["name"], order_by="name asc", limit=0)]

    # v0.0.30 — Pre-populate defaults from TOC Settings.
    #
    # The 3 pending pair fields (projection_pending_so_statuses,
    # pending_wo_statuses, pending_po_statuses) on TOC Settings store
    # `<status>|<workflow_state>` keys, one per line. We surface them as
    # `defaults.{so,wo,po}_pairs` so the report's JS controller can seed
    # the chip widgets on FIRST visit (no localStorage yet).
    #
    # Precedence on the JS side:
    #   1. localStorage (user's previous session picks) — wins if present
    #   2. server-side defaults (this payload) — first-visit fallback
    #   3. empty list — only when both above are absent
    #
    # Pre-populated keys are also FILTERED to only valid pairs (those
    # present in the dropdown options list). Stale TOC Settings entries
    # (e.g., a workflow state that no longer exists in the data) are
    # silently dropped from defaults so the chip never references an
    # option the user can't see / re-pick later.
    so_default = _read_pair_field_from_toc_settings(
        "projection_pending_so_statuses", so_pairs)
    wo_default = _read_pair_field_from_toc_settings(
        "pending_wo_statuses", wo_pairs)
    po_default = _read_pair_field_from_toc_settings(
        "pending_po_statuses", po_pairs)

    # v0.0.35 — Default Warehouses: every warehouse classified as
    # "Inventory" in TOC Settings → Warehouse Rules. Mirrors the report's
    # on-hand-stock semantics (WIP / Excluded are NEVER seeded; operator
    # can still add them manually per session).
    # DANGER: frappe.get_all("TOC Warehouse Rule", parent="TOC Settings")
    #          does NOT match child rows reliably for Singles. Always read
    #          the parent doc and walk its child list — child rows live
    #          on the in-memory doc.
    wh_default: list[str] = []
    try:
        toc_doc = frappe.get_single("TOC Settings")
        for r in (toc_doc.get("warehouse_rules") or []):
            if (r.warehouse_purpose or "").strip() == "Inventory" and r.warehouse:
                wh_default.append(r.warehouse)
    except Exception:
        # Defensive: if the child rows can't be read for any reason,
        # fall back to empty (UI keeps working, no auto-seed).
        wh_default = []

    # v0.0.35 — Default Company: Global Defaults → default_company.
    # Falls back to the only Company on file if exactly one exists.
    try:
        co_default = frappe.db.get_single_value(
            "Global Defaults", "default_company") or ""
    except Exception:
        co_default = ""
    if not co_default and len(companies) == 1:
        co_default = companies[0]

    return {
        "options": {
            # Combined pair lists — replace v0.0.25's split lists.
            "so_pairs": so_pairs,
            "wo_pairs": wo_pairs,
            "po_pairs": po_pairs,
            "companies": companies,
        },
        # v0.0.30 — defaults seeded from TOC Settings (operator override
        # per session via the chip widget).
        # v0.0.35 — adds warehouses + company defaults.
        "defaults": {
            "so_pairs": so_default,
            "wo_pairs": wo_default,
            "po_pairs": po_default,
            "warehouses": wh_default,
            "company":    co_default,
        },
        # Reminder for the UI footer / tooltip.
        "always_excluded": {
            "sales_order":    ["Closed", "Cancelled"],
            "work_order":     ["Closed", "Cancelled", "Stopped"],
            "purchase_order": ["Closed", "Cancelled"],
        },
    }


def _read_pair_field_from_toc_settings(field: str, valid_options: list) -> list[str]:
    """v0.0.30 — Read a pair-formatted Small Text field from TOC Settings
    Singles and return the list of pair KEYS that match an option in
    `valid_options`.

    Stale keys (those not in valid_options) are silently dropped — a chip
    referencing an option the user can't see / re-pick is worse than no chip.

    `valid_options` is the same list returned in `options.<X>_pairs` so
    the JS sees a strict subset.
    """
    try:
        raw = frappe.db.get_single_value("TOC Settings", field) or ""
    except Exception:
        return []
    if not raw:
        return []
    valid_keys = {opt["key"] for opt in valid_options}
    out: list[str] = []
    seen: set[str] = set()
    for line in raw.split("\n"):
        line = line.strip()
        if not line:
            continue
        # Defensive — accept legacy plain-status lines too. Convert to
        # `status|` so it matches the valid_options.key format.
        key = line if "|" in line else f"{line}|"
        if key in seen:
            continue
        seen.add(key)
        if key in valid_keys:
            out.append(key)
    return out


def _distinct(doctype: str, field: str) -> list[str]:
    """Return non-empty distinct values of <field> on <doctype>, sorted."""
    try:
        rows = frappe.db.sql(f"""
            SELECT DISTINCT `{field}` AS v
              FROM `tab{doctype}`
             WHERE `{field}` IS NOT NULL AND `{field}` != ''
             ORDER BY v
        """, as_dict=True)
        return [r["v"] for r in rows]
    except Exception:
        return []


# -----------------------------------------------------------------------------
# Endpoint 2 — get_report (the main payload)
# -----------------------------------------------------------------------------

@frappe.whitelist()
def get_report(filters: Any = None) -> dict:
    """Compute the Item Short / Surplus rows.

    Filter shape (all optional, all live-resolved):
        item:            list[str] | str    — item code(s)
        item_group:      list[str] | str    — item group(s); LIVE filter (resolves to current items)
        warehouses:      list[str]          — scope warehouses; empty = all
        company:         str                — scope to this company's WHs (when warehouses empty)
        so_statuses:     list[str]
        so_workflow_states: list[str]
        wo_statuses:     list[str]
        wo_workflow_states: list[str]
        po_statuses:     list[str]
        po_workflow_states: list[str]
        active_wo:         bool  — only items with pending WO qty > 0
        active_po:         bool  — only items with pending PO qty > 0
        active_so:         bool  — only items with pending SO qty > 0
        active_wo_consume: bool  — only items with pending WO consumption qty > 0
        no_so:             bool  — only items with 0 pending SO qty
        no_po:             bool  — only items with 0 pending PO qty
        no_wo:             bool  — only items with 0 pending WO qty
        no_wo_consume:     bool  — only items with 0 pending WO consumption qty
        so_no_wo:          bool  — pending SO > 0 AND pending WO == 0
                                   (sales demand with no production planned)
        wo_with_shortage:  bool  — pending WO > 0 AND demand_status == Shortage
                                   (WO already started but supply insufficient)
        shortage_only:     bool  — only Shortage items (demand_status == "Shortage")
        surplus_only:      bool  — only Surplus items (demand_status == "Surplus")
    """
    f = frappe.parse_json(filters) if isinstance(filters, str) else (filters or {})

    item_codes_filter = _as_list(f.get("item"))
    item_groups       = _as_list(f.get("item_group"))
    warehouses        = _as_list(f.get("warehouses"))
    company           = (f.get("company") or "").strip()

    # v0.0.26 — combined pair filter. Front-end ships an array of
    # "<status>|<workflow_state>" strings; we split into two sorted
    # uniqued lists (status[], workflow_state[]) for the existing
    # eligibility-SQL helpers. Empty halves are dropped — only non-blank
    # values reach the IN (...) predicate.
    #
    # Back-compat (v0.0.25): if a caller still ships so_statuses /
    # so_workflow_states arrays, fall back to those.
    def _split_pairs(pairs_key: str, st_key: str, wf_key: str):
        pairs = _as_list(f.get(pairs_key))
        if pairs:
            st: list[str] = []
            wf: list[str] = []
            for p in pairs:
                # Format: "status|workflow_state". Either side may be empty.
                if "|" not in p:
                    # Plain status-only key (old format) — treat as status
                    st.append(p)
                    continue
                s_part, _, w_part = p.partition("|")
                if s_part: st.append(s_part)
                if w_part: wf.append(w_part)
            # Dedup, preserve order.
            return list(dict.fromkeys(st)), list(dict.fromkeys(wf))
        # Legacy split-list path
        return _as_list(f.get(st_key)), _as_list(f.get(wf_key))

    so_st, so_wf = _split_pairs("so_pairs", "so_statuses", "so_workflow_states")
    wo_st, wo_wf = _split_pairs("wo_pairs", "wo_statuses", "wo_workflow_states")
    po_st, po_wf = _split_pairs("po_pairs", "po_statuses", "po_workflow_states")

    # ALWAYS exclude "Closed" / "Cancelled" / "Stopped" defensively, even
    # if a user puts them in the multi-select.
    BLOCK = {"Closed", "Cancelled", "Stopped"}
    so_st = [x for x in so_st if x not in BLOCK]
    wo_st = [x for x in wo_st if x not in BLOCK]
    po_st = [x for x in po_st if x not in BLOCK]

    # ── Step 1: Resolve the universe of items the report should cover.
    items = _resolve_items(item_codes_filter, item_groups)
    if not items:
        return {"rows": [], "columns": _columns(), "filters_used": f}

    # ── Step 2: Higher UOM map (single query, cached during this request).
    uom_map = _pick_higher_uoms(items)

    # ── Step 3: Bulk aggregates — five maps keyed by item_code.
    stock_map      = _agg_current_stock(items, warehouses, company)
    pending_so_map = _agg_pending_so(items, warehouses, company, so_st, so_wf)
    pending_wo_map = _agg_pending_wo(items, warehouses, company, wo_st, wo_wf)
    pending_po_map = _agg_pending_po(items, warehouses, company, po_st, po_wf)
    remain_wo_map  = _agg_remain_wo_consumption(items, warehouses, company, wo_st, wo_wf)

    # ── Step 4: Item meta — group + name.
    item_meta = _get_item_meta(items)

    # ── Step 5: Assemble + apply boolean filters.
    rows = []
    for item in items:
        meta = item_meta.get(item) or {}
        uom = uom_map.get(item) or {"higher_uom": "", "cf": 1.0, "stock_uom": ""}
        cf = uom["cf"]

        cur_stock = flt(stock_map.get(item) or 0)
        pso       = flt(pending_so_map.get(item) or 0)
        pwo       = flt(pending_wo_map.get(item) or 0)
        ppo       = flt(pending_po_map.get(item) or 0)
        rwc       = flt(remain_wo_map.get(item) or 0)

        # v0.0.35 — Compute BOTH modes in parallel so the dashboard can
        # display side-by-side columns. Demand stays constant (pending SO
        # + remaining WO consumption); only the supply side differs.
        #
        # PROJECTED mode (default columns, kept un-suffixed for back-compat):
        #   supply = current_stock + pending_wo + pending_po
        # CURRENT-STOCK mode (new "_cs" suffixed columns):
        #   supply = current_stock ONLY
        demand_stock = pso + rwc
        # ── Projected supply
        supply_stock = cur_stock + pwo + ppo
        net_stock    = supply_stock - demand_stock
        if net_stock >= 0:
            status   = "Surplus"
            shortfall_stock = 0.0
            surplus_stock   = net_stock
        else:
            status   = "Shortage"
            shortfall_stock = -net_stock
            surplus_stock   = 0.0

        # ── Current-Stock-only supply (parallel computation, "_cs" suffix)
        supply_cs = cur_stock
        net_cs    = supply_cs - demand_stock
        if net_cs >= 0:
            status_cs        = "Surplus"
            shortfall_cs_stk = 0.0
            surplus_cs_stk   = net_cs
        else:
            status_cs        = "Shortage"
            shortfall_cs_stk = -net_cs
            surplus_cs_stk   = 0.0

        row = {
            "item_group":  meta.get("item_group") or "",
            "item_code":   item,
            "item_name":   meta.get("item_name") or item,
            "higher_uom":  uom["higher_uom"],
            "stock_uom":   uom["stock_uom"],
            "cf":          cf,
            # qty pairs — stock UOM + higher UOM mirrored
            "current_stock_stock": cur_stock,
            "current_stock_higher": _to_higher(cur_stock, cf),
            "pending_so_stock":   pso,
            "pending_so_higher":  _to_higher(pso, cf),
            "pending_wo_stock":   pwo,
            "pending_wo_higher":  _to_higher(pwo, cf),
            "pending_po_stock":   ppo,
            "pending_po_higher":  _to_higher(ppo, cf),
            "remain_wo_stock":    rwc,
            "remain_wo_higher":   _to_higher(rwc, cf),
            "total_demand_stock":  demand_stock,
            "total_demand_higher": _to_higher(demand_stock, cf),
            # PROJECTED-mode columns (existing — kept un-suffixed)
            "demand_status":      status,
            "shortfall_stock":    shortfall_stock,
            "shortfall_higher":   _to_higher(shortfall_stock, cf),
            "surplus_stock":      surplus_stock,
            "surplus_higher":     _to_higher(surplus_stock, cf),
            # v0.0.35 — CURRENT-STOCK-only columns (parallel set, "_cs")
            "demand_status_cs":   status_cs,
            "shortfall_cs_stock":  shortfall_cs_stk,
            "shortfall_cs_higher": _to_higher(shortfall_cs_stk, cf),
            "surplus_cs_stock":    surplus_cs_stk,
            "surplus_cs_higher":   _to_higher(surplus_cs_stk, cf),
        }

        # Post-aggregation boolean filters (matched against zero-tol > 0).
        if cint(f.get("active_so")) and pso <= 0: continue
        if cint(f.get("active_wo")) and pwo <= 0: continue
        if cint(f.get("active_po")) and ppo <= 0: continue
        if cint(f.get("active_wo_consume")) and rwc <= 0: continue
        if cint(f.get("no_so")) and pso > 0: continue
        if cint(f.get("no_wo")) and pwo > 0: continue
        if cint(f.get("no_po")) and ppo > 0: continue
        if cint(f.get("no_wo_consume")) and rwc > 0: continue
        # Composite filters — sales-vs-production / supply-vs-demand views
        if cint(f.get("so_no_wo")) and not (pso > 0 and pwo <= 0): continue
        if cint(f.get("wo_with_shortage")) and not (pwo > 0 and status == "Shortage"): continue
        if cint(f.get("so_with_shortage")) and not (pso > 0 and status == "Shortage"): continue
        if cint(f.get("shortage_only")) and status != "Shortage": continue
        if cint(f.get("surplus_only")) and status != "Surplus": continue
        # v0.0.35 — new chips that key off the Current-Stock status
        if cint(f.get("po_with_shortage_cs")) and not (ppo > 0 and status_cs == "Shortage"): continue
        if cint(f.get("so_with_shortage_cs")) and not (pso > 0 and status_cs == "Shortage"): continue
        # Pure status-only views per mode
        if cint(f.get("shortage_only_cs")) and status_cs != "Shortage": continue
        if cint(f.get("surplus_only_cs")) and status_cs != "Surplus": continue

        rows.append(row)

    return {
        "rows": rows,
        "columns": _columns(),
        "filters_used": {
            "warehouses": warehouses,
            "company": company,
            # v0.0.26 — surface BOTH the raw pairs (what the user picked)
            # and the derived split lists (what the SQL actually filtered
            # on) so the XLSX audit sheet shows both views.
            "so_pairs":   _as_list(f.get("so_pairs")),
            "wo_pairs":   _as_list(f.get("wo_pairs")),
            "po_pairs":   _as_list(f.get("po_pairs")),
            "so_statuses": so_st,
            "so_workflow_states": so_wf,
            "wo_statuses": wo_st,
            "wo_workflow_states": wo_wf,
            "po_statuses": po_st,
            "po_workflow_states": po_wf,
            "item": item_codes_filter,
            "item_group": item_groups,
            "active_so": cint(f.get("active_so")),
            "active_wo": cint(f.get("active_wo")),
            "active_po": cint(f.get("active_po")),
            "active_wo_consume": cint(f.get("active_wo_consume")),
            "no_so": cint(f.get("no_so")),
            "no_wo": cint(f.get("no_wo")),
            "no_po": cint(f.get("no_po")),
            "no_wo_consume": cint(f.get("no_wo_consume")),
            "so_no_wo": cint(f.get("so_no_wo")),
            "wo_with_shortage": cint(f.get("wo_with_shortage")),
            "so_with_shortage": cint(f.get("so_with_shortage")),
            "shortage_only": cint(f.get("shortage_only")),
            "surplus_only": cint(f.get("surplus_only")),
            # v0.0.35 — current-stock chips
            "po_with_shortage_cs": cint(f.get("po_with_shortage_cs")),
            "so_with_shortage_cs": cint(f.get("so_with_shortage_cs")),
            "shortage_only_cs":    cint(f.get("shortage_only_cs")),
            "surplus_only_cs":     cint(f.get("surplus_only_cs")),
        },
        # v0.0.37 — friendly local format for the operator-visible footer.
        "generated_at": _fmt_local_ts(now_datetime()),
        "user": frappe.session.user,
    }


def _columns() -> list[dict]:
    """Column metadata for the XLSX export — 21 split columns.

    The dashboard JS builds its OWN merged-column view (stock + higher UOM
    stacked into one cell per category, per user spec). XLSX export uses
    this list AS-IS so analysts can pivot/SUM either UOM independently.

    Don't try to unify the two — dashboard wants merged for scanability,
    XLSX wants split for analysis.
    """
    # v0.0.35 — each column carries a `description` field used as the
    # tooltip in the dashboard column header. Layman ERPNext terms only —
    # operators reading these are warehouse / planner / sales folks, not
    # database admins. Keep formulas concrete (cite specific docs).
    return [
        {"field": "item_group",  "title": _("Item Group"),  "type": "text",
         "description": _("Item Group from the Item master (categorization used for filtering and reporting).")},
        {"field": "item_name",   "title": _("Item"),        "type": "item_link",
         "description": _("Item code + name from the Item master. Click to open the Item record.")},
        {"field": "higher_uom",  "title": _("Higher UOM"),  "type": "text",
         "description": _("Higher-level Unit of Measure (e.g. CFC / Master / Carton) configured on the Item's UOM Conversion Detail table.")},
        {"field": "stock_uom",   "title": _("Stock UOM"),   "type": "text",
         "description": _("Stock UOM from the Item master — the base unit ERPNext stores all stock balances in.")},
        {"field": "current_stock_higher", "title": _("Current Stock (Higher UOM)"), "type": "qty", "source": "stock",
         "description": _("Current on-hand stock in the Higher UOM. Sum of Bin actual_qty across the selected Warehouses / Company.")},
        {"field": "current_stock_stock",  "title": _("Current Stock (Stock UOM)"),  "type": "qty", "source": "stock",
         "description": _("Current on-hand stock in the Stock UOM. Sum of Bin actual_qty across the selected Warehouses / Company.")},
        {"field": "pending_so_higher",    "title": _("Pending SO (Higher UOM)"),    "type": "qty", "source": "so",
         "description": _("Open Sales Order quantity yet to be delivered (Sales Order Item: qty − delivered_qty), restricted to the chosen SO Status × Workflow pairs.")},
        {"field": "pending_so_stock",     "title": _("Pending SO (Stock UOM)"),     "type": "qty", "source": "so",
         "description": _("Open Sales Order quantity yet to be delivered (Sales Order Item: qty − delivered_qty), restricted to the chosen SO Status × Workflow pairs.")},
        {"field": "pending_wo_higher",    "title": _("Pending WO (Higher UOM)"),    "type": "qty", "source": "wo",
         "description": _("Quantity from open Work Orders yet to be produced (Work Order: qty − produced_qty), restricted to the chosen WO Status × Workflow pairs.")},
        {"field": "pending_wo_stock",     "title": _("Pending WO (Stock UOM)"),     "type": "qty", "source": "wo",
         "description": _("Quantity from open Work Orders yet to be produced (Work Order: qty − produced_qty), restricted to the chosen WO Status × Workflow pairs.")},
        {"field": "pending_po_higher",    "title": _("Pending PO (Higher UOM)"),    "type": "qty", "source": "po",
         "description": _("Open Purchase Order quantity yet to be received (Purchase Order Item: qty − received_qty), restricted to the chosen PO Status × Workflow pairs.")},
        {"field": "pending_po_stock",     "title": _("Pending PO (Stock UOM)"),     "type": "qty", "source": "po",
         "description": _("Open Purchase Order quantity yet to be received (Purchase Order Item: qty − received_qty), restricted to the chosen PO Status × Workflow pairs.")},
        {"field": "remain_wo_higher",     "title": _("Remain WO Consumption (Higher UOM)"), "type": "qty", "source": "wo_consume",
         "description": _("This item is a COMPONENT in open Work Orders — qty still required to be issued to production (Work Order Item: required_qty − transferred_qty).")},
        {"field": "remain_wo_stock",      "title": _("Remain WO Consumption (Stock UOM)"),  "type": "qty", "source": "wo_consume",
         "description": _("This item is a COMPONENT in open Work Orders — qty still required to be issued to production (Work Order Item: required_qty − transferred_qty).")},
        {"field": "total_demand_higher",  "title": _("Total Demand (Higher UOM)"),  "type": "qty", "source": "demand",
         "description": _("Total expected outflow = Pending SO + Remaining WO Consumption. This is what the warehouse must satisfy.")},
        {"field": "total_demand_stock",   "title": _("Total Demand (Stock UOM)"),   "type": "qty", "source": "demand",
         "description": _("Total expected outflow = Pending SO + Remaining WO Consumption. This is what the warehouse must satisfy.")},
        # PROJECTED-mode (supply = stock + pending WO + pending PO)
        {"field": "demand_status",        "title": _("Status as per Projected"), "type": "pill",
         "description": _("Shortage or Surplus assuming all pending Work Orders are produced AND all pending Purchase Orders are received. Formula: (Current Stock + Pending WO + Pending PO) − Total Demand.")},
        {"field": "shortfall_higher",     "title": _("Shortage as per Projected (Higher UOM)"), "type": "qty_proj_short",
         "description": _("How much MORE you'll still be short even AFTER all pending production + purchases land. (Demand − Stock − Pending WO − Pending PO), clamped to 0.")},
        {"field": "shortfall_stock",      "title": _("Shortage as per Projected (Stock UOM)"),  "type": "qty_proj_short",
         "description": _("How much MORE you'll still be short even AFTER all pending production + purchases land. (Demand − Stock − Pending WO − Pending PO), clamped to 0.")},
        {"field": "surplus_higher",       "title": _("Surplus as per Projected (Higher UOM)"),  "type": "qty_proj_surp",
         "description": _("How much excess you'll have AFTER all pending production + purchases land. (Stock + Pending WO + Pending PO − Demand), clamped to 0.")},
        {"field": "surplus_stock",        "title": _("Surplus as per Projected (Stock UOM)"),   "type": "qty_proj_surp",
         "description": _("How much excess you'll have AFTER all pending production + purchases land. (Stock + Pending WO + Pending PO − Demand), clamped to 0.")},
        # CURRENT-STOCK-only mode (parallel set, v0.0.35) — supply = on-hand only
        {"field": "demand_status_cs",     "title": _("Status as per Current Stock"), "type": "pill_cs",
         "description": _("Shortage or Surplus right NOW based ONLY on physical stock on the shelf — pending WO / PO not counted. Formula: Current Stock − Total Demand.")},
        {"field": "shortfall_cs_higher",  "title": _("Shortage as per Current Stock (Higher UOM)"), "type": "qty_cs_short",
         "description": _("How much you're short RIGHT NOW from on-hand stock alone, ignoring pending production / purchases. (Demand − Current Stock), clamped to 0.")},
        {"field": "shortfall_cs_stock",   "title": _("Shortage as per Current Stock (Stock UOM)"),  "type": "qty_cs_short",
         "description": _("How much you're short RIGHT NOW from on-hand stock alone, ignoring pending production / purchases. (Demand − Current Stock), clamped to 0.")},
        {"field": "surplus_cs_higher",    "title": _("Surplus as per Current Stock (Higher UOM)"),  "type": "qty_cs_surp",
         "description": _("How much excess on-hand stock you have RIGHT NOW above current demand. (Current Stock − Demand), clamped to 0.")},
        {"field": "surplus_cs_stock",     "title": _("Surplus as per Current Stock (Stock UOM)"),   "type": "qty_cs_surp",
         "description": _("How much excess on-hand stock you have RIGHT NOW above current demand. (Current Stock − Demand), clamped to 0.")},
    ]


# -----------------------------------------------------------------------------
# Item resolution
# -----------------------------------------------------------------------------

def _resolve_items(items: list[str], item_groups: list[str]) -> list[str]:
    """Resolve filter inputs to a concrete list of item codes.

    LIVE filter accuracy: item_group filter runs `Item.item_group IN (...)`
    at SELECT time — never reads a fetched mirror. Reclassifications
    propagate immediately.

    If both filters are empty, returns items that appear in ANY of:
    Bin (with actual_qty > 0), pending SO Item, pending WO, or pending PO
    Item — that's the "everything that could possibly be relevant" set.
    A scope-everything default would return thousands of irrelevant items.
    """
    if items or item_groups:
        # Explicit scope — straight Item lookup
        wh = []
        params: dict = {}
        if items:
            wh.append("name IN %(items)s")
            params["items"] = tuple(items)
        if item_groups:
            wh.append("item_group IN %(igs)s")
            params["igs"] = tuple(item_groups)
        rows = frappe.db.sql(f"""
            SELECT name FROM `tabItem`
             WHERE {' AND '.join(wh)}
             ORDER BY name
        """, params, as_dict=True)
        return [r["name"] for r in rows]

    # Default scope: items with any current stock OR any pending voucher.
    # Union of 4 sources, deduped.
    items_set: set[str] = set()
    try:
        rows = frappe.db.sql("""
            SELECT DISTINCT item_code FROM `tabBin`
             WHERE IFNULL(actual_qty, 0) > 0
        """, as_dict=True)
        items_set.update(r["item_code"] for r in rows if r["item_code"])
    except Exception:
        pass
    try:
        rows = frappe.db.sql("""
            SELECT DISTINCT item_code FROM `tabSales Order Item`
             WHERE IFNULL(qty, 0) - IFNULL(delivered_qty, 0) > 0
        """, as_dict=True)
        items_set.update(r["item_code"] for r in rows if r["item_code"])
    except Exception:
        pass
    try:
        rows = frappe.db.sql("""
            SELECT DISTINCT production_item FROM `tabWork Order`
             WHERE IFNULL(qty, 0) - IFNULL(produced_qty, 0) > 0
               AND docstatus < 2
        """, as_dict=True)
        items_set.update(r["production_item"] for r in rows if r["production_item"])
    except Exception:
        pass
    try:
        rows = frappe.db.sql("""
            SELECT DISTINCT item_code FROM `tabPurchase Order Item`
             WHERE IFNULL(qty, 0) - IFNULL(received_qty, 0) > 0
        """, as_dict=True)
        items_set.update(r["item_code"] for r in rows if r["item_code"])
    except Exception:
        pass
    return sorted(items_set)


def _get_item_meta(item_codes: list[str]) -> dict[str, dict]:
    """Live read of Item.item_group + item_name. No mirror, no cache."""
    if not item_codes:
        return {}
    rows = frappe.db.sql("""
        SELECT name, item_name, item_group
          FROM `tabItem`
         WHERE name IN %(items)s
    """, {"items": tuple(item_codes)}, as_dict=True)
    return {r["name"]: {"item_name": r["item_name"], "item_group": r["item_group"]} for r in rows}


# -----------------------------------------------------------------------------
# Bulk aggregators — one query each, joined back to items in Python.
# -----------------------------------------------------------------------------

def _wh_company_filter(alias: str, warehouses: list[str], company: str,
                       params: dict) -> str:
    """Build a (warehouse, company) WHERE fragment for a query whose
    warehouse column is `<alias>.warehouse`. Empty fragments => "1=1"."""
    parts = []
    if warehouses:
        parts.append(f"{alias}.warehouse IN %(whs)s")
        params["whs"] = tuple(warehouses)
    if company:
        parts.append(
            f"{alias}.warehouse IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)"
        )
        params["co"] = company
    return " AND ".join(parts) if parts else "1=1"


def _agg_current_stock(items: list[str], warehouses: list[str], company: str) -> dict[str, float]:
    params = {"items": tuple(items)}
    where_wh = _wh_company_filter("b", warehouses, company, params)
    rows = frappe.db.sql(f"""
        SELECT b.item_code AS item, COALESCE(SUM(b.actual_qty), 0) AS qty
          FROM `tabBin` b
         WHERE b.item_code IN %(items)s AND {where_wh}
         GROUP BY b.item_code
    """, params, as_dict=True)
    return {r["item"]: flt(r["qty"]) for r in rows}


def _so_eligibility_sql(so_statuses: list[str], so_workflow_states: list[str],
                        alias: str = "so") -> tuple[str, list]:
    """Return (sql_fragment, param_values_in_order) for SO pending check.

    Always docstatus < 2. Status path: docstatus=1 AND status IN <user>.
    Workflow path: docstatus=0 AND workflow_state IN <user> (only when
    workflow_state column exists).
    Always excludes Closed + Cancelled status regardless of user list.
    """
    parts = []
    values: list = []
    if so_statuses:
        ph = ", ".join(["%s"] * len(so_statuses))
        parts.append(
            f"({alias}.docstatus = 1 AND {alias}.status IN ({ph}) "
            f"AND {alias}.status NOT IN ('Closed','Cancelled'))"
        )
        values.extend(so_statuses)
    if so_workflow_states and _so_has_wf():
        ph = ", ".join(["%s"] * len(so_workflow_states))
        parts.append(f"({alias}.docstatus = 0 AND {alias}.workflow_state IN ({ph}))")
        values.extend(so_workflow_states)
    if not parts:
        return "1=0", []
    return " OR ".join(parts), values


def _wo_eligibility_sql_local(wo_statuses: list[str], wo_workflow_states: list[str],
                               alias: str = "wo") -> tuple[str, list]:
    """Local copy of the engine helper that returns the SQL + params list,
    so the param order is explicit at the call site."""
    parts = []
    values: list = []
    if wo_statuses:
        ph = ", ".join(["%s"] * len(wo_statuses))
        parts.append(
            f"({alias}.docstatus = 1 AND {alias}.status IN ({ph}) "
            f"AND {alias}.status NOT IN ('Closed','Cancelled','Stopped'))"
        )
        values.extend(wo_statuses)
    if wo_workflow_states and _wo_has_wf():
        ph = ", ".join(["%s"] * len(wo_workflow_states))
        parts.append(f"({alias}.docstatus = 0 AND {alias}.workflow_state IN ({ph}))")
        values.extend(wo_workflow_states)
    if not parts:
        return "1=0", []
    return " OR ".join(parts), values


def _po_eligibility_sql_local(po_statuses: list[str], po_workflow_states: list[str],
                               alias: str = "po") -> tuple[str, list]:
    parts = []
    values: list = []
    if po_statuses:
        ph = ", ".join(["%s"] * len(po_statuses))
        parts.append(
            f"({alias}.docstatus = 1 AND {alias}.status IN ({ph}) "
            f"AND {alias}.status NOT IN ('Closed','Cancelled'))"
        )
        values.extend(po_statuses)
    if po_workflow_states and _po_has_wf():
        ph = ", ".join(["%s"] * len(po_workflow_states))
        parts.append(f"({alias}.docstatus = 0 AND {alias}.workflow_state IN ({ph}))")
        values.extend(po_workflow_states)
    if not parts:
        return "1=0", []
    return " OR ".join(parts), values


def _agg_pending_so(items: list[str], warehouses: list[str], company: str,
                    so_statuses: list[str], so_workflow_states: list[str]) -> dict[str, float]:
    if not items: return {}
    elig, elig_params = _so_eligibility_sql(so_statuses, so_workflow_states, "so")
    if elig == "1=0":
        return {}
    # Build (warehouse, company) on SO item's warehouse — falls back to SO.set_warehouse
    # via COALESCE if soi.warehouse blank. Pragmatic — most chaizup SOs have
    # item-level warehouse but some legacy use the header field.
    where_parts = ["soi.item_code IN %(items)s"]
    params: dict = {"items": tuple(items)}
    if warehouses:
        where_parts.append(
            "COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) IN %(whs)s"
        )
        params["whs"] = tuple(warehouses)
    if company:
        where_parts.append(
            "COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) "
            "IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)"
        )
        params["co"] = company
    # Append eligibility params positionally (Frappe expects a tuple of
    # positional values when the SQL has both %(name)s and %s placeholders;
    # we use a workaround — embed positional values as named).
    for i, v in enumerate(elig_params):
        params[f"_elig_{i}"] = v
    # Rewrite the elig SQL to use named placeholders for safe interleave
    elig_named = elig
    for i in range(len(elig_params)):
        elig_named = elig_named.replace("%s", f"%(_elig_{i})s", 1)

    sql = f"""
        SELECT soi.item_code AS item,
               COALESCE(SUM(GREATEST(IFNULL(soi.qty,0) - IFNULL(soi.delivered_qty,0), 0)), 0) AS qty
          FROM `tabSales Order Item` soi
          JOIN `tabSales Order` so ON so.name = soi.parent
         WHERE {' AND '.join(where_parts)}
           AND so.docstatus < 2
           AND so.status NOT IN ('Closed','Cancelled')
           AND ({elig_named})
         GROUP BY soi.item_code
    """
    rows = frappe.db.sql(sql, params, as_dict=True)
    return {r["item"]: flt(r["qty"]) for r in rows}


def _agg_pending_wo(items: list[str], warehouses: list[str], company: str,
                    wo_statuses: list[str], wo_workflow_states: list[str]) -> dict[str, float]:
    if not items: return {}
    elig, elig_params = _wo_eligibility_sql_local(wo_statuses, wo_workflow_states, "wo")
    if elig == "1=0":
        return {}
    where_parts = ["wo.production_item IN %(items)s"]
    params: dict = {"items": tuple(items)}
    # WO warehouse = fg_warehouse (where the FG ends up) — that's the demand
    # signal we care about, not source_warehouse (where components come from).
    if warehouses:
        where_parts.append("wo.fg_warehouse IN %(whs)s")
        params["whs"] = tuple(warehouses)
    if company:
        where_parts.append(
            "wo.fg_warehouse IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)"
        )
        params["co"] = company
    for i, v in enumerate(elig_params):
        params[f"_elig_{i}"] = v
    elig_named = elig
    for i in range(len(elig_params)):
        elig_named = elig_named.replace("%s", f"%(_elig_{i})s", 1)
    sql = f"""
        SELECT wo.production_item AS item,
               COALESCE(SUM(GREATEST(IFNULL(wo.qty,0) - IFNULL(wo.produced_qty,0), 0)), 0) AS qty
          FROM `tabWork Order` wo
         WHERE {' AND '.join(where_parts)}
           AND ({elig_named})
         GROUP BY wo.production_item
    """
    rows = frappe.db.sql(sql, params, as_dict=True)
    return {r["item"]: flt(r["qty"]) for r in rows}


def _agg_pending_po(items: list[str], warehouses: list[str], company: str,
                    po_statuses: list[str], po_workflow_states: list[str]) -> dict[str, float]:
    if not items: return {}
    elig, elig_params = _po_eligibility_sql_local(po_statuses, po_workflow_states, "po")
    if elig == "1=0":
        return {}
    where_parts = ["poi.item_code IN %(items)s"]
    params: dict = {"items": tuple(items)}
    if warehouses:
        where_parts.append("poi.warehouse IN %(whs)s")
        params["whs"] = tuple(warehouses)
    if company:
        where_parts.append(
            "poi.warehouse IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)"
        )
        params["co"] = company
    for i, v in enumerate(elig_params):
        params[f"_elig_{i}"] = v
    elig_named = elig
    for i in range(len(elig_params)):
        elig_named = elig_named.replace("%s", f"%(_elig_{i})s", 1)
    sql = f"""
        SELECT poi.item_code AS item,
               COALESCE(SUM(GREATEST(IFNULL(poi.qty,0) - IFNULL(poi.received_qty,0), 0)), 0) AS qty
          FROM `tabPurchase Order Item` poi
          JOIN `tabPurchase Order` po ON po.name = poi.parent
         WHERE {' AND '.join(where_parts)}
           AND po.docstatus < 2
           AND po.status NOT IN ('Closed','Cancelled')
           AND ({elig_named})
         GROUP BY poi.item_code
    """
    rows = frappe.db.sql(sql, params, as_dict=True)
    return {r["item"]: flt(r["qty"]) for r in rows}


def _agg_remain_wo_consumption(items: list[str], warehouses: list[str], company: str,
                                wo_statuses: list[str], wo_workflow_states: list[str]) -> dict[str, float]:
    """Σ(required_qty − transferred_qty) on Work Order Item rows where the
    parent WO is pending. The component item (woi.item_code) is what gets
    demanded; the WO's fg_warehouse scopes "demand owned by which warehouse".
    """
    if not items: return {}
    elig, elig_params = _wo_eligibility_sql_local(wo_statuses, wo_workflow_states, "wo")
    if elig == "1=0":
        return {}
    where_parts = ["woi.item_code IN %(items)s"]
    params: dict = {"items": tuple(items)}
    if warehouses:
        where_parts.append("wo.fg_warehouse IN %(whs)s")
        params["whs"] = tuple(warehouses)
    if company:
        where_parts.append(
            "wo.fg_warehouse IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)"
        )
        params["co"] = company
    for i, v in enumerate(elig_params):
        params[f"_elig_{i}"] = v
    elig_named = elig
    for i in range(len(elig_params)):
        elig_named = elig_named.replace("%s", f"%(_elig_{i})s", 1)
    sql = f"""
        SELECT woi.item_code AS item,
               COALESCE(SUM(GREATEST(IFNULL(woi.required_qty,0) - IFNULL(woi.transferred_qty,0), 0)), 0) AS qty
          FROM `tabWork Order Item` woi
          JOIN `tabWork Order` wo ON wo.name = woi.parent
         WHERE {' AND '.join(where_parts)}
           AND ({elig_named})
         GROUP BY woi.item_code
    """
    rows = frappe.db.sql(sql, params, as_dict=True)
    return {r["item"]: flt(r["qty"]) for r in rows}


# -----------------------------------------------------------------------------
# Endpoint 3 — voucher drill-down
# -----------------------------------------------------------------------------

@frappe.whitelist()
def get_voucher_drilldown(item_code: str, source: str, filters: Any = None) -> dict:
    """Return per-voucher rows for the clicked cell.

    source ∈ {"stock", "so", "wo", "po", "wo_consume", "demand"}
      - "demand" reduces to so + wo_consume (the additive components).
    """
    f = frappe.parse_json(filters) if isinstance(filters, str) else (filters or {})
    warehouses = _as_list(f.get("warehouses"))
    company    = (f.get("company") or "").strip()

    # v0.0.26 — same combined-pairs splitter as get_report.
    def _split_pairs(pairs_key: str, st_key: str, wf_key: str):
        pairs = _as_list(f.get(pairs_key))
        if pairs:
            st, wf = [], []
            for p in pairs:
                if "|" not in p:
                    st.append(p); continue
                s_part, _, w_part = p.partition("|")
                if s_part: st.append(s_part)
                if w_part: wf.append(w_part)
            return list(dict.fromkeys(st)), list(dict.fromkeys(wf))
        return _as_list(f.get(st_key)), _as_list(f.get(wf_key))

    so_st, so_wf = _split_pairs("so_pairs", "so_statuses", "so_workflow_states")
    wo_st, wo_wf = _split_pairs("wo_pairs", "wo_statuses", "wo_workflow_states")
    po_st, po_wf = _split_pairs("po_pairs", "po_statuses", "po_workflow_states")

    uom = _pick_higher_uoms([item_code]).get(item_code) or {"higher_uom": "", "cf": 1.0, "stock_uom": ""}
    cf = uom["cf"]

    out_rows: list[dict] = []

    if source == "stock":
        # v0.0.36 — Richer per-warehouse view: pull Actual + Reserved +
        # Ordered + Planned + Projected from Bin, plus the Warehouse's
        # Company. Operator gets the full warehouse-manager dashboard in
        # one click instead of just on-hand qty.
        #
        # ERPNext Bin field meanings (layman terms):
        #   actual_qty     — what's physically on the shelf right now
        #   reserved_qty   — already committed to open Sales Orders
        #   ordered_qty    — coming in from open Purchase Orders
        #   planned_qty    — coming in from open Work Orders (FG)
        #   projected_qty  — actual + ordered + planned − reserved − indented
        #                    (ERPNext's own "what you'll really have" number)
        params: dict = {"item": item_code}
        wh = _wh_company_filter("b", warehouses, company, params)
        rows = frappe.db.sql(f"""
            SELECT b.warehouse                AS voucher_no,
                   w.company                  AS warehouse_company,
                   IFNULL(b.actual_qty, 0)    AS actual_qty,
                   IFNULL(b.reserved_qty, 0)  AS reserved_qty,
                   IFNULL(b.ordered_qty, 0)   AS ordered_qty,
                   IFNULL(b.planned_qty, 0)   AS planned_qty,
                   IFNULL(b.projected_qty, 0) AS projected_qty
              FROM `tabBin` b
              LEFT JOIN `tabWarehouse` w ON w.name = b.warehouse
             WHERE b.item_code = %(item)s AND {wh}
               AND (IFNULL(b.actual_qty, 0) != 0
                    OR IFNULL(b.reserved_qty, 0) != 0
                    OR IFNULL(b.ordered_qty, 0) != 0
                    OR IFNULL(b.planned_qty, 0) != 0)
             ORDER BY b.warehouse
        """, params, as_dict=True)
        out_rows = [{
            "voucher_type":     "Warehouse Bin",
            "voucher_no":       r["voucher_no"],
            "voucher_link":     f"/app/warehouse/{r['voucher_no']}",
            "warehouse_company": r["warehouse_company"] or "",
            "posting_date":     "",
            # Actual on-hand (canonical "Current Stock")
            "actual_stock":     flt(r["actual_qty"]),
            "actual_higher":    _to_higher(flt(r["actual_qty"]), cf),
            # v0.0.36 — extra warehouse-level columns
            "reserved_stock":   flt(r["reserved_qty"]),
            "reserved_higher":  _to_higher(flt(r["reserved_qty"]), cf),
            "ordered_stock":    flt(r["ordered_qty"]),
            "ordered_higher":   _to_higher(flt(r["ordered_qty"]), cf),
            "planned_stock":    flt(r["planned_qty"]),
            "planned_higher":   _to_higher(flt(r["planned_qty"]), cf),
            "projected_stock":  flt(r["projected_qty"]),
            "projected_higher": _to_higher(flt(r["projected_qty"]), cf),
        } for r in rows]

    elif source == "so":
        out_rows = _drill_so(item_code, warehouses, company, so_st, so_wf, cf)

    elif source == "wo":
        out_rows = _drill_wo(item_code, warehouses, company, wo_st, wo_wf, cf)

    elif source == "po":
        out_rows = _drill_po(item_code, warehouses, company, po_st, po_wf, cf)

    elif source == "wo_consume":
        out_rows = _drill_wo_consume(item_code, warehouses, company, wo_st, wo_wf, cf)

    elif source == "demand":
        out_rows = (_drill_so(item_code, warehouses, company, so_st, so_wf, cf) +
                    _drill_wo_consume(item_code, warehouses, company, wo_st, wo_wf, cf))

    return {
        "item_code": item_code,
        "source": source,
        "higher_uom": uom["higher_uom"],
        "stock_uom": uom["stock_uom"],
        "cf": cf,
        "rows": out_rows,
    }


def _drill_so(item, whs, co, st, wf, cf):
    elig, ep = _so_eligibility_sql(st, wf, "so")
    if elig == "1=0": return []
    where = ["soi.item_code = %(item)s"]
    p: dict = {"item": item}
    if whs:
        where.append("COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) IN %(whs)s")
        p["whs"] = tuple(whs)
    if co:
        where.append("COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) "
                     "IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)")
        p["co"] = co
    for i, v in enumerate(ep): p[f"_e{i}"] = v
    elig_n = elig
    for i in range(len(ep)): elig_n = elig_n.replace("%s", f"%(_e{i})s", 1)
    rows = frappe.db.sql(f"""
        SELECT soi.parent AS voucher_no, so.transaction_date AS posting_date,
               IFNULL(soi.qty, 0) AS planned,
               IFNULL(soi.delivered_qty, 0) AS delivered,
               GREATEST(IFNULL(soi.qty,0) - IFNULL(soi.delivered_qty,0), 0) AS pending
          FROM `tabSales Order Item` soi
          JOIN `tabSales Order` so ON so.name = soi.parent
         WHERE {' AND '.join(where)} AND so.docstatus < 2
           AND so.status NOT IN ('Closed','Cancelled')
           AND ({elig_n})
         ORDER BY so.transaction_date DESC
    """, p, as_dict=True)
    return [{
        "voucher_type": "Sales Order",
        "voucher_no": r["voucher_no"],
        "voucher_link": f"/app/sales-order/{r['voucher_no']}",
        "posting_date": str(r["posting_date"] or ""),
        "planned_stock": flt(r["planned"]),
        "planned_higher": _to_higher(flt(r["planned"]), cf),
        "actual_stock": flt(r["delivered"]),
        "actual_higher": _to_higher(flt(r["delivered"]), cf),
        "pending_stock": flt(r["pending"]),
        "pending_higher": _to_higher(flt(r["pending"]), cf),
    } for r in rows]


def _drill_wo(item, whs, co, st, wf, cf):
    elig, ep = _wo_eligibility_sql_local(st, wf, "wo")
    if elig == "1=0": return []
    where = ["wo.production_item = %(item)s"]
    p: dict = {"item": item}
    if whs:
        where.append("wo.fg_warehouse IN %(whs)s")
        p["whs"] = tuple(whs)
    if co:
        where.append("wo.fg_warehouse IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)")
        p["co"] = co
    for i, v in enumerate(ep): p[f"_e{i}"] = v
    elig_n = elig
    for i in range(len(ep)): elig_n = elig_n.replace("%s", f"%(_e{i})s", 1)
    rows = frappe.db.sql(f"""
        SELECT wo.name AS voucher_no, wo.planned_start_date AS posting_date,
               IFNULL(wo.qty, 0) AS planned,
               IFNULL(wo.produced_qty, 0) AS produced,
               GREATEST(IFNULL(wo.qty,0) - IFNULL(wo.produced_qty,0), 0) AS pending,
               wo.status, wo.fg_warehouse
          FROM `tabWork Order` wo
         WHERE {' AND '.join(where)} AND ({elig_n})
         ORDER BY wo.planned_start_date DESC
    """, p, as_dict=True)
    return [{
        "voucher_type": "Work Order",
        "voucher_no": r["voucher_no"],
        "voucher_link": f"/app/work-order/{r['voucher_no']}",
        "posting_date": str(r["posting_date"] or ""),
        "planned_stock": flt(r["planned"]),
        "planned_higher": _to_higher(flt(r["planned"]), cf),
        "actual_stock": flt(r["produced"]),
        "actual_higher": _to_higher(flt(r["produced"]), cf),
        "pending_stock": flt(r["pending"]),
        "pending_higher": _to_higher(flt(r["pending"]), cf),
    } for r in rows]


def _drill_po(item, whs, co, st, wf, cf):
    elig, ep = _po_eligibility_sql_local(st, wf, "po")
    if elig == "1=0": return []
    where = ["poi.item_code = %(item)s"]
    p: dict = {"item": item}
    if whs:
        where.append("poi.warehouse IN %(whs)s")
        p["whs"] = tuple(whs)
    if co:
        where.append("poi.warehouse IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)")
        p["co"] = co
    for i, v in enumerate(ep): p[f"_e{i}"] = v
    elig_n = elig
    for i in range(len(ep)): elig_n = elig_n.replace("%s", f"%(_e{i})s", 1)
    rows = frappe.db.sql(f"""
        SELECT poi.parent AS voucher_no, po.transaction_date AS posting_date,
               IFNULL(poi.qty, 0) AS planned,
               IFNULL(poi.received_qty, 0) AS received,
               GREATEST(IFNULL(poi.qty,0) - IFNULL(poi.received_qty,0), 0) AS pending
          FROM `tabPurchase Order Item` poi
          JOIN `tabPurchase Order` po ON po.name = poi.parent
         WHERE {' AND '.join(where)} AND po.docstatus < 2
           AND po.status NOT IN ('Closed','Cancelled')
           AND ({elig_n})
         ORDER BY po.transaction_date DESC
    """, p, as_dict=True)
    return [{
        "voucher_type": "Purchase Order",
        "voucher_no": r["voucher_no"],
        "voucher_link": f"/app/purchase-order/{r['voucher_no']}",
        "posting_date": str(r["posting_date"] or ""),
        "planned_stock": flt(r["planned"]),
        "planned_higher": _to_higher(flt(r["planned"]), cf),
        "actual_stock": flt(r["received"]),
        "actual_higher": _to_higher(flt(r["received"]), cf),
        "pending_stock": flt(r["pending"]),
        "pending_higher": _to_higher(flt(r["pending"]), cf),
    } for r in rows]


def _drill_wo_consume(item, whs, co, st, wf, cf):
    elig, ep = _wo_eligibility_sql_local(st, wf, "wo")
    if elig == "1=0": return []
    where = ["woi.item_code = %(item)s"]
    p: dict = {"item": item}
    if whs:
        where.append("wo.fg_warehouse IN %(whs)s")
        p["whs"] = tuple(whs)
    if co:
        where.append("wo.fg_warehouse IN (SELECT name FROM `tabWarehouse` WHERE company = %(co)s)")
        p["co"] = co
    for i, v in enumerate(ep): p[f"_e{i}"] = v
    elig_n = elig
    for i in range(len(ep)): elig_n = elig_n.replace("%s", f"%(_e{i})s", 1)
    rows = frappe.db.sql(f"""
        SELECT woi.parent AS voucher_no, wo.planned_start_date AS posting_date,
               IFNULL(woi.required_qty, 0) AS planned,
               IFNULL(woi.transferred_qty, 0) AS transferred,
               GREATEST(IFNULL(woi.required_qty,0) - IFNULL(woi.transferred_qty,0), 0) AS pending,
               wo.production_item AS for_fg
          FROM `tabWork Order Item` woi
          JOIN `tabWork Order` wo ON wo.name = woi.parent
         WHERE {' AND '.join(where)} AND ({elig_n})
         ORDER BY wo.planned_start_date DESC
    """, p, as_dict=True)
    return [{
        "voucher_type": "WO Consumption",
        "voucher_no": r["voucher_no"],
        "voucher_link": f"/app/work-order/{r['voucher_no']}",
        "posting_date": str(r["posting_date"] or ""),
        "planned_stock": flt(r["planned"]),
        "planned_higher": _to_higher(flt(r["planned"]), cf),
        "actual_stock": flt(r["transferred"]),
        "actual_higher": _to_higher(flt(r["transferred"]), cf),
        "pending_stock": flt(r["pending"]),
        "pending_higher": _to_higher(flt(r["pending"]), cf),
        "for_fg": r["for_fg"],
    } for r in rows]


# -----------------------------------------------------------------------------
# v0.0.33 — Drilldown XLSX export (per-modal)
#
# Sibling of the main report export_xlsx. Operator clicks the "Export Excel"
# button inside the voucher drilldown modal and receives a 2-sheet workbook:
#   Sheet 1: drilldown rows (voucher type pill + voucher + planned/actual/pending)
#   Sheet 2: filter + item metadata snapshot for the run
#
# RESTRICT: do NOT merge this with the main export_xlsx — the drilldown's
# row shape is per-voucher, not per-item; merging would force one of them
# to grow a discriminator column and complicate both consumers.
# -----------------------------------------------------------------------------

@frappe.whitelist()
def export_drilldown_xlsx(item_code: str, source: str, filters: Any = None) -> None:
    """Build a 2-sheet branded XLSX for a single voucher-drilldown and
    stream to browser. Reuses get_voucher_drilldown for the payload so
    the same eligibility + filter contracts apply."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = get_voucher_drilldown(item_code=item_code, source=source, filters=filters)
    rows = data.get("rows", [])
    hu = data.get("higher_uom") or "—"
    su = data.get("stock_uom") or "—"
    cf = flt(data.get("cf") or 1)

    # Style tokens (match main export_xlsx for brand consistency)
    header_fill   = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin   = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    qty_align     = Alignment(horizontal="right", vertical="center")
    text_align    = Alignment(horizontal="left",  vertical="center")
    fill_planned  = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
    fill_actual   = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fill_pending  = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

    wb = Workbook()
    ws = wb.active
    ws.title = "Drilldown"[:31]

    SOURCE_LABEL = {
        "stock":      "Current Stock by Warehouse",
        "so":         "Pending Sales Orders",
        "wo":         "Pending Work Orders",
        "po":         "Pending Purchase Orders",
        "wo_consume": "Remaining WO Consumption",
        "demand":     "Total Demand",
    }.get(source, source)

    if source == "stock":
        # v0.0.36 — Warehouse-manager view: per-warehouse Actual /
        # Reserved / Ordered (PO) / Planned (WO) / Projected, plus
        # the Warehouse's Company for multi-company filtering.
        columns = [
            ("Warehouse",             "voucher_no",        "text"),
            ("Company",               "warehouse_company", "text"),
            (f"On-Hand ({su})",       "actual_stock",      "qty_actual"),
            (f"On-Hand ({hu})",       "actual_higher",     "qty_actual"),
            (f"Reserved ({su})",      "reserved_stock",    "qty_pending"),
            (f"Reserved ({hu})",      "reserved_higher",   "qty_pending"),
            (f"Ordered PO ({su})",    "ordered_stock",     "qty_planned"),
            (f"Ordered PO ({hu})",    "ordered_higher",    "qty_planned"),
            (f"Planned WO ({su})",    "planned_stock",     "qty_planned"),
            (f"Planned WO ({hu})",    "planned_higher",    "qty_planned"),
            (f"Projected ({su})",     "projected_stock",   "qty_actual"),
            (f"Projected ({hu})",     "projected_higher",  "qty_actual"),
        ]
    else:
        columns = [
            ("Voucher",            "voucher_no",      "text"),
            ("Voucher Type",       "voucher_type",    "text"),
            ("Date",               "posting_date",    "text"),
            (f"Planned ({su})",    "planned_stock",   "qty_planned"),
            (f"Planned ({hu})",    "planned_higher",  "qty_planned"),
            (f"Actual ({su})",     "actual_stock",    "qty_actual"),
            (f"Actual ({hu})",     "actual_higher",   "qty_actual"),
            (f"Pending ({su})",    "pending_stock",   "qty_pending"),
            (f"Pending ({hu})",    "pending_higher",  "qty_pending"),
        ]

    # Header
    for c, (title, _f, _t) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.fill, cell.font, cell.alignment, cell.border = header_fill, header_font, header_align, border_thin

    # Data
    for r, row in enumerate(rows, 2):
        for c, (_t, field, kind) in enumerate(columns, 1):
            val = row.get(field, "")
            cell = ws.cell(row=r, column=c, value=(val if val != "" else None))
            cell.border = border_thin
            if kind in ("qty", "qty_planned", "qty_actual", "qty_pending"):
                cell.alignment = qty_align
                cell.number_format = "#,##0.000"
                if kind == "qty_planned": cell.fill = fill_planned
                elif kind == "qty_actual": cell.fill = fill_actual
                elif kind == "qty_pending":
                    cell.fill = fill_pending
                    cell.font = Font(bold=True)
            else:
                cell.alignment = text_align

    # Auto widths
    for c, (title, _f, _t) in enumerate(columns, 1):
        max_w = max(
            len(title),
            *(len(str(row.get(_f, "") or "")) for row in rows[:200])
        ) if rows else len(title)
        ws.column_dimensions[get_column_letter(c)].width = min(max_w + 3, 36)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2 — Run info
    ws2 = wb.create_sheet("Run Info")
    meta = [
        ("Source",           SOURCE_LABEL),
        ("Item Code",        item_code),
        ("Higher UOM",       hu),
        ("Stock UOM",        su),
        ("Conversion",       f"1 {hu} = {cf} {su}"),
        ("Row Count",        len(rows)),
        ("Generated By",     frappe.session.user),
        ("Generated At",     _fmt_local_ts(now_datetime())),
    ]
    ws2.cell(row=1, column=1, value="Field").fill = header_fill
    ws2.cell(row=1, column=1).font = header_font
    ws2.cell(row=1, column=2, value="Value").fill = header_fill
    ws2.cell(row=1, column=2).font = header_font
    for i, (k, v) in enumerate(meta, 2):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 60

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_item = item_code.replace("/", "_").replace(" ", "_")
    frappe.response.filename = f"drilldown_{safe_item}_{source}_{now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    frappe.response.filecontent = buf.getvalue()
    frappe.response.type = "binary"
    frappe.response.display_content_as = "attachment"


# -----------------------------------------------------------------------------
# Endpoint 4 — multi-sheet XLSX export
# -----------------------------------------------------------------------------

@frappe.whitelist()
def export_xlsx(filters: Any = None) -> None:
    """Build a multi-sheet branded XLSX in memory + stream to the browser.

    Sheets (v0.0.34 — 8 total):
      1. Main report (all visible columns, filters applied)
      2. Filter snapshot + run metadata (PCAOB-style audit trail)
      3. Shortage-only sorted by shortfall_higher DESC
      4. Surplus-only sorted by surplus_higher DESC
      5. Pending Sales Orders   — voucher × line item, includes party
      6. Pending Work Orders    — voucher × line item, includes prod plan
      7. Pending Purchase Orders — voucher × line item, includes vendor
      8. Pending WO Consumption — voucher × line item, includes for-FG

    CONTEXT: Operators wanted offline access to the full voucher list, not
        just per-item drilldowns. The 4 voucher sheets reuse the same
        eligibility predicates as the live report (so the export is a
        true 1:1 of what the dashboard shows), and every voucher / item /
        customer / supplier / production-plan reference is rendered as
        a clickable absolute hyperlink so users can click straight back
        into Frappe even after importing to Excel or Google Sheets.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from frappe.utils import get_url

    data = get_report(filters=filters)
    rows = data["rows"]
    cols = data["columns"]
    filters_used = data["filters_used"]
    # Absolute site URL — required so hyperlinks survive an XLSX import
    # into Google Sheets / desktop Excel (relative /app/... links open
    # nowhere from outside the browser tab).
    site_url = (get_url() or "").rstrip("/")

    wb = Workbook()
    # ── Style tokens — match chaizup brand colours
    header_fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    qty_align = Alignment(horizontal="right", vertical="center", wrap_text=True)
    text_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    red_font = Font(color="C0392B", bold=True)
    green_font = Font(color="27AE60", bold=True)
    # v0.0.35 — fills for the 4 new shortage/surplus column types + 2 status pills
    fill_proj_short = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # light red
    fill_proj_surp  = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")  # light blue
    fill_cs_short   = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")  # orange
    fill_cs_surp    = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # light green
    fill_pill_proj  = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # very light gray
    fill_pill_cs    = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")  # light gray

    def _write_main(ws, rows_subset, title):
        ws.title = title[:31]  # Excel sheet name max
        # Header row
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=str(col["title"]))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border_thin
        # Data rows
        for r, row in enumerate(rows_subset, 2):
            for c, col in enumerate(cols, 1):
                val = row.get(col["field"], "")
                cell = ws.cell(row=r, column=c, value=val if val != "" else None)
                cell.border = border_thin
                if col["type"] in ("qty", "qty_red", "qty_green",
                                   "qty_proj_short", "qty_proj_surp",
                                   "qty_cs_short", "qty_cs_surp"):
                    cell.alignment = qty_align
                    cell.number_format = "#,##0.000"
                    if col["type"] == "qty_red" and flt(val) > 0:
                        cell.font = red_font
                    elif col["type"] == "qty_green" and flt(val) > 0:
                        cell.font = green_font
                    elif col["type"] == "qty_proj_short":
                        cell.fill = fill_proj_short
                        if flt(val) > 0: cell.font = red_font
                    elif col["type"] == "qty_proj_surp":
                        cell.fill = fill_proj_surp
                    elif col["type"] == "qty_cs_short":
                        cell.fill = fill_cs_short
                        if flt(val) > 0: cell.font = red_font
                    elif col["type"] == "qty_cs_surp":
                        cell.fill = fill_cs_surp
                elif col["type"] == "pill":
                    cell.alignment = text_align
                    cell.fill = fill_pill_proj
                    cell.font = Font(bold=True, color="C0392B" if val == "Shortage" else "27AE60")
                elif col["type"] == "pill_cs":
                    cell.alignment = text_align
                    cell.fill = fill_pill_cs
                    cell.font = Font(bold=True, color="C0392B" if val == "Shortage" else "27AE60")
                elif col["type"] == "item_link":
                    # v0.0.34 — absolute URL so the hyperlink survives a
                    # round-trip into Google Sheets / desktop Excel.
                    cell.alignment = text_align
                    ic = row.get("item_code", "")
                    cell.value = f"{ic} : {row.get('item_name','')}"
                    cell.hyperlink = f"{site_url}/app/item/{ic}"
                    cell.font = Font(color="2980B9", underline="single")
                else:
                    cell.alignment = text_align
        # Auto widths (cap at 40 chars)
        for c, col in enumerate(cols, 1):
            max_w = max(
                len(str(col["title"])),
                *(len(str(row.get(col["field"], "") or "")) for row in rows_subset[:200])
            ) if rows_subset else len(str(col["title"]))
            ws.column_dimensions[get_column_letter(c)].width = min(max_w + 3, 40)
        # Freeze header
        ws.freeze_panes = "A2"
        # Filter / sort header
        if rows_subset:
            ws.auto_filter.ref = ws.dimensions

    # Sheet 1 — Main
    ws1 = wb.active
    _write_main(ws1, rows, "Main Report")

    # Sheet 2 — Filter snapshot + run metadata
    ws2 = wb.create_sheet("Filters & Run Info")
    ws2.cell(row=1, column=1, value="Field").fill = header_fill
    ws2.cell(row=1, column=1).font = header_font
    ws2.cell(row=1, column=2, value="Value").fill = header_fill
    ws2.cell(row=1, column=2).font = header_font
    meta_rows = [
        ("Generated At", data.get("generated_at", "")),
        ("Generated By", data.get("user", "")),
        ("Total Items in Report", len(rows)),
        ("Item Filter", ", ".join(filters_used.get("item") or []) or "(all)"),
        ("Item Group Filter", ", ".join(filters_used.get("item_group") or []) or "(all)"),
        ("Warehouses", ", ".join(filters_used.get("warehouses") or []) or "(all)"),
        ("Company", filters_used.get("company") or "(all)"),
        ("SO Pending Statuses", ", ".join(filters_used.get("so_statuses") or [])),
        ("SO Workflow States", ", ".join(filters_used.get("so_workflow_states") or [])),
        ("WO Pending Statuses", ", ".join(filters_used.get("wo_statuses") or [])),
        ("WO Workflow States", ", ".join(filters_used.get("wo_workflow_states") or [])),
        ("PO Pending Statuses", ", ".join(filters_used.get("po_statuses") or [])),
        ("PO Workflow States", ", ".join(filters_used.get("po_workflow_states") or [])),
        ("Active SO Only",         "Yes" if filters_used.get("active_so") else "No"),
        ("Active WO Only",         "Yes" if filters_used.get("active_wo") else "No"),
        ("Active PO Only",         "Yes" if filters_used.get("active_po") else "No"),
        ("Active WO Consume Only", "Yes" if filters_used.get("active_wo_consume") else "No"),
        ("No SO Only",             "Yes" if filters_used.get("no_so") else "No"),
        ("No WO Only",             "Yes" if filters_used.get("no_wo") else "No"),
        ("No PO Only",             "Yes" if filters_used.get("no_po") else "No"),
        ("No WO Consume Only",     "Yes" if filters_used.get("no_wo_consume") else "No"),
        ("Open SO but No WO Only",             "Yes" if filters_used.get("so_no_wo") else "No"),
        ("Open WO + Shortage (Projected)",     "Yes" if filters_used.get("wo_with_shortage") else "No"),
        ("Open SO + Shortage (Projected)",     "Yes" if filters_used.get("so_with_shortage") else "No"),
        ("Open PO + Shortage (Current Stock)", "Yes" if filters_used.get("po_with_shortage_cs") else "No"),
        ("Open SO + Shortage (Current Stock)", "Yes" if filters_used.get("so_with_shortage_cs") else "No"),
        ("Shortage as per Projection",         "Yes" if filters_used.get("shortage_only") else "No"),
        ("Surplus as per Projection",          "Yes" if filters_used.get("surplus_only") else "No"),
        ("Shortage as per Current Stock",      "Yes" if filters_used.get("shortage_only_cs") else "No"),
        ("Surplus as per Current Stock",       "Yes" if filters_used.get("surplus_only_cs") else "No"),
    ]
    for i, (k, v) in enumerate(meta_rows, 2):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=str(v))
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 60

    # Sheet 3 — Shortage only, sorted desc
    short_rows = sorted(
        [r for r in rows if r.get("demand_status") == "Shortage"],
        key=lambda r: flt(r.get("shortfall_higher") or 0),
        reverse=True,
    )
    ws3 = wb.create_sheet("Shortage (sorted)")
    _write_main(ws3, short_rows, "Shortage (sorted)")

    # Sheet 4 — Surplus only, sorted desc
    surp_rows = sorted(
        [r for r in rows if r.get("demand_status") == "Surplus"],
        key=lambda r: flt(r.get("surplus_higher") or 0),
        reverse=True,
    )
    ws4 = wb.create_sheet("Surplus (sorted)")
    _write_main(ws4, surp_rows, "Surplus (sorted)")

    # ─────────────────────────────────────────────────────────────────
    # v0.0.34 — Sheets 5-8: voucher-level breakdowns.
    #
    # For every item in the report's eligible set, expand its pending
    # SO / WO / PO / WO-Consumption vouchers into individual rows. A
    # single voucher with N line items contributes N rows (voucher_no
    # repeats), which is what the operator asked for.
    #
    # Extra columns the dashboard doesn't show:
    #   • Sales Order      → Customer  (<customer_id> : <customer_name>)
    #   • Work Order       → Production Plan (if any)
    #   • Purchase Order   → Supplier  (<supplier_id> : <supplier_name>)
    #   • WO Consumption   → For FG Item (the finished good being built)
    #
    # All voucher / item / customer / supplier / plan references are
    # rendered as clickable absolute hyperlinks (see site_url above).
    # ─────────────────────────────────────────────────────────────────

    # Re-derive filter args the drill_* helpers expect. get_report has
    # already applied the same filter parsing, so we mirror it here to
    # stay 1:1 with what the dashboard saw.
    f = frappe.parse_json(filters) if isinstance(filters, str) else (filters or {})
    _whs = _as_list(f.get("warehouses"))
    _co  = (f.get("company") or "").strip()
    # Pair-split logic — same as get_report
    def _split(pkey, skey, wkey):
        ps = _as_list(f.get(pkey))
        if ps:
            sts, wfs = [], []
            for p in ps:
                if "|" in p:
                    a, _, b = p.partition("|")
                    if a: sts.append(a)
                    if b: wfs.append(b)
                else:
                    sts.append(p)
            return list(dict.fromkeys(sts)), list(dict.fromkeys(wfs))
        return _as_list(f.get(skey)), _as_list(f.get(wkey))
    so_st, so_wf = _split("so_pairs", "so_statuses", "so_workflow_states")
    wo_st, wo_wf = _split("wo_pairs", "wo_statuses", "wo_workflow_states")
    po_st, po_wf = _split("po_pairs", "po_statuses", "po_workflow_states")
    BLOCK = {"Closed", "Cancelled", "Stopped"}
    so_st = [x for x in so_st if x not in BLOCK]
    wo_st = [x for x in wo_st if x not in BLOCK]
    po_st = [x for x in po_st if x not in BLOCK]

    # Items in report scope (already passed the boolean filters)
    scope_items = [r["item_code"] for r in rows]
    # item meta + uom for hyperlinks + qty fmt
    item_meta_map = _get_item_meta(scope_items) if scope_items else {}
    uom_map_export = _pick_higher_uoms(scope_items) if scope_items else {}

    def _link_cell(ws, r, c, label, url):
        """Helper — set a cell as a clickable, blue, underlined link."""
        cell = ws.cell(row=r, column=c, value=label)
        if url:
            cell.hyperlink = url
            cell.font = Font(color="2980B9", underline="single", bold=False)
        cell.alignment = text_align
        cell.border = border_thin
        return cell

    def _write_voucher_sheet(ws, hdef, data_rows, qty_col_keys, special_cols):
        """Generic writer for the 4 voucher sheets.

        hdef: list[(title, key)] — column definitions in order
        data_rows: list[dict] — one row per voucher line
        qty_col_keys: set[str] — keys rendered as Float/right-aligned
        special_cols: {col_idx (1-based): callable(drow)->url|None} — hyperlinks
        """
        for ci, (title, _key) in enumerate(hdef, 1):
            cell = ws.cell(row=1, column=ci, value=title)
            cell.fill, cell.font, cell.alignment, cell.border = (
                header_fill, header_font, header_align, border_thin
            )
        for ri, drow in enumerate(data_rows, 2):
            for ci, (_t, key) in enumerate(hdef, 1):
                val = drow.get(key, "")
                cell = ws.cell(row=ri, column=ci, value=(val if val != "" else None))
                cell.border = border_thin
                if key in qty_col_keys:
                    cell.alignment = qty_align
                    cell.number_format = "#,##0.000"
                else:
                    cell.alignment = text_align
                if ci in special_cols:
                    url = special_cols[ci](drow)
                    if url:
                        cell.hyperlink = url
                        cell.font = Font(color="2980B9", underline="single")
        for ci, (title, key) in enumerate(hdef, 1):
            max_w = max(
                len(str(title)),
                *(len(str(d.get(key, "") or "")) for d in data_rows[:200])
            ) if data_rows else len(str(title))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 3, 48)
        ws.freeze_panes = "A2"
        if data_rows:
            ws.auto_filter.ref = ws.dimensions

    # ───── Sheet 5 — Pending Sales Orders
    so_rows_all: list[dict] = []
    if scope_items and (so_st or so_wf):
        # Bulk-fetch customer names for the items' parent SOs
        # First gather voucher rows per item, then enrich with customer
        for item in scope_items:
            uom = uom_map_export.get(item) or {"higher_uom": "", "cf": 1.0, "stock_uom": ""}
            cf = flt(uom["cf"] or 1)
            for d in _drill_so(item, _whs, _co, so_st, so_wf, cf):
                d["item_code"] = item
                d["item_name"] = (item_meta_map.get(item) or {}).get("item_name") or item
                d["stock_uom"] = uom.get("stock_uom") or ""
                d["higher_uom"] = uom.get("higher_uom") or ""
                so_rows_all.append(d)
        if so_rows_all:
            so_names = list({d["voucher_no"] for d in so_rows_all})
            cust_map = {
                r["name"]: (r["customer"] or "", r["customer_name"] or "")
                for r in frappe.db.sql(
                    """
                    SELECT so.name, so.customer, c.customer_name
                      FROM `tabSales Order` so
                 LEFT JOIN `tabCustomer` c ON c.name = so.customer
                     WHERE so.name IN %(n)s
                    """,
                    {"n": tuple(so_names)},
                    as_dict=True,
                )
            }
            for d in so_rows_all:
                cid, cname = cust_map.get(d["voucher_no"]) or ("", "")
                d["customer_id"] = cid
                d["customer"] = f"{cid} : {cname}" if cid else ""

    ws5 = wb.create_sheet("Pending Sales Orders")
    headers_def = [
        ("Item",          "item_code"),
        ("Item Name",     "item_name"),
        ("Voucher",       "voucher_no"),
        ("Date",          "posting_date"),
        ("Customer",      "customer"),
        ("Stock UOM",     "stock_uom"),
        ("Higher UOM",    "higher_uom"),
        ("Planned (Stock UOM)",  "planned_stock"),
        ("Planned (Higher UOM)", "planned_higher"),
        ("Delivered (Stock UOM)",  "actual_stock"),
        ("Delivered (Higher UOM)", "actual_higher"),
        ("Pending (Stock UOM)",  "pending_stock"),
        ("Pending (Higher UOM)", "pending_higher"),
    ]
    _write_voucher_sheet(
        ws5,
        headers_def,
        so_rows_all,
        qty_col_keys={"planned_stock","planned_higher","actual_stock","actual_higher","pending_stock","pending_higher"},
        special_cols={
            1: lambda d: f"{site_url}/app/item/{d.get('item_code','')}",
            3: lambda d: f"{site_url}/app/sales-order/{d.get('voucher_no','')}",
            5: lambda d: f"{site_url}/app/customer/{d.get('customer_id','')}" if d.get("customer_id") else None,
        },
    )

    # ───── Sheet 6 — Pending Work Orders
    wo_rows_all: list[dict] = []
    if scope_items and (wo_st or wo_wf):
        for item in scope_items:
            uom = uom_map_export.get(item) or {"higher_uom": "", "cf": 1.0, "stock_uom": ""}
            cf = flt(uom["cf"] or 1)
            for d in _drill_wo(item, _whs, _co, wo_st, wo_wf, cf):
                d["item_code"] = item
                d["item_name"] = (item_meta_map.get(item) or {}).get("item_name") or item
                d["stock_uom"] = uom.get("stock_uom") or ""
                d["higher_uom"] = uom.get("higher_uom") or ""
                wo_rows_all.append(d)
        if wo_rows_all:
            wo_names = list({d["voucher_no"] for d in wo_rows_all})
            pp_map = {
                r["name"]: r["production_plan"] or ""
                for r in frappe.db.sql(
                    "SELECT name, production_plan FROM `tabWork Order` WHERE name IN %(n)s",
                    {"n": tuple(wo_names)},
                    as_dict=True,
                )
            }
            for d in wo_rows_all:
                d["production_plan"] = pp_map.get(d["voucher_no"]) or ""

    ws6 = wb.create_sheet("Pending Work Orders")
    headers_def = [
        ("Item",                    "item_code"),
        ("Item Name",               "item_name"),
        ("Voucher",                 "voucher_no"),
        ("Date",                    "posting_date"),
        ("Production Plan",         "production_plan"),
        ("Stock UOM",               "stock_uom"),
        ("Higher UOM",              "higher_uom"),
        ("Planned (Stock UOM)",     "planned_stock"),
        ("Planned (Higher UOM)",    "planned_higher"),
        ("Produced (Stock UOM)",    "actual_stock"),
        ("Produced (Higher UOM)",   "actual_higher"),
        ("Pending (Stock UOM)",     "pending_stock"),
        ("Pending (Higher UOM)",    "pending_higher"),
    ]
    _write_voucher_sheet(
        ws6,
        headers_def,
        wo_rows_all,
        qty_col_keys={"planned_stock","planned_higher","actual_stock","actual_higher","pending_stock","pending_higher"},
        special_cols={
            1: lambda d: f"{site_url}/app/item/{d.get('item_code','')}",
            3: lambda d: f"{site_url}/app/work-order/{d.get('voucher_no','')}",
            5: lambda d: f"{site_url}/app/production-plan/{d.get('production_plan','')}" if d.get("production_plan") else None,
        },
    )

    # ───── Sheet 7 — Pending Purchase Orders
    po_rows_all: list[dict] = []
    if scope_items and (po_st or po_wf):
        for item in scope_items:
            uom = uom_map_export.get(item) or {"higher_uom": "", "cf": 1.0, "stock_uom": ""}
            cf = flt(uom["cf"] or 1)
            for d in _drill_po(item, _whs, _co, po_st, po_wf, cf):
                d["item_code"] = item
                d["item_name"] = (item_meta_map.get(item) or {}).get("item_name") or item
                d["stock_uom"] = uom.get("stock_uom") or ""
                d["higher_uom"] = uom.get("higher_uom") or ""
                po_rows_all.append(d)
        if po_rows_all:
            po_names = list({d["voucher_no"] for d in po_rows_all})
            sup_map = {
                r["name"]: (r["supplier"] or "", r["supplier_name"] or "")
                for r in frappe.db.sql(
                    """
                    SELECT po.name, po.supplier, s.supplier_name
                      FROM `tabPurchase Order` po
                 LEFT JOIN `tabSupplier` s ON s.name = po.supplier
                     WHERE po.name IN %(n)s
                    """,
                    {"n": tuple(po_names)},
                    as_dict=True,
                )
            }
            for d in po_rows_all:
                sid, sname = sup_map.get(d["voucher_no"]) or ("", "")
                d["supplier_id"] = sid
                d["supplier"] = f"{sid} : {sname}" if sid else ""

    ws7 = wb.create_sheet("Pending Purchase Orders")
    headers_def = [
        ("Item",                "item_code"),
        ("Item Name",           "item_name"),
        ("Voucher",             "voucher_no"),
        ("Date",                "posting_date"),
        ("Supplier",            "supplier"),
        ("Stock UOM",           "stock_uom"),
        ("Higher UOM",          "higher_uom"),
        ("Planned (Stock UOM)",   "planned_stock"),
        ("Planned (Higher UOM)",  "planned_higher"),
        ("Received (Stock UOM)",  "actual_stock"),
        ("Received (Higher UOM)", "actual_higher"),
        ("Pending (Stock UOM)",   "pending_stock"),
        ("Pending (Higher UOM)",  "pending_higher"),
    ]
    _write_voucher_sheet(
        ws7,
        headers_def,
        po_rows_all,
        qty_col_keys={"planned_stock","planned_higher","actual_stock","actual_higher","pending_stock","pending_higher"},
        special_cols={
            1: lambda d: f"{site_url}/app/item/{d.get('item_code','')}",
            3: lambda d: f"{site_url}/app/purchase-order/{d.get('voucher_no','')}",
            5: lambda d: f"{site_url}/app/supplier/{d.get('supplier_id','')}" if d.get("supplier_id") else None,
        },
    )

    # ───── Sheet 8 — Pending WO Consumption
    woc_rows_all: list[dict] = []
    if scope_items and (wo_st or wo_wf):
        for item in scope_items:
            uom = uom_map_export.get(item) or {"higher_uom": "", "cf": 1.0, "stock_uom": ""}
            cf = flt(uom["cf"] or 1)
            for d in _drill_wo_consume(item, _whs, _co, wo_st, wo_wf, cf):
                d["item_code"] = item
                d["item_name"] = (item_meta_map.get(item) or {}).get("item_name") or item
                d["stock_uom"] = uom.get("stock_uom") or ""
                d["higher_uom"] = uom.get("higher_uom") or ""
                woc_rows_all.append(d)

    ws8 = wb.create_sheet("Pending WO Consumption")
    headers_def = [
        ("Required Item",          "item_code"),
        ("Required Item Name",     "item_name"),
        ("Work Order",             "voucher_no"),
        ("Date",                   "posting_date"),
        ("For FG Item",            "for_fg"),
        ("Stock UOM",              "stock_uom"),
        ("Higher UOM",             "higher_uom"),
        ("Required (Stock UOM)",      "planned_stock"),
        ("Required (Higher UOM)",     "planned_higher"),
        ("Transferred (Stock UOM)",   "actual_stock"),
        ("Transferred (Higher UOM)",  "actual_higher"),
        ("Pending (Stock UOM)",       "pending_stock"),
        ("Pending (Higher UOM)",      "pending_higher"),
    ]
    _write_voucher_sheet(
        ws8,
        headers_def,
        woc_rows_all,
        qty_col_keys={"planned_stock","planned_higher","actual_stock","actual_higher","pending_stock","pending_higher"},
        special_cols={
            1: lambda d: f"{site_url}/app/item/{d.get('item_code','')}",
            3: lambda d: f"{site_url}/app/work-order/{d.get('voucher_no','')}",
            5: lambda d: f"{site_url}/app/item/{d.get('for_fg','')}" if d.get("for_fg") else None,
        },
    )

    # ── Stream
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    frappe.response.filename = f"item_short_surplus_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"
    frappe.response.filecontent = buf.getvalue()
    frappe.response.type = "binary"
    frappe.response.display_content_as = "attachment"
