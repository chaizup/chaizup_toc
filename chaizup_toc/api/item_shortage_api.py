# =============================================================================
# CONTEXT: Item Shortage Dashboard — Page-facing whitelisted API.
#   Single surface: Frappe Page at
#   `chaizup_toc/chaizup_toc/page/item_shortage_dashboard/`. The Script
#   Report surface was removed 2026-05-14 (ISD-008); compute now lives in
#   `chaizup_toc.api.item_shortage_compute` and this file is the
#   page-friendly contract — flat JSON-serialisable payloads that the
#   Tabulator grid + drill-down modal consume directly.
#
# MEMORY: chaizup_item_shortage_dashboard.md § Page API
#
# INSTRUCTIONS:
#   - `get_dashboard_data(filters)` returns:
#         {
#           "columns": [...],         # same as Script Report columns
#           "rows":    [...],         # report rows
#           "summary": [...],         # summary cards
#           "chart":   {...},         # chart payload
#           "banner":  "<HTML…>",     # pending-status banner
#           "filters_used": {...},    # echo so client can render chips
#         }
#     `filters` is a dict-or-JSON-string, identical shape to the Report filters.
#   - `get_breakdown(column, item_code, warehouse, month, year)` delegates
#     to the report's `get_cell_breakdown` — kept here so the page never
#     reaches into a `report.` import path (cleaner separation).
#   - `get_filter_options()` returns lookup lists used by the multi-select
#     filters (item groups, warehouses, recent items) so the page does
#     not need to call ERPNext native APIs from JS.
#
# DANGER ZONE:
#   - Do NOT duplicate the column/row helpers here. ANY divergence between
#     the Page and Report will surface as "the cell shows X but the report
#     shows Y" bugs. Always import from `report.item_shortage_dashboard`.
#   - frappe.parse_json with empty string raises — guard with `or {}`.
#
# RESTRICT:
#   - Keep this file thin. If logic grows, factor into the report module,
#     not here. The shim is a routing layer, not a compute layer.
# =============================================================================

import json
import frappe
from frappe import _


def _parse_filters(filters):
    """Accept dict OR JSON string OR None and always return a dict."""
    if not filters:
        return {}
    if isinstance(filters, dict):
        return filters
    if isinstance(filters, str):
        try:
            parsed = json.loads(filters)
            return parsed if isinstance(parsed, dict) else {}
        except Exception:
            return {}
    return {}


@frappe.whitelist()
def get_dashboard_data(filters=None):
    """
    Page entry — wraps the Script Report execute() and returns a flat dict.

    INSTRUCTIONS:
      - Single backend call per page refresh; the page builds filter UI from
        `get_filter_options()` on first paint, then this endpoint thereafter.
      - Returned `columns` retain Frappe Script Report shape so the page can
        render the same labels/widths/tooltips without re-defining them.
    """
    from chaizup_toc.api import item_shortage_compute as isd_compute
    f = _parse_filters(filters)
    columns, rows, banner, chart, summary = isd_compute.execute(f)
    return {
        "columns":       columns,
        "rows":          rows or [],
        "banner":        banner or "",
        "chart":         chart,
        "summary":       summary or [],
        "filters_used":  f,
        "row_count":     len(rows or []),
    }


@frappe.whitelist()
def get_breakdown(column, item_code, warehouse=None, month=None, year=None):
    """
    Page-friendly drill-down shim — delegates to the Script Report's
    `get_cell_breakdown` (already whitelisted there). Re-exposed here so the
    page JS calls a single `chaizup_toc.api.item_shortage_api.get_breakdown`
    namespace instead of a report-path call (cleaner permission boundary).
    """
    from chaizup_toc.api import item_shortage_compute as isd_compute
    return isd_compute.get_cell_breakdown(
        column=column,
        item_code=item_code,
        warehouse=warehouse,
        month=month,
        year=year,
    )


@frappe.whitelist()
def get_filter_options():
    """
    Bootstrap lookup payload for the page filters.

    Returns:
      {
        "companies":   [name, ...],
        "warehouses":  [name, ...],   # active, non-group
        "item_groups": [name, ...],
        "uoms":        [name, ...],   # for future tooltip enrichment
        "default_company": user_default_company or first one,
      }

    INSTRUCTIONS:
      - List sizes are intentionally not paginated; multi-select UI in the
        page uses substring filter on the client. If your site has 5000+
        items, switch the item-name MultiSelectList to a frappe.db.get_link
        lookup like Production Overview does.
      - All four lookups are read in a single round-trip — keeps page
        first-paint snappy.
    """
    companies = frappe.get_all("Company", pluck="name", order_by="name")
    warehouses = frappe.get_all(
        "Warehouse",
        filters={"disabled": 0, "is_group": 0},
        pluck="name",
        order_by="name",
    )
    item_groups = frappe.get_all(
        "Item Group", pluck="name", order_by="name",
    )
    default_company = frappe.defaults.get_user_default("Company") or (
        companies[0] if companies else ""
    )
    # Pending status payload — single source of truth (TS-001).
    try:
        from chaizup_toc.api.wo_kitting_api import get_toc_pending_filters
        pending = get_toc_pending_filters()
    except Exception:
        pending = {"wo": [], "so": [], "po": [], "edit_route": "/app/toc-settings"}
    return {
        "companies":       companies,
        "warehouses":      warehouses,
        "item_groups":     item_groups,
        "default_company": default_company,
        "pending":         pending,
    }


# =============================================================================
# CONTEXT: export_xlsx — creative branded multi-sheet workbook.
#   Replaces Tabulator's default XLSX export (flat single-sheet dump) with
#   a server-side openpyxl workbook that mirrors the production_overview
#   Excel pattern: Cover + Items + Shortage Drivers + Pending SO Items
#   + filter / banner context. Conditional formatting on the shortage
#   columns + frozen header + auto filter + alt-row tint.
#
# MEMORY: chaizup_item_shortage_dashboard.md § XLSX Export | ISD-015
#
# INSTRUCTIONS:
#   - Returns the workbook bytes via frappe.local.response so the browser
#     downloads it directly (no extra round-trip / no File doctype).
#   - Filters mirror get_dashboard_data so the user gets exactly what
#     they see on screen.
#   - The "Cover" sheet is always sheet #1 — KPIs + the active pending
#     status lists from TOC Settings + the universal-search / month-year
#     context. Gives the reader a self-contained snapshot.
#
# DANGER ZONE:
#   - frappe.local.response["type"] MUST be set to "binary" AND
#     ["filecontent"] to the bytes AND ["filename"] to a string ending
#     in .xlsx — any one missing and Frappe responds with JSON.
#   - openpyxl Worksheet.column_dimensions[col_letter].width is the only
#     width property that survives a save round-trip. Setting widths on
#     the cell object is silently dropped.
#   - Conditional formatting rules attach to RANGES, not cells. Always
#     pass an explicit "A2:A99" style string.
#
# RESTRICT:
#   - Do NOT round numeric values before writing — Excel handles
#     precision better with raw floats + number_format on the cell.
#   - Do NOT bundle the inline UOM ladder text in numeric cells — write
#     the stock-UOM number and put alt UOMs in adjacent columns. The
#     reader expects numbers to remain sortable / pivotable.
# =============================================================================
@frappe.whitelist()
def export_xlsx(filters=None):
    """Generate a creative branded XLSX workbook of the dashboard.

    Returns: streams the workbook as the HTTP response body. Frontend
    triggers it via window.open() / form submission to the URL
    `/api/method/chaizup_toc.api.item_shortage_api.export_xlsx`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, NamedStyle,
    )
    from openpyxl.formatting.rule import (
        ColorScaleRule, CellIsRule, DataBarRule,
    )
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    data = get_dashboard_data(filters)
    rows    = data.get("rows") or []
    cols    = [c for c in (data.get("columns") or []) if not c.get("hidden")]
    summary = data.get("summary") or []
    pending = (data.get("filters_used") or {})
    fopts = get_filter_options() or {}
    pending_status = (fopts.get("pending") or {})

    wb = Workbook()
    # ── Style primitives ────────────────────────────────────────────────
    fill_header_brand = PatternFill("solid", fgColor="1F2937")
    fill_header_accent = PatternFill("solid", fgColor="4F46E5")
    fill_subheader   = PatternFill("solid", fgColor="334155")
    fill_zebra       = PatternFill("solid", fgColor="F8FAFC")
    fill_shortage    = PatternFill("solid", fgColor="FEE2E2")
    fill_warn        = PatternFill("solid", fgColor="FEF3C7")
    fill_ok          = PatternFill("solid", fgColor="D1FAE5")
    fill_cover_title = PatternFill("solid", fgColor="0F172A")
    font_h1     = Font(name="Calibri", size=20, bold=True, color="FFFFFF")
    font_h2     = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    font_label  = Font(name="Calibri", size=10, bold=True, color="475569")
    font_value  = Font(name="Calibri", size=11, color="0F172A")
    font_th     = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_normal = Font(name="Calibri", size=10, color="0F172A")
    thin = Side(border_style="thin", color="E2E8F0")
    border_all = Border(top=thin, bottom=thin, left=thin, right=thin)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    align_right  = Alignment(horizontal="right",  vertical="center")
    align_wrap   = Alignment(horizontal="left",   vertical="top",    wrap_text=True)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1 — COVER
    # ═══════════════════════════════════════════════════════════════════════
    cover = wb.active
    cover.title = "Cover"
    cover.sheet_view.showGridLines = False
    # Wide first columns for banner layout
    cover.column_dimensions["A"].width = 4
    for col, w in zip("BCDEFGHIJ", [22, 22, 22, 22, 22, 22, 22, 22, 4]):
        cover.column_dimensions[col].width = w

    cover.merge_cells("B2:I3")
    cover["B2"] = "Item Shortage Dashboard"
    cover["B2"].font = font_h1
    cover["B2"].fill = fill_cover_title
    cover["B2"].alignment = align_center

    cover.merge_cells("B4:I4")
    cover["B4"] = (
        f"Snapshot generated {frappe.utils.now_datetime():%Y-%m-%d %H:%M} "
        f"by {frappe.session.user}"
    )
    cover["B4"].font = Font(name="Calibri", size=10, italic=True, color="64748B")
    cover["B4"].alignment = align_center

    # KPIs as 4-card row
    cover.merge_cells("B6:I6")
    cover["B6"] = "At-a-glance KPIs"
    cover["B6"].font = font_h2
    cover["B6"].fill = fill_subheader
    cover["B6"].alignment = align_left

    kpi_row = 7
    kpi_cols = ["B", "D", "F", "H"]
    for i, s in enumerate(summary[:4]):
        if i >= len(kpi_cols):
            break
        col1 = kpi_cols[i]
        col2 = chr(ord(col1) + 1)
        cover.merge_cells(f"{col1}{kpi_row}:{col2}{kpi_row}")
        cell = cover[f"{col1}{kpi_row}"]
        cell.value = str(s.get("label", ""))
        cell.font = font_label
        cell.fill = fill_zebra
        cell.alignment = align_center
        cell.border = border_all
        cover.merge_cells(f"{col1}{kpi_row+1}:{col2}{kpi_row+1}")
        cv = cover[f"{col1}{kpi_row+1}"]
        cv.value = s.get("value", 0)
        cv.font = Font(name="Calibri", size=18, bold=True,
                       color={"red":"DC2626","orange":"D97706","yellow":"CA8A04",
                              "green":"059669","blue":"4F46E5"}.get(
                                  (s.get("indicator") or "").lower(), "0F172A"))
        cv.alignment = align_center
        cv.border = border_all
    cover.row_dimensions[kpi_row].height = 28
    cover.row_dimensions[kpi_row + 1].height = 40

    # Pending status banner
    cover.merge_cells("B10:I10")
    cover["B10"] = "Pending statuses (single source of truth — TOC Settings)"
    cover["B10"].font = font_h2
    cover["B10"].fill = fill_subheader
    cover["B10"].alignment = align_left
    banner_rows = [
        ("Sales Order",    ", ".join(pending_status.get("so") or []) or "—"),
        ("Work Order",     ", ".join(pending_status.get("wo") or []) or "—"),
        ("Purchase Order", ", ".join(pending_status.get("po") or []) or "—"),
    ]
    r = 11
    for lbl, val in banner_rows:
        cover.merge_cells(f"B{r}:C{r}")
        cover[f"B{r}"] = lbl
        cover[f"B{r}"].font = font_label
        cover[f"B{r}"].alignment = align_left
        cover[f"B{r}"].fill = fill_zebra
        cover.merge_cells(f"D{r}:I{r}")
        cover[f"D{r}"] = val
        cover[f"D{r}"].font = font_value
        cover[f"D{r}"].alignment = align_wrap
        cover.row_dimensions[r].height = 22
        r += 1

    # Filters used
    cover.merge_cells(f"B{r+1}:I{r+1}")
    cover[f"B{r+1}"] = "Filters applied"
    cover[f"B{r+1}"].font = font_h2
    cover[f"B{r+1}"].fill = fill_subheader
    cover[f"B{r+1}"].alignment = align_left
    r += 2
    filter_pairs = [
        ("Company",          pending.get("company") or "(all)"),
        ("Item Group",       ", ".join(pending.get("item_group") or []) or "(all)"),
        ("Item Name",        ", ".join(pending.get("item_name") or []) or "(all)"),
        ("Warehouse",        ", ".join(pending.get("warehouse") or []) or "(all)"),
        ("Will Consumed In", f"{pending.get('month') or ''} {pending.get('year') or ''}".strip() or "(default)"),
        ("Universal Search", pending.get("universal_search") or "(none)"),
    ]
    for lbl, val in filter_pairs:
        cover.merge_cells(f"B{r}:C{r}")
        cover[f"B{r}"] = lbl
        cover[f"B{r}"].font = font_label
        cover[f"B{r}"].fill = fill_zebra
        cover[f"B{r}"].alignment = align_left
        cover.merge_cells(f"D{r}:I{r}")
        cover[f"D{r}"] = val
        cover[f"D{r}"].font = font_value
        cover[f"D{r}"].alignment = align_wrap
        r += 1

    # Sheet guide
    r += 1
    cover.merge_cells(f"B{r}:I{r}")
    cover[f"B{r}"] = "Sheet guide"
    cover[f"B{r}"].font = font_h2
    cover[f"B{r}"].fill = fill_subheader
    cover[f"B{r}"].alignment = align_left
    r += 1
    sheet_guide = [
        ("Cover", "KPIs + active filters + status snapshot."),
        ("Items", "Full grid — every column in the dashboard, with auto-filter and frozen header. Conditional fill on shortage column."),
        ("Shortage Drivers", "Subset of Items where Total Shortage incl. Expected (o) > 0. Sorted by shortage desc."),
        ("Pending SO Items", "Subset of Items with any pending Sales Order. Highlights customer-driven demand."),
        ("UOM Conversions", "Item × UOM × Factor matrix sourced from each row's UOM Conversion ladder. Lets you cross-check unit math."),
    ]
    for lbl, val in sheet_guide:
        cover.merge_cells(f"B{r}:C{r}")
        cover[f"B{r}"] = lbl
        cover[f"B{r}"].font = font_label
        cover[f"B{r}"].fill = fill_zebra
        cover[f"B{r}"].alignment = align_left
        cover.merge_cells(f"D{r}:I{r}")
        cover[f"D{r}"] = val
        cover[f"D{r}"].font = font_value
        cover[f"D{r}"].alignment = align_wrap
        cover.row_dimensions[r].height = 32
        r += 1

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2 — ITEMS  (full grid)
    # ═══════════════════════════════════════════════════════════════════════
    def _write_data_sheet(ws, rows_in, title_text):
        ws.sheet_view.showGridLines = False
        # Title banner
        if cols:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        ws.cell(row=1, column=1, value=title_text)
        ws.cell(row=1, column=1).font = font_h1
        ws.cell(row=1, column=1).fill = fill_header_brand
        ws.cell(row=1, column=1).alignment = align_center
        ws.row_dimensions[1].height = 36
        # Header row
        for idx, c in enumerate(cols, start=1):
            cell = ws.cell(row=2, column=idx, value=c.get("label") or c.get("fieldname"))
            cell.font = font_th
            cell.fill = fill_header_accent
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            # Decent default width by fieldtype
            ftype = c.get("fieldtype")
            width = 14
            if ftype in ("Float", "Currency"):    width = 16
            if ftype == "Percent":                 width = 12
            if ftype == "Int":                     width = 10
            if c.get("fieldname") == "item_name":  width = 36
            if c.get("fieldname") in ("item_code", "warehouse", "item_group"):
                width = 20
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.row_dimensions[2].height = 38
        # Data rows
        numeric_fmt = '#,##0.00'
        int_fmt     = '#,##0'
        pct_fmt     = '0.0%'
        for ridx, row in enumerate(rows_in, start=3):
            zebra = (ridx % 2 == 0)
            is_short = (row.get("total_shortage_with_expected") or 0) > 0
            is_so    = bool(row.get("_has_pending_so"))
            for cidx, c in enumerate(cols, start=1):
                v = row.get(c.get("fieldname"))
                cell = ws.cell(row=ridx, column=cidx, value=v)
                cell.font = font_normal
                cell.border = border_all
                ftype = c.get("fieldtype")
                if ftype in ("Float", "Currency"):
                    cell.alignment = align_right
                    cell.number_format = numeric_fmt
                elif ftype == "Int":
                    cell.alignment = align_right
                    cell.number_format = int_fmt
                elif ftype == "Percent":
                    cell.alignment = align_right
                    # Percent values come as 0..100; Excel format wants 0..1.
                    if isinstance(v, (int, float)):
                        cell.value = float(v) / 100.0
                    cell.number_format = pct_fmt
                else:
                    cell.alignment = align_left
                # Highlight precedence: pending SO row > shortage > zebra
                if is_so:
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")
                elif is_short and c.get("fieldname") in (
                    "total_shortage_with_expected",
                    "total_shortage_stock_only",
                    "need_as_per_max_level",
                    "decision_qty",
                ):
                    cell.fill = fill_shortage
                elif zebra:
                    cell.fill = fill_zebra
        # Freeze + autofilter
        ws.freeze_panes = "B3"
        if cols:
            last_col = get_column_letter(len(cols))
            last_row = max(2, len(rows_in) + 2)
            ws.auto_filter.ref = f"A2:{last_col}{last_row}"
        # Conditional data bars on the shortage columns
        target_fields = {
            "total_shortage_with_expected": "FCA5A5",
            "total_shortage_stock_only":    "FCD34D",
            "decision_qty":                 "A78BFA",
        }
        last_row = max(3, len(rows_in) + 2)
        for cidx, c in enumerate(cols, start=1):
            colour = target_fields.get(c.get("fieldname"))
            if not colour:
                continue
            col_letter = get_column_letter(cidx)
            rng = f"{col_letter}3:{col_letter}{last_row}"
            try:
                ws.conditional_formatting.add(rng, DataBarRule(
                    start_type="min", end_type="max", color=colour, showValue=True,
                ))
            except Exception:
                pass

    ws_items = wb.create_sheet("Items")
    _write_data_sheet(ws_items, rows, "Items — Full Grid")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3 — Shortage Drivers
    # ═══════════════════════════════════════════════════════════════════════
    short_rows = [r for r in rows if (r.get("total_shortage_with_expected") or 0) > 0]
    short_rows.sort(key=lambda x: -(x.get("total_shortage_with_expected") or 0))
    ws_short = wb.create_sheet("Shortage Drivers")
    _write_data_sheet(ws_short, short_rows, "Shortage Drivers (o > 0)")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4 — Pending SO Items
    # ═══════════════════════════════════════════════════════════════════════
    so_rows = [r for r in rows if r.get("_has_pending_so")]
    so_rows.sort(key=lambda x: -(x.get("will_dispatch_pending_so") or 0))
    ws_so = wb.create_sheet("Pending SO Items")
    _write_data_sheet(ws_so, so_rows, "Pending SO Items")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 5 — UOM Conversions matrix
    # ═══════════════════════════════════════════════════════════════════════
    ws_uom = wb.create_sheet("UOM Conversions")
    ws_uom.sheet_view.showGridLines = False
    headers = ["Item Code", "Item Name", "Stock UOM", "Alt UOM", "Conversion Factor",
               "Current Stock (Stock UOM)", "Current Stock (Alt UOM)"]
    ws_uom.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws_uom.cell(row=1, column=1, value="Item × UOM Conversion Ladder")
    ws_uom.cell(row=1, column=1).font = font_h1
    ws_uom.cell(row=1, column=1).fill = fill_header_brand
    ws_uom.cell(row=1, column=1).alignment = align_center
    ws_uom.row_dimensions[1].height = 36
    for i, h in enumerate(headers, start=1):
        cell = ws_uom.cell(row=2, column=i, value=h)
        cell.font = font_th
        cell.fill = fill_header_accent
        cell.alignment = align_center
        cell.border = border_all
    for col_letter, w in zip("ABCDEFG", [18, 36, 14, 14, 18, 22, 22]):
        ws_uom.column_dimensions[col_letter].width = w
    ws_uom.row_dimensions[2].height = 32
    r = 3
    for row in rows:
        try:
            conv_list = frappe.parse_json(row.get("_uom_conversions") or "[]") or []
        except Exception:
            conv_list = []
        stock_qty = row.get("current_stock") or 0
        for c in conv_list:
            if not c or not c.get("factor"):
                continue
            ws_uom.cell(row=r, column=1, value=row.get("item_code"))
            ws_uom.cell(row=r, column=2, value=row.get("item_name"))
            ws_uom.cell(row=r, column=3, value=row.get("stock_uom"))
            ws_uom.cell(row=r, column=4, value=c.get("uom"))
            ws_uom.cell(row=r, column=5, value=c.get("factor"))
            ws_uom.cell(row=r, column=6, value=stock_qty)
            try:
                in_alt = float(stock_qty) / float(c.get("factor")) if c.get("factor") else 0
            except Exception:
                in_alt = 0
            ws_uom.cell(row=r, column=7, value=in_alt)
            for col in range(1, len(headers) + 1):
                cell = ws_uom.cell(row=r, column=col)
                cell.font = font_normal
                cell.border = border_all
                if col in (5, 6, 7):
                    cell.alignment = align_right
                    cell.number_format = '#,##0.0000'
                else:
                    cell.alignment = align_left
                if r % 2 == 0:
                    cell.fill = fill_zebra
            r += 1
    ws_uom.freeze_panes = "A3"
    ws_uom.auto_filter.ref = f"A2:G{max(2, r - 1)}"

    # ── Stream the workbook ─────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    filename = f"Item-Shortage-Dashboard-{frappe.utils.now()[:10]}.xlsx"
    frappe.local.response.update({
        "type":        "binary",
        "filename":    filename,
        "filecontent": buf.getvalue(),
    })


@frappe.whitelist()
def send_email_snapshot(recipients, subject, message=None, filters=None, cc=None):
    """
    Send the current dashboard snapshot as an HTML email.

    INSTRUCTIONS:
      - Backend builds the HTML server-side rather than trusting client to
        compose it — avoids markup injection and lets us render a richer
        table (server has all 27 columns, page might be virtualized).
      - Capped at 500 rows in the inline body; full export uses CSV/XLSX.
    """
    if not recipients:
        frappe.throw(_("Recipients required"))
    data = get_dashboard_data(filters)
    rows = (data.get("rows") or [])[:500]
    cols = [c for c in (data.get("columns") or []) if not c.get("hidden")]

    style_th = (
        "padding:6px;text-align:left;background:#1F2937;color:#FFF;"
        "border:1px solid #334155;font-size:11px"
    )
    style_td_num = (
        "padding:4px 6px;text-align:right;border:1px solid #E2E8F0;"
        "font-size:11px;font-family:'Geist Mono',monospace"
    )
    style_td_txt = (
        "padding:4px 6px;text-align:left;border:1px solid #E2E8F0;font-size:11px"
    )
    html = f"<div>{message or ''}</div>"
    html += (
        data.get("banner") or ""
    )
    html += (
        "<table style='border-collapse:collapse;font-family:Arial,sans-serif;"
        "margin-top:12px;width:100%'>"
        "<thead><tr>"
    )
    for c in cols:
        html += f"<th style=\"{style_th}\">{frappe.utils.escape_html(c.get('label') or c.get('fieldname'))}</th>"
    html += "</tr></thead><tbody>"
    numeric_types = {"Float", "Int", "Percent", "Currency"}
    for r in rows:
        # Light-red row for items with pending SO.
        row_style = ""
        if r.get("_has_pending_so"):
            row_style = " style='background:#FEE2E2'"
        html += f"<tr{row_style}>"
        for c in cols:
            v = r.get(c.get("fieldname"))
            if v is None:
                v = ""
            cell_style = style_td_num if c.get("fieldtype") in numeric_types else style_td_txt
            html += f"<td style=\"{cell_style}\">{frappe.utils.escape_html(str(v))}</td>"
        html += "</tr>"
    html += "</tbody></table>"
    if data.get("row_count", 0) > 500:
        html += (
            "<p style='font-size:11px;color:#64748B;margin-top:8px'>"
            f"Showing first 500 of {data['row_count']} rows. "
            "Use CSV / XLSX export for the full dataset.</p>"
        )

    frappe.sendmail(
        recipients=[r.strip() for r in str(recipients).split(",") if r.strip()],
        cc=[c.strip() for c in str(cc or "").split(",") if c.strip()] or None,
        subject=subject or "Item Shortage Dashboard",
        message=html,
        now=False,
    )
    return {"queued": True, "row_count": len(rows)}
