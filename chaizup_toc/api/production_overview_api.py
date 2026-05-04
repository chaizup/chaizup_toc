# =============================================================================
# CONTEXT: Production Overview Page — full backend API.
#   Serves the 3-tab Production Overview page:
#   Tab 1 → item-level production metrics (plan vs actual, orders, dispatch,
#            sales projection, shortage check, possible qty, cost breakup).
#   Tab 2 → DeepSeek AI advisor (same key hierarchy as wo_kitting_api).
#   Tab 3 → Chart data (aggregated summaries for pie + bar charts).
#
# MEMORY: app_chaizup_toc.md § Production Overview Page
#
# INSTRUCTIONS:
#   - Items included: auto_manufacture=1 OR in open WO OR in open SO OR in
#     current Sales Projection OR appearing as component in active BOM.
#   - All quantities returned in stock_uom. UOM conversions returned alongside.
#   - Month/Year identify current period. Prev month = month-1 (handles Jan→Dec).
#   - Projection month uses string names ("April") matching Sales Projection.projection_month.
#   - "Sub Assembly" = appears as component in any active BOM (tabBOM Item).
#   - Cost breakup = BOM standard vs actual STE consumed vs 6-month historical avg.
#   - AI: shares DEEPSEEK_API_KEY + helper functions from wo_kitting_api.
#   - Session Redis key prefix: "por:chat:" (distinct from wkp: prefix).
#
# DANGER ZONE:
#   - Joining tabBOM Item can return many rows — always filter is_active=1 BOMs.
#   - Sales Projected Items uses field "item" (Link→Item), NOT "item_code".
#   - projection_month is a string ("April"), not an integer.
#   - Do NOT call _execute_chat_with_tools with "rows" key in system message.
#   - Warehouse list can be empty = all warehouses; never crash on None.
#   - tabWork Order column for "Qty To Manufacture" is `qty` (label only is the
#     long form). Using `qty_to_manufacture` in SQL → OperationalError 1054.
#     Convention here: `SELECT qty AS qty_to_manufacture` so Python attribute
#     access stays descriptive while SQL stays correct.
#     (Bug 2026-04-30: _get_planned_qty crashed page load.)
#   - Production Plan Item DOES expose `qty_to_manufacture` — ONLY Work Order
#     uses the short `qty` column. Do not assume one schema based on the other.
#
# RESTRICT:
#   - Do NOT bypass frappe.only_for() on any write endpoint.
#   - Do NOT hardcode month names — use _MONTH_NAMES list for indexing.
#   - Do NOT add frappe.db.commit() inside loops — single commit at end.
#   - Do NOT replace `qty AS qty_to_manufacture` aliasing with raw `qty` unless
#     you also rename every `wo.qty_to_manufacture` / `r.qty_to_manufacture`
#     reader and the output dict key in the same change.
# =============================================================================

import json
import frappe
from frappe.utils import cint, flt, getdate, today

# Month name → number and reverse
_MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]
_MONTH_MAP = {name: i + 1 for i, name in enumerate(_MONTH_NAMES)}

# Default pending statuses per DocType
_DEFAULT_WO_STATUSES   = ["Not Started", "In Process", "Material Transferred"]
_DEFAULT_SO_STATUSES   = ["To Deliver and Bill", "To Deliver", "To Bill", "Partially Delivered"]
_DEFAULT_PP_STATUSES   = ["Submitted", "In Process"]
_DEFAULT_PO_STATUSES   = ["To Receive and Bill", "To Receive"]

# ── Shared DeepSeek helpers (imported from existing integration) ──
from chaizup_toc.api.wo_kitting_api import (   # noqa: E402
    _get_api_key,
    _call_deepseek,
    _execute_chat_with_tools,
    DEEPSEEK_MODELS,
    _AI_SESSION_TTL,
    _AI_MAX_HISTORY,
)

_POR_AI_SESSION_PREFIX = "por:chat:"

# =============================================================================
# MAIN DATA API
# =============================================================================

@frappe.whitelist()
def get_production_overview(
    company=None,
    month=None,
    year=None,
    warehouses=None,
    wo_statuses=None,
    so_statuses=None,
    pp_statuses=None,
    po_statuses=None,
    stock_mode="physical",
    planning_mode=0,
):
    """
    Main data for Production Overview Tab 1.

    Returns:
        {
          "items": [ { ...per-item metrics... } ],
          "summary": { total_items, items_with_shortage, ... },
          "period": { month_name, month_num, year, prev_month_name, prev_month_num, prev_year }
        }
    """
    # ── Parse inputs ──
    if isinstance(warehouses, str):
        warehouses = frappe.parse_json(warehouses) or []
    if isinstance(wo_statuses, str):
        wo_statuses = frappe.parse_json(wo_statuses) or _DEFAULT_WO_STATUSES
    if isinstance(so_statuses, str):
        so_statuses = frappe.parse_json(so_statuses) or _DEFAULT_SO_STATUSES
    if isinstance(pp_statuses, str):
        pp_statuses = frappe.parse_json(pp_statuses) or _DEFAULT_PP_STATUSES
    if isinstance(po_statuses, str):
        po_statuses = frappe.parse_json(po_statuses) or _DEFAULT_PO_STATUSES

    wo_statuses  = wo_statuses  or _DEFAULT_WO_STATUSES
    so_statuses  = so_statuses  or _DEFAULT_SO_STATUSES
    pp_statuses  = pp_statuses  or _DEFAULT_PP_STATUSES
    po_statuses  = po_statuses  or _DEFAULT_PO_STATUSES
    warehouses   = warehouses   or []

    today_date   = getdate(today())
    curr_month   = cint(month)  if month  else today_date.month
    curr_year    = cint(year)   if year   else today_date.year

    if curr_month < 1 or curr_month > 12:
        curr_month = today_date.month

    prev_month = curr_month - 1 if curr_month > 1 else 12
    prev_year  = curr_year if curr_month > 1 else curr_year - 1

    curr_month_name = _MONTH_NAMES[curr_month - 1]
    prev_month_name = _MONTH_NAMES[prev_month - 1]

    period = {
        "month_name":      curr_month_name,
        "month_num":       curr_month,
        "year":            curr_year,
        "prev_month_name": prev_month_name,
        "prev_month_num":  prev_month,
        "prev_year":       prev_year,
    }

    company_filter = f"AND i.company = {frappe.db.escape(company)}" if company else ""
    wh_sql, wh_params = _build_wh_filter(warehouses)

    # ═══════════════════════════════════════════════════════
    # STEP 1 — Collect all qualifying item codes
    # ═══════════════════════════════════════════════════════
    qualifying_items = _get_qualifying_items(
        wo_statuses, so_statuses, curr_month, curr_year,
        prev_month, prev_year, company, pp_statuses=pp_statuses,
    )
    if not qualifying_items:
        return {"items": [], "summary": _empty_summary(), "period": period}

    items_sql = _sql_in(qualifying_items)

    # ═══════════════════════════════════════════════════════
    # STEP 2 — Fetch item master details
    # ═══════════════════════════════════════════════════════
    item_rows = frappe.db.sql(f"""
        SELECT
            i.name           AS item_code,
            i.item_name,
            i.item_group,
            i.stock_uom,
            i.description,
            COALESCE(i.custom_toc_auto_manufacture, 0) AS auto_manufacture,
            COALESCE(i.custom_toc_auto_purchase, 0)    AS auto_purchase,
            COALESCE(i.is_purchase_item, 0)            AS is_purchase_item,
            i.standard_rate
        FROM `tabItem` i
        WHERE i.name IN {items_sql}
          AND i.disabled = 0
        ORDER BY i.item_name
    """, as_dict=True)

    if not item_rows:
        return {"items": [], "summary": _empty_summary(), "period": period}

    all_codes = [r.item_code for r in item_rows]
    codes_sql  = _sql_in(all_codes)

    # ═══════════════════════════════════════════════════════
    # STEP 3 — Bulk queries (all return dicts keyed by item_code)
    # ═══════════════════════════════════════════════════════
    uom_map          = _get_uom_conversions(all_codes)
    sub_asm_map      = _get_sub_assembly_info(all_codes, wo_statuses)
    wo_count_map     = _get_wo_counts(all_codes, wo_statuses, warehouses)
    planned_info_map = _get_planned_qty(all_codes, wo_statuses, warehouses)
    actual_qty_map   = _get_actual_qty(all_codes, curr_month, curr_year, warehouses)
    prev_order_map   = _get_so_qty(all_codes, so_statuses, prev_month, prev_year, warehouses)
    curr_order_map   = _get_so_qty(all_codes, so_statuses, curr_month, curr_year, warehouses)
    dispatch_map     = _get_dispatch_qty(all_codes, curr_month, curr_year, warehouses)
    prev_dispatch_map= _get_dispatch_qty(all_codes, prev_month, prev_year, warehouses)
    projection_map   = _get_projection_qty(all_codes, curr_month_name, curr_year, warehouses)
    total_sales_map  = _get_total_sales_qty(all_codes, curr_month, curr_year, warehouses)
    stock_map        = _get_stock(all_codes, warehouses, stock_mode)
    bom_map          = _get_active_bom(all_codes)
    shortage_map     = _get_shortage_summary(
        all_codes, bom_map, stock_map,
        planned_qty_map=planned_info_map,
        warehouses=warehouses, stock_mode=stock_mode
    )
    possible_map     = _get_possible_qty(all_codes, bom_map, stock_map)
    cost_summary_map = _get_cost_summary(all_codes, curr_month, curr_year, bom_map)
    pp_count_map     = _get_active_pp_count(all_codes, pp_statuses)
    wo_pp_group_map  = _get_wo_pp_groups(all_codes, wo_statuses, warehouses)
    has_open_so_map  = _get_has_open_so_map(all_codes, so_statuses, warehouses)
    # Total pending SO (any month) — used by Excel Simple Report sheet + AI ctx.
    total_pending_so_map = _get_all_pending_so_qty(all_codes, so_statuses, warehouses)

    # ═══════════════════════════════════════════════════════
    # STEP 4 — Assemble output rows
    # ═══════════════════════════════════════════════════════
    items_out = []
    for r in item_rows:
        code = r.item_code
        uoms = uom_map.get(code, [])
        pinfo = planned_info_map.get(code, {})
        plan         = flt(pinfo.get("planned", 0))
        pending_wo   = flt(pinfo.get("pending", 0))
        act          = flt(actual_qty_map.get(code, 0))
        prev_o = flt(prev_order_map.get(code, 0))
        curr_o = flt(curr_order_map.get(code, 0))
        disp  = flt(dispatch_map.get(code, 0))
        prev_disp = flt(prev_dispatch_map.get(code, 0))
        proj  = flt(projection_map.get(code, 0))
        total_s = flt(total_sales_map.get(code, 0))
        stock_qty = flt(stock_map.get(code, 0))
        proj_vs = round(total_s / proj, 3) if proj > 0 else 0

        # Coverage = how much of the current-month sales is "covered" by the
        # combined demand inputs (Sales Projection + Previous Month carryover).
        coverage_input = proj + prev_o
        coverage_pct   = round((coverage_input / total_s) * 100, 1) if total_s > 0 else 0

        # Target Production — the higher of (Sales Projection) and (Total
        # current month order). Whichever is greater drives what must be made.
        # Per spec: "if the sales projection is greater so the projection
        # will be target, if order is greater then the order qty will be the
        # production target."
        target_production = max(proj, total_s)

        # % Target Achieved — based on the user formula:
        #   Total order pending qty − (total pending qty of work orders + stock)
        # This is the *gap*. Positive = under-supplied, negative = over-supplied.
        # Achievement is the inverse of gap, normalised to target_production:
        #   gap        = max(curr_o - (pending_wo + stock), 0)
        #   achieved   = max(target - gap, 0)
        #   target_achieved_pct = achieved / target * 100
        # Computed warehouse-scoped because curr_o, pending_wo, and stock are
        # already filtered by the user-selected warehouse list.
        gap_qty = max(curr_o - (pending_wo + stock_qty), 0)
        achieved_qty = max(target_production - gap_qty, 0)
        target_achieved_pct = (
            round(achieved_qty / target_production * 100, 1)
            if target_production > 0 else 0
        )

        # Item type still computed and returned (used by AI context + Excel)
        # but the FRONTEND no longer renders the column (per user request).
        item_type = _classify_item_type(r, sub_asm_map.get(code, {}).get("is_sub_assembly", False))

        items_out.append({
            "item_code":         code,
            "item_name":         r.item_name,
            "item_group":        r.item_group,
            "stock_uom":         r.stock_uom,
            "item_type":         item_type,
            "is_sub_assembly":   sub_asm_map.get(code, {}).get("is_sub_assembly", False),
            "sub_assembly_wos":  sub_asm_map.get(code, {}).get("wo_list", []),
            "open_wo_count":     wo_count_map.get(code, 0),
            "planned_qty":       plan,
            "pending_wo_qty":    pending_wo,
            "actual_qty":        act,
            "prev_month_order":  prev_o,
            "curr_month_order":  curr_o,
            "curr_dispatch":     disp,
            "prev_dispatch":     prev_disp,
            "curr_projection":   proj,
            "total_curr_sales":  total_s,
            "projection_vs_sales": proj_vs,
            "coverage_pct":      coverage_pct,
            "coverage_input":    coverage_input,
            "target_production": target_production,
            "target_gap":        gap_qty,
            "target_achieved_qty": achieved_qty,
            "target_achieved_pct": target_achieved_pct,
            "stock":             stock_qty,
            "has_shortage":      shortage_map.get(code, {}).get("has_shortage", False),
            "shortage_components": shortage_map.get(code, {}).get("short_components", []),
            "possible_qty":      flt(possible_map.get(code, 0)),
            "uom_conversions":   uoms,
            "active_bom":        bom_map.get(code, ""),
            "cost_summary":      cost_summary_map.get(code, {}),
            "pp_count":          pp_count_map.get(code, 0),
            # Smart PP indicator (POR-007): see _get_wo_pp_groups docstring
            "wo_pp_count":       wo_pp_group_map.get(code, {}).get("wo_pp_count", 0),
            "wo_pp_names":       wo_pp_group_map.get(code, {}).get("wo_pp_names", []),
            "wo_pp_siblings":    wo_pp_group_map.get(code, {}).get("wo_pp_siblings", 0),
            "has_open_so":       has_open_so_map.get(code, False),
            "total_pending_so":  flt(total_pending_so_map.get(code, 0)),
        })

    # ═══════════════════════════════════════════════════════
    # STEP 5 — Summary stats
    # ═══════════════════════════════════════════════════════
    # Summary metrics — COUNTS only on the dashboard cards.
    # Reason: items use heterogeneous UOMs (Pcs, Kg, Gram, Master Carton...);
    # summing `qty` across UOMs would be mathematically meaningless.
    # Total qty per UOM is exposed only on the Excel "UOM Comparison" sheet.
    items_with_open_wos     = sum(1 for x in items_out if x.get("open_wo_count", 0) > 0)
    items_with_active_pp    = sum(1 for x in items_out if x.get("pp_count", 0) > 0)
    items_with_curr_so      = sum(1 for x in items_out if x.get("curr_month_order", 0) > 0)
    items_dispatched        = sum(1 for x in items_out if x.get("curr_dispatch", 0) > 0)
    items_with_projection   = sum(1 for x in items_out if x.get("curr_projection", 0) > 0)
    items_target_hit        = sum(1 for x in items_out if x.get("target_achieved_pct", 0) >= 100)
    items_blocked           = sum(
        1 for x in items_out
        if x.get("has_shortage") and x.get("possible_qty", 0) == 0
    )
    items_sub_assembly      = sum(1 for x in items_out if x.get("is_sub_assembly"))

    summary = {
        # === COUNTS — used on dashboard summary cards ============
        "total_items":               len(items_out),
        "items_with_shortage":       sum(1 for x in items_out if x["has_shortage"]),
        "items_no_shortage":         sum(1 for x in items_out if not x["has_shortage"]),
        "items_with_open_wos":       items_with_open_wos,
        "items_with_active_pp":      items_with_active_pp,
        "items_with_curr_so":        items_with_curr_so,
        "items_dispatched":          items_dispatched,
        "items_with_projection":     items_with_projection,
        "items_target_hit":          items_target_hit,
        "items_blocked":             items_blocked,
        "items_sub_assembly":        items_sub_assembly,
        # === Aggregate qty — for Excel summary sheet only ========
        # Labelled in the Excel render with an explicit warning that mixed
        # UOMs make these numbers context-dependent.
        "total_planned_qty":         sum(x["planned_qty"]       for x in items_out),
        "total_actual_qty":          sum(x["actual_qty"]        for x in items_out),
        "total_curr_orders":         sum(x["curr_month_order"]  for x in items_out),
        "total_dispatch":            sum(x["curr_dispatch"]     for x in items_out),
        "total_projection":          sum(x["curr_projection"]   for x in items_out),
    }

    return {"items": items_out, "summary": summary, "period": period}


# =============================================================================
# ITEM DETAIL MODAL
# =============================================================================

@frappe.whitelist()
def get_item_detail(item_code, company=None, month=None, year=None,
                    warehouses=None, stock_mode="physical", wo_statuses=None):
    """
    Full detail for the item click modal.
    Returns: active WOs with sub-assemblies, component shortage, material batch consumption.
    """
    if isinstance(warehouses, str):
        warehouses = frappe.parse_json(warehouses) or []
    if isinstance(wo_statuses, str):
        wo_statuses = frappe.parse_json(wo_statuses) or _DEFAULT_WO_STATUSES

    today_date = getdate(today())
    curr_month = cint(month) if month else today_date.month
    curr_year  = cint(year)  if year  else today_date.year

    # Active work orders for this item
    # NOTE: ERPNext Work Order column is `qty` (label "Qty To Manufacture").
    # We alias to `qty_to_manufacture` so downstream Python keeps the descriptive name.
    wo_rows = frappe.db.sql("""
        SELECT wo.name, wo.status, wo.qty AS qty_to_manufacture, wo.produced_qty,
               wo.bom_no, wo.production_item, wo.planned_start_date, wo.planned_end_date,
               wo.production_plan, wo.company
        FROM `tabWork Order` wo
        WHERE wo.production_item = %s
          AND wo.docstatus = 1
        ORDER BY wo.planned_start_date ASC
        LIMIT 20
    """, item_code, as_dict=True)

    stock_map = _get_stock([item_code], warehouses, stock_mode)

    wo_details = []
    for wo in wo_rows:
        # BOM components with shortage
        components = _walk_bom_shallow(wo.bom_no, flt(wo.qty_to_manufacture), warehouses, stock_mode)
        # Batch consumption from completed STEs
        batch_data = _get_batch_consumption(wo.name)
        # Sub-assembly WOs if this is linked to a production plan
        sub_asm_wos = _get_sub_assembly_wos(wo.production_plan) if wo.production_plan else []

        wo_details.append({
            "wo_name":           wo.name,
            "status":            wo.status,
            "qty_to_manufacture": flt(wo.qty_to_manufacture),
            "produced_qty":      flt(wo.produced_qty),
            "bom_no":            wo.bom_no,
            "planned_start":     str(wo.planned_start_date or ""),
            "planned_end":       str(wo.planned_end_date or ""),
            "production_plan":   wo.production_plan or "",
            "components":        components,
            "batch_consumption": batch_data,
            "sub_assembly_wos":  sub_asm_wos,
        })

    uom_conversions = _get_uom_conversions([item_code]).get(item_code, [])
    item_doc        = frappe.db.get_value("Item", item_code,
        ["item_name", "item_group", "stock_uom", "description", "standard_rate"],
        as_dict=True) or {}

    return {
        "item_code":      item_code,
        "item_name":      item_doc.get("item_name", ""),
        "item_group":     item_doc.get("item_group", ""),
        "stock_uom":      item_doc.get("stock_uom", ""),
        "description":    item_doc.get("description", ""),
        "current_stock":  flt(stock_map.get(item_code, 0)),
        "uom_conversions":uom_conversions,
        "work_orders":    wo_details,
    }


# =============================================================================
# SHORTAGE DETAIL (View button)
# =============================================================================

@frappe.whitelist()
def get_shortage_detail(item_code, warehouses=None, stock_mode="physical", wo_statuses=None):
    """
    Full shortage breakdown across all work order + sub-assembly combinations.
    Called when user clicks 'View' in the Overall Shortage column.
    """
    if isinstance(warehouses, str):
        warehouses = frappe.parse_json(warehouses) or []
    if isinstance(wo_statuses, str):
        wo_statuses = frappe.parse_json(wo_statuses) or _DEFAULT_WO_STATUSES

    # NOTE: Work Order column is `qty`; aliased to `qty_to_manufacture` for clarity.
    wo_rows = frappe.db.sql("""
        SELECT name, qty AS qty_to_manufacture, bom_no, status, production_plan
        FROM `tabWork Order`
        WHERE production_item = %s AND docstatus = 1
        LIMIT 20
    """, item_code, as_dict=True)

    shortage_by_wo = []
    all_shortage_items = {}  # item_code → aggregated shortage

    for wo in wo_rows:
        components = _walk_bom_deep(wo.bom_no, flt(wo.qty_to_manufacture), warehouses, stock_mode, depth=0)
        short_comps = [c for c in components if c["shortage"] > 0]

        for comp in short_comps:
            c = comp["item_code"]
            if c not in all_shortage_items:
                all_shortage_items[c] = {
                    "item_code":   c,
                    "item_name":   comp["item_name"],
                    "stock_uom":   comp["uom"],
                    "total_required": 0,
                    "in_stock":    comp["in_stock"],
                    "shortage":    0,
                    "wo_count":    0,
                    "open_docs":   comp.get("open_docs", []),
                }
            all_shortage_items[c]["total_required"] += comp["required"]
            all_shortage_items[c]["shortage"] += comp["shortage"]
            all_shortage_items[c]["wo_count"] += 1

        shortage_by_wo.append({
            "wo_name":       wo.name,
            "qty_to_manufacture": flt(wo.qty_to_manufacture),
            "status":        wo.status,
            "short_components": short_comps,
        })

    aggregated = sorted(
        all_shortage_items.values(),
        key=lambda x: x["shortage"] * _get_valuation_rate(x["item_code"]),
        reverse=True
    )

    # Add UOM conversions + open docs per aggregated item
    agg_codes = [x["item_code"] for x in aggregated]
    uom_map   = _get_uom_conversions(agg_codes) if agg_codes else {}
    for item in aggregated:
        item["uom_conversions"] = uom_map.get(item["item_code"], [])
        item["open_docs"]       = _get_open_docs_for_item(item["item_code"])

    return {
        "by_wo":     shortage_by_wo,
        "aggregated":aggregated,
    }


# =============================================================================
# COST BREAKUP (3-way comparison)
# =============================================================================

@frappe.whitelist()
def get_cost_breakup(item_code, company=None, month=None, year=None):
    """
    3-way cost comparison:
      1. BOM Standard Cost  — BOM item qty × valuation_rate
      2. Actual STE Cost    — current month manufacture STEs
      3. Historical Average — last 6 months STE average
    """
    today_date = getdate(today())
    curr_month = cint(month) if month else today_date.month
    curr_year  = cint(year)  if year  else today_date.year

    bom = _get_active_bom([item_code]).get(item_code, "")

    # 1. BOM standard cost
    bom_components = []
    bom_total = 0.0
    if bom:
        bom_items = frappe.db.sql("""
            SELECT bi.item_code, bi.item_name, bi.qty, bi.stock_qty,
                   bi.uom, bi.rate, bi.amount
            FROM `tabBOM Item` bi
            WHERE bi.parent = %s AND bi.parenttype = 'BOM'
            ORDER BY bi.idx
        """, bom, as_dict=True)
        for bi in bom_items:
            bom_components.append({
                "item_code": bi.item_code,
                "item_name": bi.item_name,
                "qty":       flt(bi.stock_qty),
                "uom":       bi.uom,
                "rate":      flt(bi.rate),
                "amount":    flt(bi.amount),
            })
            bom_total += flt(bi.amount)

    # 2. Actual STE cost (current month)
    actual_components = []
    actual_total = 0.0
    actual_qty_produced = 0.0
    ste_rows = frappe.db.sql("""
        SELECT sed.item_code, sed.item_name, SUM(sed.qty) AS total_qty,
               AVG(sed.valuation_rate) AS avg_rate, SUM(sed.basic_amount) AS total_amount,
               sed.uom
        FROM `tabStock Entry Detail` sed
        JOIN `tabStock Entry` se ON se.name = sed.parent
        WHERE se.docstatus = 1
          AND se.stock_entry_type = 'Manufacture'
          AND sed.is_finished_item = 0
          AND MONTH(se.posting_date) = %s
          AND YEAR(se.posting_date) = %s
          AND se.work_order IN (
              SELECT name FROM `tabWork Order`
              WHERE production_item = %s AND docstatus = 1
          )
        GROUP BY sed.item_code, sed.uom
    """, (curr_month, curr_year, item_code), as_dict=True)

    # Get produced qty for per-unit cost calc
    prod_qty_row = frappe.db.sql("""
        SELECT SUM(sed.qty) AS produced
        FROM `tabStock Entry Detail` sed
        JOIN `tabStock Entry` se ON se.name = sed.parent
        WHERE se.docstatus = 1
          AND se.stock_entry_type = 'Manufacture'
          AND sed.is_finished_item = 1
          AND sed.item_code = %s
          AND MONTH(se.posting_date) = %s
          AND YEAR(se.posting_date) = %s
    """, (item_code, curr_month, curr_year))
    actual_qty_produced = flt(prod_qty_row[0][0]) if prod_qty_row else 0

    for row in ste_rows:
        actual_components.append({
            "item_code": row.item_code,
            "item_name": row.item_name,
            "qty":       flt(row.total_qty),
            "uom":       row.uom or "Nos",
            "rate":      flt(row.avg_rate),
            "amount":    flt(row.total_amount),
        })
        actual_total += flt(row.total_amount)

    # 3. Historical average (last 6 months)
    hist_total = 0.0
    hist_runs  = 0
    hist_data = frappe.db.sql("""
        SELECT MONTH(se.posting_date) AS mo, YEAR(se.posting_date) AS yr,
               SUM(sed.basic_amount) AS total_cost
        FROM `tabStock Entry Detail` sed
        JOIN `tabStock Entry` se ON se.name = sed.parent
        WHERE se.docstatus = 1
          AND se.stock_entry_type = 'Manufacture'
          AND sed.is_finished_item = 0
          AND se.posting_date >= DATE_SUB(%s, INTERVAL 6 MONTH)
          AND se.work_order IN (
              SELECT name FROM `tabWork Order`
              WHERE production_item = %s AND docstatus = 1
          )
        GROUP BY MONTH(se.posting_date), YEAR(se.posting_date)
    """, (f"{curr_year}-{curr_month:02d}-01", item_code), as_dict=True)

    for h in hist_data:
        hist_total += flt(h.total_cost)
        hist_runs  += 1
    hist_avg = round(hist_total / hist_runs, 2) if hist_runs else 0

    bom_per_unit   = round(bom_total, 2)
    actual_per_unit = round(actual_total / actual_qty_produced, 2) if actual_qty_produced > 0 else 0
    variance_pct   = round((actual_per_unit - bom_per_unit) / bom_per_unit * 100, 1) if bom_per_unit else 0

    # Per-UOM cost breakdown (BOM std + actual) for every UOM the item supports.
    # Lets the user see "₹/Kg vs ₹/Pcs" without manual math.
    uom_conv = _get_uom_conversions([item_code]).get(item_code, [])
    item_stock_uom = frappe.db.get_value("Item", item_code, "stock_uom") or ""
    cost_per_uom = []
    # Stock UOM row (factor 1)
    cost_per_uom.append({
        "uom":           item_stock_uom,
        "factor":        1.0,
        "bom_std":       round(bom_per_unit, 4),
        "actual":        round(actual_per_unit, 4),
        "variance_pct":  variance_pct,
    })
    for u in uom_conv:
        f = flt(u["factor"])
        if f > 0 and (u.get("uom") or "") and u.get("uom") != item_stock_uom:
            cost_per_uom.append({
                "uom":           u["uom"],
                "factor":        f,
                "bom_std":       round(bom_per_unit * f, 4),
                "actual":        round(actual_per_unit * f, 4),
                "variance_pct":  variance_pct,
            })

    return {
        "item_code":           item_code,
        "stock_uom":           item_stock_uom,
        "bom":                 bom,
        "bom_standard": {
            "components": bom_components,
            "total":      round(bom_total, 2),
            "per_unit":   bom_per_unit,
        },
        "actual_consumed": {
            "components":     actual_components,
            "total":          round(actual_total, 2),
            "qty_produced":   actual_qty_produced,
            "per_unit":       actual_per_unit,
        },
        "historical_avg": {
            "avg_per_run": hist_avg,
            "months_data": hist_runs,
            "period":      "Last 6 months",
        },
        "variance_pct":    variance_pct,
        "variance_amount": round(actual_per_unit - bom_per_unit, 2) if bom_per_unit else 0,
        # NEW — per-UOM display
        "cost_per_uom":    cost_per_uom,
    }


# =============================================================================
# PENDING STATUSES (for filter population)
# =============================================================================

@frappe.whitelist()
def get_default_statuses():
    """
    Return all distinct statuses that actually exist in the user's data plus
    sensible "pending" defaults. Nothing is hardcoded — the universe of
    selectable values comes from `SELECT DISTINCT status FROM tab<DocType>`,
    union the active workflow states (if a Workflow is assigned to that doc).

    This handles two real-world cases the user explicitly called out:
      1. Custom workflow on Sales Order ("Draft + workflow_state = Confirmed")
         is treated as pending — workflow states are surfaced as
         "Workflow: <state>" entries that the JS can pass back to the
         backend. The backend interprets these in `_status_filter_clause`.
      2. Sites with custom statuses not in the ERPNext default list still
         appear in the dropdown.
    """
    def _distinct_status(doctype):
        try:
            rows = frappe.db.sql(
                f"SELECT DISTINCT status FROM `tab{doctype}` WHERE status IS NOT NULL"
            )
            return sorted({r[0] for r in rows if r[0]})
        except Exception:
            return []

    def _distinct_workflow(doctype):
        # Workflow states only exist if a Workflow is assigned to that DocType.
        if not frappe.db.has_column(doctype, "workflow_state"):
            return []
        try:
            rows = frappe.db.sql(
                f"SELECT DISTINCT workflow_state FROM `tab{doctype}` "
                f"WHERE workflow_state IS NOT NULL AND workflow_state != ''"
            )
            return sorted({r[0] for r in rows if r[0]})
        except Exception:
            return []

    all_so_status = _distinct_status("Sales Order")
    so_workflow   = _distinct_workflow("Sales Order")
    all_wo_status = _distinct_status("Work Order")
    all_pp_status = _distinct_status("Production Plan")
    all_po_status = _distinct_status("Purchase Order")
    all_mr_status = _distinct_status("Material Request")

    # "Pending" defaults — intersect existing data with our preferred set so we
    # never offer a tick-box that maps to zero rows.
    def _intersect(actual, preferred):
        actual_set = set(actual)
        return [s for s in preferred if s in actual_set]

    so_pending_default = _intersect(all_so_status, _DEFAULT_SO_STATUSES) \
                          or [s for s in all_so_status
                              if s not in ("Cancelled", "Completed", "Closed", "Draft")]
    wo_pending_default = _intersect(all_wo_status, _DEFAULT_WO_STATUSES) \
                          or [s for s in all_wo_status
                              if s not in ("Cancelled", "Completed", "Stopped", "Draft", "Closed")]
    pp_pending_default = _intersect(all_pp_status, _DEFAULT_PP_STATUSES) \
                          or [s for s in all_pp_status
                              if s not in ("Cancelled", "Completed", "Closed")]
    po_pending_default = _intersect(all_po_status, _DEFAULT_PO_STATUSES) \
                          or [s for s in all_po_status
                              if s not in ("Cancelled", "Completed", "Closed", "Draft")]

    # Surface workflow states as parallel "Workflow: <state>" entries so the
    # user can flag draft-but-confirmed orders as pending without contradicting
    # the docstatus model.
    so_universe = list(all_so_status) + [f"Workflow: {w}" for w in so_workflow]

    return {
        # Defaults JS pre-checks on load
        "wo_statuses":  wo_pending_default,
        "so_statuses":  so_pending_default,
        "pp_statuses":  pp_pending_default,
        "po_statuses":  po_pending_default,
        # Universe of choices the multi-select shows (dynamic, never hardcoded)
        "all_wo_statuses": all_wo_status,
        "all_so_statuses": so_universe,
        "all_pp_statuses": all_pp_status,
        "all_po_statuses": all_po_status,
        "all_mr_statuses": all_mr_status,
        # Diagnostic — JS can show "(includes workflow states)" hint
        "so_has_workflow_column": frappe.db.has_column("Sales Order", "workflow_state"),
        "so_workflow_states":     so_workflow,
    }


def _split_status_and_workflow(status_list):
    """
    Split an incoming status list into (plain_statuses, workflow_states).
    Workflow states arrive as "Workflow: Confirmed" entries from JS.
    """
    plain, wf = [], []
    for s in (status_list or []):
        if isinstance(s, str) and s.startswith("Workflow: "):
            wf.append(s[len("Workflow: "):])
        else:
            plain.append(s)
    return plain, wf


def _so_status_clause(so_statuses):
    """
    Build a SQL fragment that matches:
      (so.docstatus = 1 AND so.status IN (plain_statuses))
      OR (so.docstatus = 0 AND so.workflow_state IN (wf_states))   ← only if column exists
    Returns SQL text — caller embeds inside an existing WHERE.
    """
    plain, wf = _split_status_and_workflow(so_statuses)
    parts = []
    if plain:
        parts.append(f"(so.docstatus = 1 AND so.status IN {_sql_in(plain)})")
    if wf and frappe.db.has_column("Sales Order", "workflow_state"):
        parts.append(f"(so.docstatus = 0 AND so.workflow_state IN {_sql_in(wf)})")
    if not parts:
        # No usable filter — match nothing rather than everything (safe default).
        return "1=0"
    return "(" + " OR ".join(parts) + ")"


# =============================================================================
# EXPORT DATA (for CSV/Excel)
# =============================================================================

@frappe.whitelist()
def get_export_data(company=None, month=None, year=None, warehouses=None,
                    wo_statuses=None, so_statuses=None, stock_mode="physical"):
    """
    Returns flat rows suitable for CSV/Excel export.
    UOM conversions are included as extra columns.
    """
    result = get_production_overview(
        company=company, month=month, year=year,
        warehouses=warehouses, wo_statuses=wo_statuses,
        so_statuses=so_statuses, stock_mode=stock_mode
    )

    rows = []
    for item in result.get("items", []):
        code = item["item_code"]
        row = {
            "Item Code":          code,
            "Item Name":          item["item_name"],
            "Item Group":         item["item_group"],
            "Item Type":          item["item_type"],
            "Stock UOM":          item["stock_uom"],
            "Is Sub Assembly":    "Yes" if item["is_sub_assembly"] else "No",
            "Open WO Count":      item["open_wo_count"],
            "Planned Qty":        item["planned_qty"],
            "Actual Qty":         item["actual_qty"],
            "Prev Month Orders":  item["prev_month_order"],
            "Curr Month Orders":  item["curr_month_order"],
            "Curr Dispatch":      item["curr_dispatch"],
            "Curr Projection":    item["curr_projection"],
            "Total Curr Sales":   item["total_curr_sales"],
            "Projection vs Sales (%)": round(item["projection_vs_sales"] * 100, 1),
            "Stock":              item["stock"],
            "Has Shortage":       "Yes" if item["has_shortage"] else "No",
            "Possible Qty":       item["possible_qty"],
            "BOM Standard Cost":  item.get("cost_summary", {}).get("bom_total", 0),
            "Actual Cost":        item.get("cost_summary", {}).get("actual_total", 0),
            "Cost Variance %":    item.get("cost_summary", {}).get("variance_pct", 0),
        }
        # Append UOM conversions as extra columns
        for uom in item.get("uom_conversions", []):
            col_pfx = uom["uom"]
            factor  = flt(uom["factor"])
            if factor > 0:
                row[f"Planned Qty ({col_pfx})"]   = round(item["planned_qty"] / factor, 3)
                row[f"Actual Qty ({col_pfx})"]    = round(item["actual_qty"] / factor, 3)
                row[f"Curr Orders ({col_pfx})"]   = round(item["curr_month_order"] / factor, 3)
                row[f"Curr Dispatch ({col_pfx})"] = round(item["curr_dispatch"] / factor, 3)
        rows.append(row)

    return {"rows": rows, "period": result.get("period", {})}


# =============================================================================
# EXCEL EXPORT (server-side, multi-sheet, formatted)
# =============================================================================

@frappe.whitelist()
def export_excel(company=None, month=None, year=None, warehouses=None,
                 wo_statuses=None, so_statuses=None, pp_statuses=None,
                 po_statuses=None, stock_mode="physical"):
    """
    Server-side Excel export — six sheets, fully formatted, Google-Sheets-import friendly.

    Sheets:
      1. Cover           - title + period + filters applied + how to read this workbook
      2. Overview        - the main item table with description, all qtys, alt-row
                           shading, autofilter on every column, frozen panes, conditional
                           tinting (red = shortage, pink = open SO)
      3. UOM Comparison  - one row per (item, metric, UOM) so a planner can look up
                           any qty in any UOM. Stock UOM rows are highlighted.
      4. Item Master     - reference catalogue: item_code / name / group / stock_uom /
                           description / standard_rate / valuation_rate / TOC flags /
                           is_purchase_item / is_stock_item / disabled, plus a list of
                           every UOM conversion in `Conversions` column.
      5. Group Pivot     - pivot-style aggregate: per Item Group, count of items with
                           open WO/SO/projection/dispatch/shortage/blocked. Hand-rolled
                           rather than a real pivot so Google Sheets imports cleanly.
      6. Shortage Drivers- procurement priority list: components causing shortages,
                           how many parents they block, total shortage qty + UOM.

    Google Sheets compatibility:
      - openpyxl AutoFilter is preserved on import.
      - Pre-aggregated "pivot-style" sheet replaces a real pivot table (real pivots
        are sometimes flattened during import).
      - Alt-row shading uses static fills (not table styles) so the colours survive.
      - Frozen panes are honoured by Google Sheets.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # IMPORTANT: parse JSON-encoded params BEFORE rendering them on the Cover
    # sheet. The URL form sends `wo_statuses=["In Process","Not Started"]`
    # as a string; iterating it as-is renders character-by-character.
    # `get_production_overview` parses internally too, but we need parsed
    # lists here for the cover-sheet "filters applied" rendering.
    def _as_list(v, default):
        if v is None or v == "":
            return list(default)
        if isinstance(v, str):
            parsed = frappe.parse_json(v)
            if isinstance(parsed, list):
                return parsed
            # Single-string fallthrough — wrap in a 1-elem list
            return [parsed] if parsed not in (None, "") else list(default)
        if isinstance(v, list):
            return list(v)
        return list(default)

    warehouses_list = _as_list(warehouses, [])
    wo_list         = _as_list(wo_statuses, _DEFAULT_WO_STATUSES)
    so_list         = _as_list(so_statuses, _DEFAULT_SO_STATUSES)
    pp_list         = _as_list(pp_statuses, _DEFAULT_PP_STATUSES)
    po_list         = _as_list(po_statuses, _DEFAULT_PO_STATUSES)

    result = get_production_overview(
        company=company, month=month, year=year,
        warehouses=warehouses_list, wo_statuses=wo_list,
        so_statuses=so_list, pp_statuses=pp_list,
        po_statuses=po_list,
        stock_mode=stock_mode,
    )
    items   = result.get("items", [])
    summary = result.get("summary", {})
    period  = result.get("period", {})

    wb = Workbook()

    # ── Style palette (Tiger / Indigo theme — matches the page) ──────────
    HEAD_FILL    = PatternFill("solid", fgColor="4F46E5")
    HEAD_FONT    = Font(bold=True, color="FFFFFF", size=11)
    SUB_HDR_FILL = PatternFill("solid", fgColor="0F172A")
    SUB_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="F8FAFC")  # zebra row
    SUB_FILL     = PatternFill("solid", fgColor="EEF2FF")
    OK_FILL      = PatternFill("solid", fgColor="D1FAE5")
    WARN_FILL    = PatternFill("solid", fgColor="FEF3C7")
    ERR_FILL     = PatternFill("solid", fgColor="FEE2E2")
    SO_FILL      = PatternFill("solid", fgColor="FFF1F2")
    COVER_TITLE_FONT = Font(bold=True, size=18, color="4F46E5")
    NOTE_FONT    = Font(italic=True, color="64748B", size=9)
    THIN         = Side(border_style="thin", color="E2E8F0")
    BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTRE       = Alignment(horizontal="center", vertical="center")
    LEFT_WRAP    = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def _fmt_header(ws, row, headers, fill=HEAD_FILL, font=HEAD_FONT):
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = CENTRE
            cell.border    = BORDER

    def _autosize(ws, headers, sample_count=20, max_w=42):
        for col_idx, h in enumerate(headers, start=1):
            max_len = len(str(h))
            for row in range(2, min(2 + sample_count, ws.max_row + 1)):
                v = ws.cell(row=row, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, min(len(str(v)), max_w + 5))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_w)

    def _alt_shade(ws, start_row, end_row, n_cols):
        # Zebra striping — every even row gets the alt fill.
        for r in range(start_row, end_row + 1):
            if (r - start_row) % 2 == 1:
                for c in range(1, n_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.fill is None or cell.fill.fgColor is None or cell.fill.fgColor.rgb in (None, "00000000"):
                        cell.fill = ALT_FILL

    def _set_autofilter(ws, n_rows, n_cols, header_row=1):
        if n_rows <= header_row:
            return
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(n_cols)}{n_rows}"
        )

    # ── Bulk fetch Item Master fields for sheet 4 ────────────────────────
    item_codes = [it["item_code"] for it in items]
    master_rows = []
    if item_codes:
        codes_sql = _sql_in(item_codes)
        master_rows = frappe.db.sql(f"""
            SELECT name, item_name, item_group, stock_uom, description,
                   COALESCE(standard_rate, 0)             AS standard_rate,
                   COALESCE(valuation_rate, 0)            AS valuation_rate,
                   COALESCE(custom_toc_auto_manufacture,0) AS auto_manufacture,
                   COALESCE(custom_toc_auto_purchase, 0)   AS auto_purchase,
                   COALESCE(is_purchase_item, 0)           AS is_purchase_item,
                   COALESCE(is_stock_item, 0)              AS is_stock_item,
                   COALESCE(disabled, 0)                   AS disabled
            FROM `tabItem`
            WHERE name IN {codes_sql}
            ORDER BY item_group, name
        """, as_dict=True)
    master_by_code = {r.name: r for r in master_rows}
    uom_by_code    = {it["item_code"]: (it.get("uom_conversions") or []) for it in items}

    # ────────────────────────────────────────────────────────────────────
    # SHEET 1 — COVER (filters applied, how to read the workbook)
    # ────────────────────────────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "Cover"
    ws_cover["A1"] = f"Production Overview — {period.get('month_name','')} {period.get('year','')}"
    ws_cover["A1"].font = COVER_TITLE_FONT
    ws_cover.merge_cells("A1:E1")
    ws_cover["A2"] = (
        "All quantities are in stock UOM unless noted. Per-UOM breakdown lives "
        "on the UOM Comparison sheet. Counts on the dashboard cards are item "
        "counts (not aggregated qty)."
    )
    ws_cover["A2"].font = NOTE_FONT
    ws_cover.merge_cells("A2:E2")
    ws_cover.row_dimensions[2].height = 28

    _fmt_header(ws_cover, 4, ["Filter", "Value"], fill=SUB_HDR_FILL, font=SUB_HDR_FONT)
    cover_rows = [
        ("Company",        company or "(all)"),
        ("Period",         f"{period.get('month_name','')} {period.get('year','')}"),
        ("Warehouses",     ", ".join(warehouses_list) if warehouses_list else "(all)"),
        ("WO Statuses",    ", ".join(wo_list)         if wo_list         else "(none)"),
        ("SO Statuses",    ", ".join(so_list)         if so_list         else "(none)"),
        ("PP Statuses",    ", ".join(pp_list)         if pp_list         else "(none)"),
        ("PO Statuses",    ", ".join(po_list)         if po_list         else "(none)"),
        ("Stock View",     stock_mode),
        ("Generated At",   getdate(today()).strftime("%Y-%m-%d")),
    ]
    for i, (k, v) in enumerate(cover_rows, start=5):
        ws_cover.cell(row=i, column=1, value=k).border = BORDER
        c = ws_cover.cell(row=i, column=2, value=str(v))
        c.border = BORDER
        c.alignment = LEFT_WRAP

    _fmt_header(ws_cover, len(cover_rows) + 7, ["Item Counts", "Value"], fill=SUB_HDR_FILL, font=SUB_HDR_FONT)
    cnt_rows = [
        ("Total Items shown",                summary.get("total_items", 0)),
        ("Items with Shortage",              summary.get("items_with_shortage", 0)),
        ("Items with NO Shortage",           summary.get("items_no_shortage", 0)),
        ("Items with Open Work Orders",      summary.get("items_with_open_wos", 0)),
        ("Items in Active Production Plans", summary.get("items_with_active_pp", 0)),
        ("Items with Curr Month SO",         summary.get("items_with_curr_so", 0)),
        ("Items Dispatched this Month",      summary.get("items_dispatched", 0)),
        ("Items with Sales Projection",      summary.get("items_with_projection", 0)),
        ("Items at >= 100% Target",          summary.get("items_target_hit", 0)),
        ("BLOCKED items (short + 0 possible)", summary.get("items_blocked", 0)),
        ("Sub-assembly items",               summary.get("items_sub_assembly", 0)),
    ]
    for i, (k, v) in enumerate(cnt_rows, start=len(cover_rows) + 8):
        ws_cover.cell(row=i, column=1, value=k).border = BORDER
        c = ws_cover.cell(row=i, column=2, value=v)
        c.border = BORDER
        c.alignment = Alignment(horizontal="right")
    ws_cover.column_dimensions["A"].width = 36
    ws_cover.column_dimensions["B"].width = 32

    # ────────────────────────────────────────────────────────────────────
    # SHEET 2 — SHEET GUIDE (use cases for every sheet, the workbook map)
    # ────────────────────────────────────────────────────────────────────
    ws_guide = wb.create_sheet("Sheet Guide")
    ws_guide["A1"] = "Sheet Guide — what each tab in this workbook is for"
    ws_guide["A1"].font = COVER_TITLE_FONT
    ws_guide.merge_cells("A1:D1")
    ws_guide["A2"] = (
        "Use this guide to know which sheet to open for each question. "
        "Every data sheet has the indigo header row with AutoFilter enabled."
    )
    ws_guide["A2"].font = NOTE_FONT
    ws_guide.merge_cells("A2:D2")
    ws_guide.row_dimensions[2].height = 30

    guide_headers = ["Sheet", "Purpose", "Use Cases (real-world)", "Key Columns / Notes"]
    _fmt_header(ws_guide, 4, guide_headers, fill=SUB_HDR_FILL, font=SUB_HDR_FONT)
    guide_rows = [
        (
            "Cover",
            "Title, applied filters, item-count summary.",
            "First place to look — confirms WHAT you exported (which Company, "
            "Period, Warehouses, Statuses, Stock View) and the headline "
            "counts of items in each state.",
            "Filter rows + Item Counts table.",
        ),
        (
            "Sheet Guide",
            "(this sheet) Workbook map.",
            "When you forget which sheet has which data — look here.",
            "—",
        ),
        (
            "Simple Report",
            "One-line-per-item planner snapshot. Pending SO vs WO vs stock vs "
            "target vs shortage.",
            "Daily quick-look for the production planner. \"Can I cover the "
            "pending orders with what I have + what's already in WO, or do I "
            "need to start more?\". The two SHORTAGE columns answer that "
            "directly: Shortage vs Physical Stock = pure firefight; Shortage "
            "vs (Stock + Pending WO) = remaining gap after open WOs land.",
            "Item Code | Item Name | Stock UOM | Total Pending SO | "
            "Open WO Pending | Sales Projection | Stock on Hand (Physical) | "
            "Target Production | Target Calc | Shortage vs Physical | "
            "Shortage vs Stock+WO.",
        ),
        (
            "Overview",
            "Full 30-column item table — same columns visible on the page.",
            "When you need every metric. Use AutoFilter to slice by item "
            "group / has shortage / curr SO etc. Description column lets "
            "you read what the item actually is.",
            "Frozen pane on Item Name (C2). Conditional tints: red = has "
            "shortage, pink = open SO. Alt rows for scan readability.",
        ),
        (
            "UOM Comparison",
            "Same qty viewed in EVERY UOM the item supports.",
            "When stock UOM is Gram but you bought in Kg — look up the Kg "
            "row. Or when sales is in Pcs but you produce in Cartons — see "
            "both. Stock UOM rows highlighted in indigo.",
            "Item Code | Field (Planned/Produced/SO/Dispatch/Projection/"
            "Stock/Possible/Target) | Stock UOM Qty | UOM | Factor | "
            "Converted Qty.",
        ),
        (
            "Item Master",
            "Reference catalogue — the item dictionary.",
            "When you need item description, valuation rate, TOC flags, or "
            "the FULL list of UOM conversions for an item. Useful for "
            "cross-checking the costing or item-master setup.",
            "Item Code | Name | Group | Stock UOM | Description | Standard "
            "Rate | Valuation Rate | TOC Auto-Manufacture | TOC Auto-"
            "Purchase | Is Purchase | Is Stock | Disabled | All UOM "
            "Conversions (single string).",
        ),
        (
            "Group Pivot",
            "Pre-aggregated counts per Item Group.",
            "When you want a one-page management summary by group: how many "
            "items in each group have shortage / open WO / curr SO / are "
            "blocked. TOTAL row at the bottom.",
            "Item Group | Total Items | With Shortage | With Open WO | "
            "With Active PP | With Curr SO | Dispatched | With Projection | "
            "Target Hit | Blocked | Sub-Assembly Count.",
        ),
        (
            "Shortage Drivers",
            "Procurement priority list.",
            "When components are blocking production — this sheet says which "
            "components to chase first. Rows tinted RED if the component "
            "blocks 5+ parents, AMBER if 2–4. Buying these first closes the "
            "biggest production logjam.",
            "Component Item | Item Name | UOM | # Parents Blocked | Total "
            "Shortage Qty.",
        ),
    ]
    for r, (s, p, uc, k) in enumerate(guide_rows, start=5):
        cells = [
            ws_guide.cell(row=r, column=1, value=s),
            ws_guide.cell(row=r, column=2, value=p),
            ws_guide.cell(row=r, column=3, value=uc),
            ws_guide.cell(row=r, column=4, value=k),
        ]
        for c in cells:
            c.border = BORDER
            c.alignment = LEFT_WRAP
        cells[0].font = Font(bold=True, color="4F46E5", size=11)
        if (r - 5) % 2 == 1:
            for c in cells:
                c.fill = ALT_FILL
        ws_guide.row_dimensions[r].height = 60
    ws_guide.column_dimensions["A"].width = 20
    ws_guide.column_dimensions["B"].width = 36
    ws_guide.column_dimensions["C"].width = 60
    ws_guide.column_dimensions["D"].width = 56

    # ────────────────────────────────────────────────────────────────────
    # SHEET 3 — SIMPLE REPORT (planner one-liner per item)
    # ────────────────────────────────────────────────────────────────────
    # Spec (2026-05-01 #8): Item name, stock UOM, total pending SO (per
    # filter), total open WO pending qty, total projection, total physical
    # stock, target production (with formula in a header row note),
    # shortage vs physical stock, shortage vs (stock + pending WO).
    #
    # Why a separate sheet: the full Overview has 30 columns. The planner
    # asks "do I need to start production today?" — that's answered in 9
    # columns. This sheet exposes only those 9.
    #
    # Anchor for shortage: TARGET PRODUCTION (max(Projection, Total Sales)).
    # Two variants:
    #   Shortage vs Physical Stock        = max(Target − Stock, 0)
    #   Shortage vs (Stock + Pending WO)  = max(Target − (Stock + Pending WO), 0)
    # First answers "can I cover today with what I have?".
    # Second answers "after open WOs land, what is left?".

    # Recompute physical stock independent of stock_mode (so this sheet is
    # always strict physical, even if user picked Physical+Expected on the page).
    physical_stock_map = _get_stock(item_codes, warehouses, "physical")

    ws_simple = wb.create_sheet("Simple Report")

    # Pre-header row with the formulas in plain English. Helps a layman planner.
    ws_simple["A1"] = "Simple Report — pending demand vs supply per item"
    ws_simple["A1"].font = COVER_TITLE_FONT
    ws_simple.merge_cells("A1:K1")
    ws_simple["A2"] = (
        "Total Pending SO  = sum of pending qty across ALL pending Sales Orders that match your SO Status filter (any month). "
        "Open WO Pending Qty = SUM(Work Order.qty − produced_qty) on open WOs. "
        "Total Projection  = qty_in_stock_uom from this month's Sales Projection. "
        "Stock on Hand (Physical) = Bin.actual_qty (this report uses PHYSICAL only, regardless of Stock View). "
        "Target Production = max(Sales Projection, Total Curr Month Sales). "
        "Shortage vs Physical Stock = max(Target − Stock, 0). "
        "Shortage vs Stock + Pending WO = max(Target − (Stock + Pending WO), 0)."
    )
    ws_simple["A2"].font = NOTE_FONT
    ws_simple["A2"].alignment = LEFT_WRAP
    ws_simple.merge_cells("A2:K2")
    ws_simple.row_dimensions[2].height = 56

    simple_headers = [
        "Item Code", "Item Name", "Item Group", "Stock UOM",
        "Total Pending SO",
        "Open WO Pending Qty",
        "Total Projection",
        "Stock on Hand (Physical)",
        "Target Production",
        "Shortage vs Physical Stock",
        "Shortage vs (Stock + Pending WO)",
    ]
    _fmt_header(ws_simple, 4, simple_headers)
    for r, item in enumerate(items, start=5):
        code = item["item_code"]
        target  = flt(item.get("target_production", 0))
        stock   = flt(physical_stock_map.get(code, 0))
        wo_pend = flt(item.get("pending_wo_qty", 0))
        short_phys  = max(target - stock, 0)
        short_total = max(target - (stock + wo_pend), 0)
        row = [
            code,
            item.get("item_name", ""),
            item.get("item_group", ""),
            item.get("stock_uom", ""),
            flt(item.get("total_pending_so", 0)),
            wo_pend,
            flt(item.get("curr_projection", 0)),
            stock,
            target,
            round(short_phys, 3),
            round(short_total, 3),
        ]
        for col_idx, val in enumerate(row, start=1):
            cell = ws_simple.cell(row=r, column=col_idx, value=val)
            cell.border = BORDER
            if col_idx >= 5 and isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        # Tinting: red if any shortage, alt-row otherwise
        if short_total > 0:
            for c in range(1, len(simple_headers) + 1):
                ws_simple.cell(row=r, column=c).fill = ERR_FILL
        elif short_phys > 0:
            for c in range(1, len(simple_headers) + 1):
                ws_simple.cell(row=r, column=c).fill = WARN_FILL
        elif (r - 5) % 2 == 1:
            for c in range(1, len(simple_headers) + 1):
                ws_simple.cell(row=r, column=c).fill = ALT_FILL
    ws_simple.freeze_panes = "C5"
    _autosize(ws_simple, simple_headers, max_w=30)
    # AutoFilter on the data table only (rows 4..end). Header row is row 4.
    if ws_simple.max_row >= 5:
        ws_simple.auto_filter.ref = (
            f"A4:{get_column_letter(len(simple_headers))}{ws_simple.max_row}"
        )

    # ────────────────────────────────────────────────────────────────────
    # SHEET 4 — OVERVIEW (the data table)
    # ────────────────────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Overview")
    headers = [
        "Item Code", "Item Name", "Description", "Item Group", "Stock UOM",
        "Sub-Assembly", "Has Open SO", "Open WOs", "Active PPs",
        "Planned Qty", "Pending WO Qty", "Produced",
        "Prev Month SO", "Curr Month SO",
        "Prev Month Dispatch", "Curr Month Dispatch",
        "Sales Projection", "Total Curr Sales",
        "Proj vs Sales %", "Coverage %",
        "Target Production", "Target Gap", "% Target Achieved",
        "Stock on Hand", "Possible Qty",
        "Has Shortage", "Active BOM",
        "BOM Std Cost (per unit)", "Actual Cost (per unit)", "Variance %",
    ]
    _fmt_header(ws1, 1, headers)
    for i, item in enumerate(items, start=2):
        master = master_by_code.get(item["item_code"]) or {}
        row = [
            item["item_code"], item["item_name"], (master.get("description") or "")[:500],
            item.get("item_group", ""),
            item["stock_uom"],
            "Yes" if item["is_sub_assembly"] else "No",
            "Yes" if item.get("has_open_so") else "No",
            item["open_wo_count"], item.get("pp_count", 0),
            item["planned_qty"], item.get("pending_wo_qty", 0), item["actual_qty"],
            item["prev_month_order"], item["curr_month_order"],
            item.get("prev_dispatch", 0), item["curr_dispatch"],
            item["curr_projection"], item["total_curr_sales"],
            round(item["projection_vs_sales"] * 100, 1) if item["projection_vs_sales"] else 0,
            item.get("coverage_pct", 0),
            item.get("target_production", 0),
            item.get("target_gap", 0),
            item.get("target_achieved_pct", 0),
            item["stock"], item["possible_qty"],
            "Yes" if item["has_shortage"] else "No",
            item.get("active_bom", ""),
            item.get("cost_summary", {}).get("bom_total", 0),
            item.get("cost_summary", {}).get("actual_total", 0),
            item.get("cost_summary", {}).get("variance_pct", 0),
        ]
        for col_idx, val in enumerate(row, start=1):
            cell = ws1.cell(row=i, column=col_idx, value=val)
            cell.border = BORDER
            if col_idx >= 8 and isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")
        if item["has_shortage"]:
            for col_idx in range(1, len(headers) + 1):
                ws1.cell(row=i, column=col_idx).fill = ERR_FILL
        elif item.get("has_open_so"):
            for col_idx in range(1, len(headers) + 1):
                ws1.cell(row=i, column=col_idx).fill = SO_FILL
    # Zebra rows on the rest (skip already-tinted shortage/SO rows)
    for i, item in enumerate(items, start=2):
        if not item["has_shortage"] and not item.get("has_open_so") and (i - 2) % 2 == 1:
            for c in range(1, len(headers) + 1):
                ws1.cell(row=i, column=c).fill = ALT_FILL
    ws1.freeze_panes = "C2"
    _autosize(ws1, headers, max_w=36)
    _set_autofilter(ws1, ws1.max_row, len(headers))

    # ────────────────────────────────────────────────────────────────────
    # SHEET 3 — UOM COMPARISON
    # ────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("UOM Comparison")
    uom_headers = ["Item Code", "Item Name", "Item Group", "Stock UOM",
                   "Field", "Stock UOM Qty", "UOM", "Factor", "Converted Qty"]
    _fmt_header(ws2, 1, uom_headers)
    fields = [
        ("Planned Qty",     "planned_qty"),
        ("Produced",        "actual_qty"),
        ("Prev Month SO",   "prev_month_order"),
        ("Curr Month SO",   "curr_month_order"),
        ("Curr Dispatch",   "curr_dispatch"),
        ("Sales Projection","curr_projection"),
        ("Stock on Hand",   "stock"),
        ("Possible Qty",    "possible_qty"),
        ("Target Production","target_production"),
    ]
    r = 2
    for item in items:
        uom_list = item.get("uom_conversions") or [{"uom": item["stock_uom"], "factor": 1.0}]
        # Always include stock UOM as factor 1 row even if conversions exist
        if item["stock_uom"] and not any((u.get("uom") == item["stock_uom"]) for u in uom_list):
            uom_list = [{"uom": item["stock_uom"], "factor": 1.0}] + list(uom_list)
        for label, key in fields:
            qty = flt(item.get(key, 0))
            for u in uom_list:
                f = flt(u.get("factor"))
                if f <= 0:
                    continue
                ws2.append([
                    item["item_code"], item["item_name"], item.get("item_group", ""),
                    item["stock_uom"], label, qty,
                    u.get("uom") or item["stock_uom"], f,
                    round(qty / f, 4),
                ])
                for col_idx in range(1, len(uom_headers) + 1):
                    ws2.cell(row=r, column=col_idx).border = BORDER
                if u.get("uom") == item["stock_uom"]:
                    for col_idx in range(1, len(uom_headers) + 1):
                        ws2.cell(row=r, column=col_idx).fill = SUB_FILL
                elif (r % 2) == 1:
                    for col_idx in range(1, len(uom_headers) + 1):
                        ws2.cell(row=r, column=col_idx).fill = ALT_FILL
                r += 1
    ws2.freeze_panes = "B2"
    _autosize(ws2, uom_headers)
    _set_autofilter(ws2, ws2.max_row, len(uom_headers))

    # ────────────────────────────────────────────────────────────────────
    # SHEET 4 — ITEM MASTER (reference catalogue)
    # ────────────────────────────────────────────────────────────────────
    ws_master = wb.create_sheet("Item Master")
    master_headers = [
        "Item Code", "Item Name", "Item Group", "Stock UOM", "Description",
        "Standard Rate (INR)", "Valuation Rate (INR)",
        "TOC Auto-Manufacture", "TOC Auto-Purchase",
        "Is Purchase Item", "Is Stock Item", "Disabled",
        "All UOM Conversions",
    ]
    _fmt_header(ws_master, 1, master_headers)
    r = 2
    for code in item_codes:
        m   = master_by_code.get(code, {})
        uom = uom_by_code.get(code, [])
        uom_str = "; ".join(f"{u.get('uom')} (×{flt(u.get('factor'))})" for u in uom if u.get("uom"))
        if not uom_str and m.get("stock_uom"):
            uom_str = f"{m.get('stock_uom')} (×1)"
        row = [
            code, m.get("item_name", ""), m.get("item_group", ""),
            m.get("stock_uom", ""), (m.get("description") or "")[:1000],
            flt(m.get("standard_rate", 0)), flt(m.get("valuation_rate", 0)),
            "Yes" if cint(m.get("auto_manufacture", 0)) else "No",
            "Yes" if cint(m.get("auto_purchase", 0))    else "No",
            "Yes" if cint(m.get("is_purchase_item", 0)) else "No",
            "Yes" if cint(m.get("is_stock_item", 0))    else "No",
            "Yes" if cint(m.get("disabled", 0))         else "No",
            uom_str,
        ]
        for col_idx, val in enumerate(row, start=1):
            cell = ws_master.cell(row=r, column=col_idx, value=val)
            cell.border = BORDER
            if col_idx == 5 or col_idx == 13:
                cell.alignment = LEFT_WRAP
        if (r - 2) % 2 == 1:
            for c in range(1, len(master_headers) + 1):
                ws_master.cell(row=r, column=c).fill = ALT_FILL
        r += 1
    ws_master.freeze_panes = "B2"
    _autosize(ws_master, master_headers, max_w=44)
    _set_autofilter(ws_master, ws_master.max_row, len(master_headers))

    # ────────────────────────────────────────────────────────────────────
    # SHEET 5 — GROUP PIVOT (pre-aggregated, Google-Sheets-friendly)
    # ────────────────────────────────────────────────────────────────────
    ws_pivot = wb.create_sheet("Group Pivot")
    pivot_headers = [
        "Item Group", "Total Items", "With Shortage",
        "With Open WO", "With Active PP", "With Curr SO",
        "Dispatched", "With Projection",
        "Target Hit (>=100%)", "Blocked", "Sub-Assembly Count",
    ]
    _fmt_header(ws_pivot, 1, pivot_headers)
    by_group = {}
    for it in items:
        g = it.get("item_group") or "(no group)"
        b = by_group.setdefault(g, {
            "total":0, "shortage":0, "wo":0, "pp":0, "curr_so":0,
            "disp":0, "proj":0, "hit":0, "blocked":0, "sub":0
        })
        b["total"]    += 1
        if it["has_shortage"]:                          b["shortage"] += 1
        if it.get("open_wo_count", 0) > 0:              b["wo"]       += 1
        if it.get("pp_count", 0) > 0:                   b["pp"]       += 1
        if it.get("curr_month_order", 0) > 0:           b["curr_so"]  += 1
        if it.get("curr_dispatch", 0) > 0:              b["disp"]     += 1
        if it.get("curr_projection", 0) > 0:            b["proj"]     += 1
        if it.get("target_achieved_pct", 0) >= 100:     b["hit"]      += 1
        if it.get("has_shortage") and it.get("possible_qty", 0) == 0:
            b["blocked"] += 1
        if it.get("is_sub_assembly"):                   b["sub"]      += 1

    pivot_sorted = sorted(by_group.items(), key=lambda kv: kv[1]["total"], reverse=True)
    for r, (g, b) in enumerate(pivot_sorted, start=2):
        row = [g, b["total"], b["shortage"], b["wo"], b["pp"],
               b["curr_so"], b["disp"], b["proj"], b["hit"], b["blocked"], b["sub"]]
        for col_idx, val in enumerate(row, start=1):
            cell = ws_pivot.cell(row=r, column=col_idx, value=val)
            cell.border = BORDER
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="right")
        if (r - 2) % 2 == 1:
            for c in range(1, len(pivot_headers) + 1):
                ws_pivot.cell(row=r, column=c).fill = ALT_FILL
    # Total row (bold, indigo bg)
    total_r = ws_pivot.max_row + 1
    totals = ["TOTAL"]
    for k in ("total","shortage","wo","pp","curr_so","disp","proj","hit","blocked","sub"):
        totals.append(sum(b[k] for _, b in pivot_sorted))
    for col_idx, val in enumerate(totals, start=1):
        cell = ws_pivot.cell(row=total_r, column=col_idx, value=val)
        cell.fill = SUB_HDR_FILL
        cell.font = SUB_HDR_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="right" if col_idx > 1 else "left")
    ws_pivot.freeze_panes = "B2"
    _autosize(ws_pivot, pivot_headers)
    _set_autofilter(ws_pivot, ws_pivot.max_row - 1, len(pivot_headers))  # exclude TOTAL row

    # ────────────────────────────────────────────────────────────────────
    # SHEET 6 — SHORTAGE DRIVERS (procurement priority)
    # ────────────────────────────────────────────────────────────────────
    ws_short = wb.create_sheet("Shortage Drivers")
    sd_headers = ["Component Item", "Item Name", "UOM",
                  "# Parents Blocked", "Total Shortage Qty"]
    _fmt_header(ws_short, 1, sd_headers)
    drv = {}
    for it in items:
        for c in (it.get("shortage_components") or []):
            code = c.get("item_code")
            if not code:
                continue
            d = drv.setdefault(code, {
                "name": c.get("item_name") or master_by_code.get(code, {}).get("item_name", ""),
                "uom":  c.get("uom") or master_by_code.get(code, {}).get("stock_uom", ""),
                "blocks": 0,
                "total_short": 0.0,
            })
            d["blocks"]      += 1
            d["total_short"] += flt(c.get("shortage", 0))
    drv_sorted = sorted(drv.items(),
                        key=lambda t: (t[1]["blocks"], t[1]["total_short"]),
                        reverse=True)
    for r, (code, v) in enumerate(drv_sorted, start=2):
        row = [code, v["name"], v["uom"], v["blocks"], round(v["total_short"], 3)]
        for col_idx, val in enumerate(row, start=1):
            cell = ws_short.cell(row=r, column=col_idx, value=val)
            cell.border = BORDER
            if col_idx >= 4:
                cell.alignment = Alignment(horizontal="right")
        if (r - 2) % 2 == 1:
            for c in range(1, len(sd_headers) + 1):
                ws_short.cell(row=r, column=c).fill = ALT_FILL
        if v["blocks"] >= 5:
            for c in range(1, len(sd_headers) + 1):
                ws_short.cell(row=r, column=c).fill = ERR_FILL
        elif v["blocks"] >= 2:
            for c in range(1, len(sd_headers) + 1):
                ws_short.cell(row=r, column=c).fill = WARN_FILL
    ws_short.freeze_panes = "B2"
    _autosize(ws_short, sd_headers)
    _set_autofilter(ws_short, ws_short.max_row, len(sd_headers))

    # ── Save into Frappe response ─────────────────────────────────────
    fname = (
        f"production_overview_{period.get('month_name','')}_{period.get('year','')}.xlsx"
    )
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    frappe.response.filename     = fname
    frappe.response.filecontent  = bio.getvalue()
    frappe.response.type         = "binary"
    frappe.response.display_content_as = "attachment"


# =============================================================================
# CHART DATA (Tab 3)
# =============================================================================

@frappe.whitelist()
def get_chart_data(company=None, month=None, year=None, warehouses=None,
                   wo_statuses=None, so_statuses=None, pp_statuses=None,
                   po_statuses=None, stock_mode="physical"):
    """
    Aggregated data for the Charts tab — production-team-meaningful insights.

    Returned series (each keyed for Chart.js consumption):

      shortage_pie         pie  — items with vs without shortage
      bar_orders           bar  — top 10 by current-month order, planned/dispatch overlay
      bar_projection       bar  — projection vs actual sales for top 10 by projection
      bar_wo_by_group      bar  — open WO count per item group (top 10)

      bar_priority_action  bar  — top 10 items by Target Gap. Each bar shows what
                                  is blocking shipment: shortfall after stock + pending WOs.
                                  This is "today's priority action board".
      bar_daily_need       bar  — top 10 items with the largest remaining production
                                  need this month (target − produced). For each item,
                                  also returns a `daily_need` value (remaining ÷ working
                                  days left in the month) so the production manager
                                  can size each day's batch.
      readiness_pie        pie  — production readiness mix: Ready (possible_qty > 0
                                  and demand exists) / Partial / Blocked / No demand.
      bar_coverage_health  bar  — distribution of items across coverage% buckets
                                  (<50, 50-99, 100-149, ≥150). Shows how well
                                  planning inputs (Projection + Prev SO) cover
                                  current month sales.
      bar_shortage_drivers bar  — top 10 component items causing shortages, ranked
                                  by how many parent items they block. Procurement
                                  priority list.
      summary              dict — counts/sums copied from Overview tab for the
                                  metric cards above the charts.
    """
    import calendar as _cal
    from datetime import date as _date

    result = get_production_overview(
        company=company, month=month, year=year,
        warehouses=warehouses, wo_statuses=wo_statuses,
        so_statuses=so_statuses, pp_statuses=pp_statuses,
        po_statuses=po_statuses, stock_mode=stock_mode,
    )
    items = result.get("items", [])
    period = result.get("period", {})

    # ── Working days remaining for daily-need calc ───────────────────────
    today_d  = getdate(today())
    p_month  = cint(period.get("month_num") or today_d.month)
    p_year   = cint(period.get("year")      or today_d.year)
    last_day = _cal.monthrange(p_year, p_month)[1]
    if today_d.year == p_year and today_d.month == p_month:
        start_day = today_d.day
    else:
        start_day = 1
    working_days_left = 0
    for d in range(start_day, last_day + 1):
        wd = _date(p_year, p_month, d).weekday()  # 0=Mon..6=Sun
        if wd < 6:                                  # exclude Sunday
            working_days_left += 1
    if working_days_left < 1:
        working_days_left = 1

    # ── Existing pies / bars ──────────────────────────────────────────────
    type_counts = {}
    for item in items:
        t = item["item_type"]
        type_counts[t] = type_counts.get(t, 0) + 1
    shortage_yes = sum(1 for x in items if x["has_shortage"])
    shortage_no  = len(items) - shortage_yes

    top_order = sorted(items, key=lambda x: x["curr_month_order"], reverse=True)[:10]
    bar_orders = {
        "labels":     [x["item_code"]        for x in top_order],
        "item_names": [x.get("item_name","") for x in top_order],
        "planned":    [x["planned_qty"]      for x in top_order],
        "actual":     [x["actual_qty"]       for x in top_order],
        "orders":     [x["curr_month_order"] for x in top_order],
        "dispatch":   [x["curr_dispatch"]    for x in top_order],
    }

    top_proj = sorted(items, key=lambda x: x["curr_projection"], reverse=True)[:10]
    bar_proj = {
        "labels":     [x["item_code"]        for x in top_proj],
        "item_names": [x.get("item_name","") for x in top_proj],
        "projection": [x["curr_projection"]  for x in top_proj],
        "sales":      [x["total_curr_sales"] for x in top_proj],
        "stock":      [x["stock"]            for x in top_proj],
    }

    group_wo = {}
    for item in items:
        g = item["item_group"] or "Other"
        group_wo[g] = group_wo.get(g, 0) + item["open_wo_count"]
    group_wo_sorted = sorted(group_wo.items(), key=lambda x: x[1], reverse=True)[:10]

    # ── NEW: Priority Action Board ────────────────────────────────────────
    # Top 10 items by Target Gap (descending). Bar split into:
    #   covered_by_stock  (green)     = min(target, stock)
    #   covered_by_wo     (blue)      = pending WO qty contribution after stock
    #   gap               (red)       = remaining = target − stock − pending_wo
    # = the daily punch list. Manager looks at this first thing each morning.
    items_with_gap = [x for x in items if (x.get("target_gap") or 0) > 0]
    top_priority   = sorted(items_with_gap,
                            key=lambda x: x.get("target_gap") or 0,
                            reverse=True)[:10]
    bar_priority = {
        "labels":         [x["item_code"]                 for x in top_priority],
        "item_names":     [x.get("item_name","")          for x in top_priority],
        "target":         [x.get("target_production",0)   for x in top_priority],
        "stock":          [min(flt(x.get("target_production",0)), flt(x.get("stock",0))) for x in top_priority],
        "wo_pending":     [
            min(flt(x.get("pending_wo_qty",0)),
                max(flt(x.get("target_production",0)) - flt(x.get("stock",0)), 0))
            for x in top_priority
        ],
        "gap":            [x.get("target_gap",0)          for x in top_priority],
        "achieved_pct":   [x.get("target_achieved_pct",0) for x in top_priority],
        "stock_uoms":     [x.get("stock_uom","")          for x in top_priority],
    }

    # ── NEW: Daily Production Need ────────────────────────────────────────
    # Remaining = max(target − produced, 0). Daily = Remaining ÷ working_days_left.
    # Top 10 items by Remaining. Drives "what to make today" sizing.
    items_with_remaining = []
    for x in items:
        rem = max(flt(x.get("target_production",0)) - flt(x.get("actual_qty",0)), 0)
        if rem > 0:
            items_with_remaining.append((x, rem))
    items_with_remaining.sort(key=lambda t: t[1], reverse=True)
    top_daily = items_with_remaining[:10]
    bar_daily_need = {
        "labels":          [x["item_code"]            for x, _ in top_daily],
        "item_names":      [x.get("item_name","")     for x, _ in top_daily],
        "remaining":       [round(rem, 3)             for _, rem in top_daily],
        "daily_need":      [round(rem / working_days_left, 3) for _, rem in top_daily],
        "produced":        [x.get("actual_qty",0)     for x, _ in top_daily],
        "target":          [x.get("target_production",0) for x, _ in top_daily],
        "achieved_pct":    [x.get("target_achieved_pct",0) for x, _ in top_daily],
        "stock_uoms":      [x.get("stock_uom","")     for x, _ in top_daily],
        "working_days_left": working_days_left,
    }

    # ── NEW: Production Readiness pie ─────────────────────────────────────
    # Ready    = possible_qty > 0 AND has demand (curr_so or projection or target>0)
    # Partial  = possible_qty > 0 but possible_qty < target
    # Blocked  = possible_qty = 0 AND has_shortage
    # No-demand= no curr_so, no projection, no target
    ready = partial = blocked = no_demand = 0
    for x in items:
        target = flt(x.get("target_production", 0))
        possible = flt(x.get("possible_qty", 0))
        has_short = x.get("has_shortage", False)
        if target <= 0 and flt(x.get("curr_month_order",0)) <= 0:
            no_demand += 1
        elif possible <= 0 and has_short:
            blocked += 1
        elif possible >= target and target > 0:
            ready += 1
        else:
            partial += 1
    readiness_pie = {
        "Ready":     ready,
        "Partial":   partial,
        "Blocked":   blocked,
        "No Demand": no_demand,
    }

    # ── NEW: Coverage Health distribution ─────────────────────────────────
    # Bucket items by coverage_pct so the manager sees how planning inputs
    # (Projection + Prev SO carryover) line up with this month's actual sales.
    cov_buckets = {"<50%": 0, "50–99%": 0, "100–149%": 0, "≥150%": 0, "No Sales": 0}
    for x in items:
        if not flt(x.get("total_curr_sales", 0)):
            cov_buckets["No Sales"] += 1
            continue
        c = flt(x.get("coverage_pct", 0))
        if   c <  50:  cov_buckets["<50%"]    += 1
        elif c <  100: cov_buckets["50–99%"]  += 1
        elif c <  150: cov_buckets["100–149%"] += 1
        else:          cov_buckets["≥150%"]   += 1
    bar_coverage_health = {
        "labels": list(cov_buckets.keys()),
        "values": list(cov_buckets.values()),
    }

    # ── NEW: Shortage Drivers — components blocking the most parents ──────
    # For each row, every short_components entry contributes 1 vote to that
    # component's "blocks" score. Rank top 10 — that's the procurement
    # priority list. Useful even if a component has small qty short for
    # one item but small qty short for many items.
    short_drivers = {}
    for it in items:
        for c in (it.get("shortage_components") or []):
            code = c.get("item_code")
            if not code:
                continue
            d = short_drivers.setdefault(code, {
                "blocks": 0,
                "item_name": c.get("item_name") or "",
                "uom": c.get("uom") or "",
                "total_short": 0.0,
            })
            d["blocks"]      += 1
            d["total_short"] += flt(c.get("shortage", 0))
    drivers_sorted = sorted(short_drivers.items(),
                            key=lambda t: (t[1]["blocks"], t[1]["total_short"]),
                            reverse=True)[:10]
    bar_shortage_drivers = {
        "labels":     [code for code, _ in drivers_sorted],
        "item_names": [v["item_name"]  for _, v in drivers_sorted],
        "blocks":     [v["blocks"]     for _, v in drivers_sorted],
        "total_short":[round(v["total_short"], 3) for _, v in drivers_sorted],
        "uoms":       [v["uom"]        for _, v in drivers_sorted],
    }

    return {
        "type_pie":            type_counts,
        "shortage_pie":        {"Yes": shortage_yes, "No": shortage_no},
        "bar_orders":          bar_orders,
        "bar_projection":      bar_proj,
        "bar_wo_by_group": {
            "labels": [x[0] for x in group_wo_sorted],
            "values": [x[1] for x in group_wo_sorted],
        },
        # NEW production-team series
        "bar_priority_action":  bar_priority,
        "bar_daily_need":       bar_daily_need,
        "readiness_pie":        readiness_pie,
        "bar_coverage_health":  bar_coverage_health,
        "bar_shortage_drivers": bar_shortage_drivers,
        # Meta
        "working_days_left":    working_days_left,
        "summary":              result.get("summary", {}),
        "period":               period,
    }


# =============================================================================
# AI ADVISOR (Tab 2) — DeepSeek Integration
# =============================================================================

_POR_AI_SYSTEM_PROMPT = (
    "You are a production planning advisor for an ERPNext-based factory.\n"
    "Data covers Work Orders, Sales Orders, Sales Projections, stock, "
    "BOM-derived shortages, and cost variances.\n\n"
    "STRICT RULES (do not break):\n"
    " - NO greetings. NO 'Hi', 'Hello', 'Sure', 'Certainly', 'Of course'.\n"
    " - NO preamble. NO 'Based on the data...', 'Here is...'.\n"
    " - Lead with the answer or action. Skip everything before it.\n"
    " - Output is ALWAYS HTML (never plain Markdown, never plain text).\n"
    " - Be direct, factual, and actionable. The reader is a factory manager.\n\n"
    "OUTPUT FORMAT (HTML ONLY):\n"
    " - For status text:\n"
    "   <span class=\"por-ai-ok\">good / on track</span>\n"
    "   <span class=\"por-ai-warn\">warning</span>\n"
    "   <span class=\"por-ai-err\">critical</span>\n"
    " - For data: WRAP every table in a scroll container so wide tables\n"
    "   never overflow the chat bubble:\n"
    "   <div class=\"por-ai-tablewrap\" style=\"overflow:auto; max-height:340px;\">\n"
    "     <table class=\"por-ai-table\"><thead>...</thead><tbody>...</tbody></table>\n"
    "   </div>\n"
    "   Tables: keep to <=6 columns and <=20 rows; if more rows are needed,\n"
    "   sort by impact (descending) and append a footer row 'Top N of M shown'.\n"
    " - For action steps:\n"
    "   <ol class=\"por-ai-actions\"><li>verb-led sentence</li>...</ol>\n"
    "   Use 3 steps unless the question explicitly asks for more or fewer.\n"
    " - For lists with sub-detail use <details><summary>...</summary>...</details>.\n\n"
    "VALUE FORMATTING:\n"
    " - ALWAYS include item_code in parentheses: 'Masala Blend (MBLND-500G)'.\n"
    " - Show qty with both UOMs when conversion is meaningful: '5,000 Gram (5 Kg)'.\n"
    " - Currency is INR. Use ₹ symbol. Two decimals max.\n"
    " - Projection vs Sales ratio: >1.0 = outperforming, <1.0 = below projection.\n"
    " - Shortage = active BOM components are short to meet demand.\n"
    " - Coverage % = (Projection + Prev Month SO) / Total Curr Sales * 100.\n\n"
    "RESPONSE LENGTH:\n"
    " - One-sentence questions get one-sentence answers (no table).\n"
    " - 'Top N' / 'Which items' / 'List' questions get a wrapped table.\n"
    " - 'Plan' / 'What should I do' questions end with the action ordered list.\n"
)


@frappe.whitelist()
def get_deepseek_models_por():
    """Return available DeepSeek models for the model selector."""
    return {
        model_id: {
            "name":             cfg["name"],
            "description":      cfg["description"],
            "est_cost_per_call":cfg["est_cost_per_call"],
        }
        for model_id, cfg in DEEPSEEK_MODELS.items()
    }


@frappe.whitelist()
def get_ai_overview_insight(context_json, model=None):
    """
    Auto-insight: run once after data loads.
    Returns a structured briefing (no session, no tool calling).
    """
    if isinstance(context_json, str):
        context = frappe.parse_json(context_json) or {}
    else:
        context = context_json or {}

    api_key = _get_api_key()
    if not api_key:
        return {
            "insight": "<span class='por-ai-warn'>DeepSeek API key not configured. "
                       "Set it in TOC Settings → AI Advisor → DeepSeek API Key.</span>",
            "is_html": True,
        }

    summary = context.get("summary", {})
    items   = context.get("items", [])[:20]  # top 20 only to stay within token budget

    insight_prompt = (
        "Analyse this production overview data and give a structured briefing.\n\n"
        f"SUMMARY: {json.dumps(summary)}\n"
        f"TOP ITEMS (by curr month order qty): {json.dumps(items)}\n\n"
        "OUTPUT FORMAT (HTML):\n"
        "1. One sentence overall status.\n"
        "2. Table of top 3 critical issues: Item | Issue | Impact | Recommended Action\n"
        "3. Exactly 3 action steps as <ol class='por-ai-actions'>.\n"
        "No preamble. Lead with status."
    )

    messages = [
        {"role": "system",  "content": _POR_AI_SYSTEM_PROMPT},
        {"role": "user",    "content": insight_prompt},
    ]

    try:
        resp   = _call_deepseek(messages, api_key=api_key, model=model)
        reply  = resp["choices"][0]["message"]["content"] or ""
    except Exception as exc:
        frappe.log_error(str(exc), "POR AI Auto-Insight")
        reply  = "<span class='por-ai-warn'>AI insight unavailable. Check Error Log → POR AI.</span>"

    return {"insight": reply, "is_html": True}


@frappe.whitelist()
def chat_with_overview_advisor(message, session_id, context_json, model=None):
    """
    Session-persistent AI chat for the Production Overview Advisor.
    Redis key: por:chat:{user}:{session_id}
    """
    if isinstance(context_json, str):
        context = frappe.parse_json(context_json) or {}
    else:
        context = context_json or {}

    api_key = _get_api_key()
    if not api_key:
        return {
            "reply":      "<span class='por-ai-warn'>DeepSeek API key not configured.</span>",
            "session_id": session_id,
            "is_html":    True,
        }

    user      = frappe.session.user or "Guest"
    redis_key = f"{_POR_AI_SESSION_PREFIX}{user}:{session_id}"

    # Load or initialise session
    history_raw = frappe.cache().get_value(redis_key)
    history = json.loads(history_raw) if history_raw else []

    # Build system context (strip large keys before sending to LLM)
    ctx_for_ai = {k: v for k, v in context.items() if k not in ("items", "raw_rows")}
    system_content = (
        f"{_POR_AI_SYSTEM_PROMPT}\n\n"
        f"CURRENT PRODUCTION CONTEXT:\n{json.dumps(ctx_for_ai, default=str)}"
    )

    messages = (
        [{"role": "system", "content": system_content}]
        + history[-_AI_MAX_HISTORY:]
        + [{"role": "user", "content": message}]
    )

    reply_text, updated_messages, _ = _execute_chat_with_tools(
        messages, context, api_key, model=model
    )

    # Save only user + assistant turns
    new_history = [m for m in updated_messages if m.get("role") in ("user", "assistant")]
    frappe.cache().set_value(redis_key, json.dumps(new_history[-_AI_MAX_HISTORY:]),
                             expires_in_sec=_AI_SESSION_TTL)

    return {"reply": reply_text, "session_id": session_id, "is_html": True}


@frappe.whitelist()
def test_ai_connection_por():
    """Diagnostic: verify API key and DeepSeek connectivity."""
    api_key = _get_api_key()
    if not api_key:
        return {"ok": False, "message": "No API key found. Configure in TOC Settings."}
    try:
        resp = _call_deepseek(
            [{"role": "user", "content": "Reply: ok"}],
            api_key=api_key, model="deepseek-chat"
        )
        reply = resp["choices"][0]["message"]["content"]
        return {"ok": True, "message": reply, "model": "deepseek-chat"}
    except Exception as exc:
        return {"ok": False, "message": str(exc)}


# =============================================================================
# PRIVATE HELPERS
# =============================================================================

def _empty_summary():
    return {
        "total_items": 0, "items_with_shortage": 0, "items_no_shortage": 0,
        "total_planned_qty": 0, "total_actual_qty": 0,
        "total_curr_orders": 0, "total_dispatch": 0, "total_projection": 0,
    }


def _sql_in(lst):
    """Build a safe SQL IN(...) clause string from a list."""
    if not lst:
        return "('')"
    escaped = ", ".join(frappe.db.escape(str(x)) for x in lst)
    return f"({escaped})"


def _build_wh_filter(warehouses):
    if not warehouses:
        return "", {}
    sql = _sql_in(warehouses)
    return f"AND warehouse IN {sql}", {}


def _wh_clause(warehouses, column_expr):
    """
    Reusable helper: return ' AND <column_expr> IN (...)' if warehouses set,
    otherwise the empty string. Caller embeds inside an existing WHERE.
    `column_expr` is the SQL expression naming the warehouse column
    (e.g. 'wo.fg_warehouse', 'soi.warehouse').
    """
    if not warehouses:
        return ""
    return f" AND {column_expr} IN {_sql_in(warehouses)}"


def _get_qualifying_items(wo_statuses, so_statuses, curr_month, curr_year,
                          prev_month, prev_year, company, pp_statuses=None):
    """
    Return set of item_codes that qualify for the overview dashboard.

    Per the user-supplied 2026-05-01 spec, EXACTLY four conditions qualify:

      (1) `custom_toc_auto_manufacture = 1` on Item master (TOC App flag).
      (2) Item is the `production_item` of an open Work Order
          OR appears in `tabProduction Plan Item.item_code` of an open
          Production Plan.
      (3) Sub-assembly proper: the item has its own open Work Order AND is
          consumed by a parent Work Order via a shared Production Plan
          (= "will be used in a parent WO but the item itself has its own
          open WO"). Independent BOM components are NOT included.
      (4) Item has at least one OPEN PENDING Sales Order line (no month
          filter — any pending SO surfaces the item).

    Dropped (per spec — these were broader and cluttered the dashboard):
      - Items only present in Sales Projection.
      - Items used purely as BOM components without their own open WO.
    """
    codes = set()
    wo_sql = _sql_in(wo_statuses)
    so_clause = _so_status_clause(so_statuses)
    pp_statuses = pp_statuses or _DEFAULT_PP_STATUSES
    pp_sql = _sql_in(pp_statuses)

    # (1) Auto TOC manufacture items
    rows = frappe.db.sql_list("""
        SELECT name FROM `tabItem`
        WHERE disabled = 0 AND custom_toc_auto_manufacture = 1
    """)
    codes.update(rows)

    # (2a) Items in open Work Orders
    rows = frappe.db.sql_list(f"""
        SELECT DISTINCT production_item FROM `tabWork Order`
        WHERE docstatus = 1 AND status IN {wo_sql}
    """)
    codes.update(rows)

    # (2b) Items in open Production Plans
    rows = frappe.db.sql_list(f"""
        SELECT DISTINCT ppi.item_code
        FROM `tabProduction Plan Item` ppi
        JOIN `tabProduction Plan` pp ON pp.name = ppi.parent
        WHERE pp.docstatus = 1 AND pp.status IN {pp_sql}
    """)
    codes.update(rows)

    # (3) Sub-assembly proper — child WO + same-PP parent that consumes it.
    #     Implements: "will be used in WO but that item itself has WO".
    rows = frappe.db.sql_list(f"""
        SELECT DISTINCT child.production_item
        FROM `tabWork Order` child
        WHERE child.docstatus = 1 AND child.status IN {wo_sql}
          AND IFNULL(child.production_plan, '') != ''
          AND EXISTS (
              SELECT 1
              FROM `tabWork Order` parent
              JOIN `tabBOM Item` bi ON bi.parent = parent.bom_no
              WHERE parent.production_plan = child.production_plan
                AND parent.docstatus = 1
                AND parent.production_item != child.production_item
                AND bi.item_code = child.production_item
          )
    """)
    codes.update(rows)

    # (4) Items with open pending Sales Order line (any month).
    #     Pending = stock_qty − delivered_qty * conversion_factor > 0.
    rows = frappe.db.sql_list(f"""
        SELECT DISTINCT soi.item_code
        FROM `tabSales Order Item` soi
        JOIN `tabSales Order` so ON so.name = soi.parent
        WHERE {so_clause}
          AND (soi.stock_qty - IFNULL(soi.delivered_qty,0) * IFNULL(soi.conversion_factor,1)) > 0
    """)
    codes.update(rows)

    return list(codes)


def _get_uom_conversions(item_codes):
    """Returns {item_code: [{uom, factor}, ...]} for all conversions."""
    if not item_codes:
        return {}
    sql = _sql_in(item_codes)
    rows = frappe.db.sql(f"""
        SELECT parent AS item_code, uom, conversion_factor AS factor
        FROM `tabUOM Conversion Detail`
        WHERE parent IN {sql}
          AND conversion_factor > 0
        ORDER BY conversion_factor ASC
    """, as_dict=True)
    result = {}
    for r in rows:
        result.setdefault(r.item_code, []).append({
            "uom": r.uom, "factor": flt(r.factor)
        })
    return result


def _get_sub_assembly_info(item_codes, wo_statuses):
    """
    Sub-assembly detection — strict definition per user requirement:

      "the item is a child of a parent work order but it itself has open
       work order (identify from production planning doctype with connection)"

    Two qualifying conditions joined together:

      (A) Item is produced by an open child WO that links to a Production Plan
          (`Work Order.production_plan` is not null, status open, docstatus=1).
      (B) The same Production Plan also has a parent WO whose `production_item`
          differs from this item — i.e. there's a real parent → child chain
          via PP, not a stand-alone WO.

    The returned `wo_list` contains the PARENT WOs (the ones consuming this item
    as a sub-assembly) along with shortage qty estimated from BOM × parent qty
    minus current stock — so the user can see "Will be used in PARENT_WO with
    shortage 1200 g".

    Independent (own BOM, no parent linkage) WOs are intentionally NOT marked
    as sub-assembly.
    """
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wo_sql    = _sql_in(wo_statuses)

    # 1. Find PP-linked open WOs grouped by Production Plan.
    pp_wo_rows = frappe.db.sql(f"""
        SELECT wo.name AS wo_name, wo.production_plan, wo.production_item,
               wo.bom_no, wo.qty AS planned_qty, wo.status
        FROM `tabWork Order` wo
        WHERE wo.docstatus = 1
          AND wo.status IN {wo_sql}
          AND IFNULL(wo.production_plan, '') != ''
    """, as_dict=True)

    pp_groups = {}
    for r in pp_wo_rows:
        pp_groups.setdefault(r.production_plan, []).append(r)

    # 2. For each item, find PPs where this item is a CHILD WO and there is at
    #    least one OTHER WO in the same PP with a different production_item
    #    (= the parent).
    sub_asm_set = set()
    parent_wo_map = {ic: [] for ic in item_codes}

    for pp, wos in pp_groups.items():
        items_in_pp = {w.production_item for w in wos}
        for w in wos:
            if w.production_item in item_codes and len(items_in_pp - {w.production_item}) > 0:
                sub_asm_set.add(w.production_item)
                # Record parent WOs for this sub-assembly
                for parent in wos:
                    if parent.production_item != w.production_item:
                        parent_wo_map[w.production_item].append({
                            "wo_name":         parent.wo_name,
                            "bom_no":          parent.bom_no,
                            "parent_item":     parent.production_item,
                            "parent_qty":      flt(parent.planned_qty),
                            "production_plan": pp,
                            "status":          parent.status,
                        })

    # 3. Estimate "shortage qty in this sub-assembly required by each parent WO".
    #    Uses BOM Item.stock_qty × parent_qty / bom.quantity − current stock map
    #    is computed lazily by caller (we just supply the required component qty).
    if sub_asm_set:
        # Bulk fetch BOM line for this component in each parent BOM.
        sub_codes_sql = _sql_in(list(sub_asm_set))
        bom_lines = frappe.db.sql(f"""
            SELECT bi.parent AS bom_no, bi.item_code, bi.stock_qty AS qty_per_unit
            FROM `tabBOM Item` bi
            WHERE bi.item_code IN {sub_codes_sql}
        """, as_dict=True)
        bom_qty_map = {(b.bom_no, b.item_code): flt(b.qty_per_unit) for b in bom_lines}

        # Annotate parent WO entries with required_qty for the sub-assembly.
        for ic, parents in parent_wo_map.items():
            for p in parents:
                qpu = bom_qty_map.get((p["bom_no"], ic), 0)
                p["required_qty"] = round(qpu * p["parent_qty"], 3)

    result = {}
    for code in item_codes:
        result[code] = {
            "is_sub_assembly": code in sub_asm_set,
            "wo_list": parent_wo_map.get(code, []),
        }
    return result


def _get_pp_list(item_code, pp_statuses=None):
    """
    Native ERPNext call: list active Production Plans containing this item
    as a Production Plan Item, plus all child Work Orders for the same plan.
    Used by `get_active_production_plans` modal.

    NOTE: pp.custom_created_by and pp.custom_creation_reason are chaizup_toc
    custom fields installed via fixtures/custom_field.json. We probe each
    column with frappe.db.has_column() so the query still works on bench
    sites where the fixtures haven't been synced yet.
    """
    pp_statuses = pp_statuses or _DEFAULT_PP_STATUSES
    pp_sql      = _sql_in(pp_statuses)

    # ── Probe optional custom columns (graceful when fixtures not synced) ──
    has_created_by      = frappe.db.has_column("Production Plan", "custom_created_by")
    has_creation_reason = frappe.db.has_column("Production Plan", "custom_creation_reason")
    extra_cols = []
    if has_created_by:      extra_cols.append("pp.custom_created_by")
    if has_creation_reason: extra_cols.append("pp.custom_creation_reason")
    extra_select = ("," + ", ".join(extra_cols)) if extra_cols else ""

    pps = frappe.db.sql(f"""
        SELECT DISTINCT pp.name, pp.status, pp.posting_date{extra_select}
        FROM `tabProduction Plan` pp
        JOIN `tabProduction Plan Item` ppi ON ppi.parent = pp.name
        WHERE pp.docstatus = 1
          AND pp.status IN {pp_sql}
          AND ppi.item_code = %s
        ORDER BY pp.posting_date DESC
        LIMIT 25
    """, item_code, as_dict=True)

    if not pps:
        return []
    pp_names_sql = _sql_in([p.name for p in pps])
    wo_rows = frappe.db.sql(f"""
        SELECT name, production_item, qty AS planned_qty, produced_qty,
               status, production_plan, bom_no, planned_start_date
        FROM `tabWork Order`
        WHERE docstatus = 1
          AND production_plan IN {pp_names_sql}
        ORDER BY production_plan, planned_start_date
    """, as_dict=True)

    by_pp = {}
    for w in wo_rows:
        by_pp.setdefault(w.production_plan, []).append({
            "wo_name":            w.name,
            "production_item":    w.production_item,
            "planned_qty":        flt(w.planned_qty),
            "produced_qty":       flt(w.produced_qty),
            "status":             w.status,
            "bom_no":             w.bom_no,
            "is_target_item":     w.production_item == item_code,
            "planned_start_date": str(w.planned_start_date or ""),
        })

    out = []
    for pp in pps:
        out.append({
            "pp_name":          pp.name,
            "status":           pp.status,
            "posting_date":     str(pp.posting_date or ""),
            "created_by":       (pp.get("custom_created_by") if has_created_by else None) or "User",
            "creation_reason":  (pp.get("custom_creation_reason") if has_creation_reason else "") or "",
            "work_orders":      by_pp.get(pp.name, []),
        })
    return out


@frappe.whitelist()
def get_active_production_plans(item_code, pp_statuses=None):
    """
    Returns active Production Plans containing this item (as PPI) AND all WOs
    of those plans grouped by parent/child. Used by the "View Plan" modal.
    """
    if isinstance(pp_statuses, str):
        pp_statuses = frappe.parse_json(pp_statuses) or _DEFAULT_PP_STATUSES
    return {"item_code": item_code, "plans": _get_pp_list(item_code, pp_statuses)}


@frappe.whitelist()
def get_pp_tree(pp_name):
    """
    Hierarchical PP tree for the click-into-PP-id modal. Layers:
      - Each Work Order in the plan (from `tabWork Order.production_plan`).
      - Each WO's BOM components (one level — typical FG → SFG/RM/PM split).
      - For each component, ERPNext-native supply/demand snapshot:
          * required           : qty needed from BOM × WO qty
          * consumed           : Stock Entry Detail.qty already consumed for this WO
          * remaining          : max(required - consumed, 0)
          * stock              : Bin.actual_qty (sum across all warehouses, since
                                 PP doesn't pin a warehouse on the component side)
          * supply_from_po     : Σ open Purchase Order Item qty (purchased items)
          * supply_from_wo     : Σ open Work Order pending qty (manufactured items)
          * supply_from_mr     : Σ open Material Request qty (RM-side)
          * shortage           : max(remaining - stock - supply_from_po
                                       - supply_from_wo - supply_from_mr, 0)
      - Items resolved to manufactured vs purchased via
        `Item.is_purchase_item`, `Item.default_material_request_type`,
        or active default BOM presence.

    Performance note: bulk-fetches all component data with single SQL calls.
    """
    if not pp_name or not frappe.db.exists("Production Plan", pp_name):
        return {"pp_name": pp_name, "found": False, "work_orders": []}

    pp_doc_meta = frappe.db.get_value("Production Plan", pp_name,
        ["status", "posting_date", "company", "for_warehouse"], as_dict=True) or {}

    # Step 1: WOs of this plan
    wo_rows = frappe.db.sql("""
        SELECT name, production_item, qty AS qty_to_manufacture, produced_qty,
               status, bom_no, fg_warehouse, planned_start_date
        FROM `tabWork Order`
        WHERE production_plan = %s AND docstatus = 1
        ORDER BY planned_start_date, name
    """, pp_name, as_dict=True)

    # Step 2: collect all component item_codes across all WO BOMs
    bom_set = list({w.bom_no for w in wo_rows if w.bom_no})
    components_by_bom = {}
    if bom_set:
        bom_in = _sql_in(bom_set)
        rows = frappe.db.sql(f"""
            SELECT bi.parent AS bom_no, bi.item_code, bi.item_name,
                   bi.stock_qty AS qty_per_unit, bi.stock_uom AS uom
            FROM `tabBOM Item` bi
            WHERE bi.parent IN {bom_in}
            ORDER BY bi.parent, bi.idx
        """, as_dict=True)
        for r in rows:
            components_by_bom.setdefault(r.bom_no, []).append({
                "item_code":    r.item_code,
                "item_name":    r.item_name,
                "qty_per_unit": flt(r.qty_per_unit),
                "uom":          r.uom or "",
            })

    all_comp_codes = sorted({c["item_code"] for comps in components_by_bom.values() for c in comps})

    # Step 3: bulk fetch supply data for all components
    consumed_by_wo_item = {}  # {(wo, item_code): qty}
    if wo_rows:
        wo_names = _sql_in([w.name for w in wo_rows])
        if all_comp_codes:
            comp_in = _sql_in(all_comp_codes)
            rows = frappe.db.sql(f"""
                SELECT se.work_order AS wo, sed.item_code, SUM(sed.qty) AS qty
                FROM `tabStock Entry Detail` sed
                JOIN `tabStock Entry` se ON se.name = sed.parent
                WHERE se.docstatus = 1
                  AND se.work_order IN {wo_names}
                  AND sed.item_code IN {comp_in}
                  AND sed.is_finished_item = 0
                GROUP BY se.work_order, sed.item_code
            """, as_dict=True)
            for r in rows:
                consumed_by_wo_item[(r.wo, r.item_code)] = flt(r.qty)

    stock_map = _get_stock(all_comp_codes, [], "physical")
    item_meta = {}
    if all_comp_codes:
        comp_in = _sql_in(all_comp_codes)
        meta_rows = frappe.db.sql(f"""
            SELECT name, item_name, stock_uom,
                   COALESCE(is_purchase_item, 0)     AS is_purchase_item,
                   COALESCE(custom_toc_auto_manufacture, 0) AS auto_manufacture,
                   default_material_request_type
            FROM `tabItem`
            WHERE name IN {comp_in}
        """, as_dict=True)
        item_meta = {r.name: r for r in meta_rows}

        # Open POs
        po_rows = frappe.db.sql(f"""
            SELECT poi.item_code,
                   SUM(GREATEST(poi.stock_qty - IFNULL(poi.received_qty,0)
                       * IFNULL(poi.conversion_factor,1), 0)) AS qty
            FROM `tabPurchase Order Item` poi
            JOIN `tabPurchase Order` po ON po.name = poi.parent
            WHERE po.docstatus = 1
              AND po.status NOT IN ('Closed','Completed','Cancelled','Delivered')
              AND poi.item_code IN {comp_in}
            GROUP BY poi.item_code
        """, as_dict=True)
        po_supply_map = {r.item_code: flt(r.qty) for r in po_rows}

        # Open MRs (Material Request)
        mr_rows = frappe.db.sql(f"""
            SELECT mri.item_code,
                   SUM(GREATEST(mri.stock_qty - IFNULL(mri.ordered_qty,0)
                       * IFNULL(mri.conversion_factor,1), 0)) AS qty
            FROM `tabMaterial Request Item` mri
            JOIN `tabMaterial Request` mr ON mr.name = mri.parent
            WHERE mr.docstatus = 1
              AND mr.status NOT IN ('Stopped','Cancelled','Issued','Received')
              AND mri.item_code IN {comp_in}
            GROUP BY mri.item_code
        """, as_dict=True)
        mr_supply_map = {r.item_code: flt(r.qty) for r in mr_rows}

        # Open WOs producing these components (i.e. they are SFG made elsewhere)
        wo_supply_rows = frappe.db.sql(f"""
            SELECT production_item AS item_code,
                   SUM(GREATEST(qty - IFNULL(produced_qty,0), 0)) AS qty
            FROM `tabWork Order`
            WHERE docstatus = 1
              AND status IN ('Not Started','In Process','Material Transferred')
              AND production_item IN {comp_in}
            GROUP BY production_item
        """, as_dict=True)
        wo_supply_map = {r.item_code: flt(r.qty) for r in wo_supply_rows}

        # Open WO names (for "show work orders" detail)
        wo_detail_rows = frappe.db.sql(f"""
            SELECT name, production_item, qty, produced_qty, status, fg_warehouse
            FROM `tabWork Order`
            WHERE docstatus = 1
              AND status IN ('Not Started','In Process','Material Transferred')
              AND production_item IN {comp_in}
            ORDER BY production_item, planned_start_date
        """, as_dict=True)
        wo_detail_by_item = {}
        for r in wo_detail_rows:
            wo_detail_by_item.setdefault(r.production_item, []).append({
                "wo_name":      r.name,
                "qty":          flt(r.qty),
                "produced_qty": flt(r.produced_qty),
                "pending":      flt(r.qty) - flt(r.produced_qty),
                "status":       r.status,
                "fg_warehouse": r.fg_warehouse or "",
            })

        # Open PO names (for "show po list" detail)
        po_detail_rows = frappe.db.sql(f"""
            SELECT poi.parent AS po, poi.item_code, poi.stock_qty,
                   poi.received_qty, poi.conversion_factor, poi.uom, poi.warehouse,
                   po.transaction_date, po.schedule_date, po.supplier
            FROM `tabPurchase Order Item` poi
            JOIN `tabPurchase Order` po ON po.name = poi.parent
            WHERE po.docstatus = 1
              AND po.status NOT IN ('Closed','Completed','Cancelled','Delivered')
              AND poi.item_code IN {comp_in}
            ORDER BY poi.item_code, po.schedule_date
        """, as_dict=True)
        po_detail_by_item = {}
        for r in po_detail_rows:
            pending = flt(r.stock_qty) - flt(r.received_qty) * flt(r.conversion_factor or 1)
            if pending <= 0:
                continue
            po_detail_by_item.setdefault(r.item_code, []).append({
                "po_name":      r.po,
                "supplier":     r.supplier or "",
                "pending_qty":  pending,
                "uom":          r.uom or "",
                "warehouse":    r.warehouse or "",
                "schedule":     str(r.schedule_date or ""),
            })

        # Open MR names
        mr_detail_rows = frappe.db.sql(f"""
            SELECT mri.parent AS mr, mri.item_code, mri.stock_qty,
                   mri.ordered_qty, mri.conversion_factor, mri.uom, mri.warehouse,
                   mr.transaction_date, mr.material_request_type
            FROM `tabMaterial Request Item` mri
            JOIN `tabMaterial Request` mr ON mr.name = mri.parent
            WHERE mr.docstatus = 1
              AND mr.status NOT IN ('Stopped','Cancelled','Issued','Received')
              AND mri.item_code IN {comp_in}
            ORDER BY mri.item_code, mr.transaction_date DESC
        """, as_dict=True)
        mr_detail_by_item = {}
        for r in mr_detail_rows:
            pending = flt(r.stock_qty) - flt(r.ordered_qty) * flt(r.conversion_factor or 1)
            if pending <= 0:
                continue
            mr_detail_by_item.setdefault(r.item_code, []).append({
                "mr_name":      r.mr,
                "type":         r.material_request_type or "",
                "pending_qty":  pending,
                "uom":          r.uom or "",
                "warehouse":    r.warehouse or "",
            })
    else:
        po_supply_map = mr_supply_map = wo_supply_map = {}
        po_detail_by_item = mr_detail_by_item = wo_detail_by_item = {}

    # Step 4: build the tree
    out_wos = []
    for w in wo_rows:
        comps = components_by_bom.get(w.bom_no, [])
        wo_qty = flt(w.qty_to_manufacture)
        produced = flt(w.produced_qty)
        comp_out = []
        for c in comps:
            required = round(c["qty_per_unit"] * wo_qty, 3)
            consumed = flt(consumed_by_wo_item.get((w.name, c["item_code"]), 0))
            remaining = max(required - consumed, 0)
            stock = flt(stock_map.get(c["item_code"], 0))
            po_qty = flt(po_supply_map.get(c["item_code"], 0))
            mr_qty = flt(mr_supply_map.get(c["item_code"], 0))
            wo_qty_inb = flt(wo_supply_map.get(c["item_code"], 0))
            # Classify supply mode for "show WOs vs POs"
            meta = item_meta.get(c["item_code"]) or {}
            is_purchase = bool(cint(meta.get("is_purchase_item", 0)))
            shortage = max(remaining - stock - po_qty - mr_qty - wo_qty_inb, 0)
            comp_out.append({
                "item_code":         c["item_code"],
                "item_name":         c["item_name"],
                "uom":               c["uom"],
                "qty_per_unit":      c["qty_per_unit"],
                "required":          required,
                "consumed":          consumed,
                "remaining":         remaining,
                "stock":             stock,
                "supply_from_po":    po_qty,
                "supply_from_wo":    wo_qty_inb,
                "supply_from_mr":    mr_qty,
                "shortage":          round(shortage, 3),
                "is_purchase":       is_purchase,
                "supply_pos":        po_detail_by_item.get(c["item_code"], [])[:10],
                "supply_wos":        wo_detail_by_item.get(c["item_code"], [])[:10],
                "supply_mrs":        mr_detail_by_item.get(c["item_code"], [])[:10],
            })
        out_wos.append({
            "wo_name":            w.name,
            "production_item":    w.production_item,
            "bom_no":             w.bom_no,
            "qty_to_manufacture": wo_qty,
            "produced_qty":       produced,
            "remaining_qty":      max(wo_qty - produced, 0),
            "status":             w.status,
            "fg_warehouse":       w.fg_warehouse or "",
            "planned_start_date": str(w.planned_start_date or ""),
            "components":         comp_out,
        })

    return {
        "pp_name":      pp_name,
        "found":        True,
        "status":       pp_doc_meta.get("status"),
        "company":      pp_doc_meta.get("company"),
        "for_warehouse": pp_doc_meta.get("for_warehouse"),
        "posting_date": str(pp_doc_meta.get("posting_date") or ""),
        "work_orders":  out_wos,
    }


def _get_wo_counts(item_codes, wo_statuses, warehouses=None):
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wo_sql    = _sql_in(wo_statuses)
    wh_clause = _wh_clause(warehouses, "fg_warehouse")
    rows = frappe.db.sql(f"""
        SELECT production_item, COUNT(*) AS cnt
        FROM `tabWork Order`
        WHERE docstatus = 1 AND status IN {wo_sql}
          AND production_item IN {codes_sql}
          {wh_clause}
        GROUP BY production_item
    """, as_dict=True)
    return {r.production_item: r.cnt for r in rows}


def _get_planned_qty(item_codes, wo_statuses, warehouses=None):
    """Sum of planned manufacturing qty for open WOs (warehouse-scoped).

    The warehouse here is `Work Order.fg_warehouse` — where the finished
    item is targeted to land. Matches what the user means by "warehouse
    target" on this report.

    DANGER: ERPNext Work Order's "Qty To Manufacture" column is named `qty`
    in the database (label only is "Qty To Manufacture"). Do NOT use
    `qty_to_manufacture` in SQL on `tabWork Order` — that column does not exist
    and raises OperationalError 1054. The fieldname `qty_to_manufacture` exists
    on Production Plan Item, NOT on Work Order.
    """
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wo_sql    = _sql_in(wo_statuses)
    wh_clause = _wh_clause(warehouses, "fg_warehouse")
    rows = frappe.db.sql(f"""
        SELECT production_item,
               SUM(qty) AS total,
               SUM(IFNULL(produced_qty, 0)) AS produced,
               SUM(GREATEST(qty - IFNULL(produced_qty, 0), 0)) AS pending_wo_qty
        FROM `tabWork Order`
        WHERE docstatus = 1 AND status IN {wo_sql}
          AND production_item IN {codes_sql}
          {wh_clause}
        GROUP BY production_item
    """, as_dict=True)
    out = {}
    for r in rows:
        out[r.production_item] = {
            "planned":   flt(r.total),
            "produced":  flt(r.produced),
            "pending":   flt(r.pending_wo_qty),
        }
    return out


def _get_actual_qty(item_codes, month, year, warehouses=None):
    """Sum of finished item qty from Manufacture STEs in the given month.
    Warehouse filter is `t_warehouse` (target warehouse for the FG)."""
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wh_clause = _wh_clause(warehouses, "sed.t_warehouse")
    rows = frappe.db.sql(f"""
        SELECT sed.item_code, SUM(sed.transfer_qty) AS total
        FROM `tabStock Entry Detail` sed
        JOIN `tabStock Entry` se ON se.name = sed.parent
        WHERE se.docstatus = 1
          AND se.stock_entry_type = 'Manufacture'
          AND sed.is_finished_item = 1
          AND MONTH(se.posting_date) = {cint(month)}
          AND YEAR(se.posting_date) = {cint(year)}
          AND sed.item_code IN {codes_sql}
          {wh_clause}
        GROUP BY sed.item_code
    """, as_dict=True)
    return {r.item_code: flt(r.total) for r in rows}


def _get_so_qty(item_codes, so_statuses, month, year, warehouses=None):
    """
    Sum of pending qty in stock_uom from open SOs with delivery_date in given month.

    Warehouse scoping uses BOTH `soi.warehouse` (line-level) AND
    `so.set_warehouse` (header-level), so legacy SOs that only set the header
    warehouse still match. ERPNext-native UOM:
      pending_stock = soi.stock_qty - delivered_qty * conversion_factor
    (matches `production_plan_engine.py` and `wo_kitting_api.py`.)
    """
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    so_clause = _so_status_clause(so_statuses)
    wh_clause = ""
    if warehouses:
        wh_in = _sql_in(warehouses)
        wh_clause = (
            f" AND (COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) IN {wh_in})"
        )
    rows = frappe.db.sql(f"""
        SELECT soi.item_code,
               SUM(soi.stock_qty - IFNULL(soi.delivered_qty,0) * IFNULL(soi.conversion_factor,1)) AS total
        FROM `tabSales Order Item` soi
        JOIN `tabSales Order` so ON so.name = soi.parent
        WHERE {so_clause}
          AND MONTH(soi.delivery_date) = {cint(month)}
          AND YEAR(soi.delivery_date) = {cint(year)}
          AND soi.item_code IN {codes_sql}
          {wh_clause}
        GROUP BY soi.item_code
    """, as_dict=True)
    return {r.item_code: flt(r.total) for r in rows}


def _get_all_pending_so_qty(item_codes, so_statuses, warehouses=None):
    """
    Sum of TOTAL pending SO qty (no month filter) per item, in stock UOM.
    Used by the Excel "Simple Report" sheet — the planner wants to see
    every pending order regardless of delivery month.

    pending_stock = stock_qty - delivered_qty * conversion_factor (clamped >= 0).
    Filtered by selected SO statuses + warehouses (line OR header).
    """
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    so_clause = _so_status_clause(so_statuses)
    wh_clause = ""
    if warehouses:
        wh_in = _sql_in(warehouses)
        wh_clause = f" AND (COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) IN {wh_in})"
    rows = frappe.db.sql(f"""
        SELECT soi.item_code,
               SUM(GREATEST(soi.stock_qty - IFNULL(soi.delivered_qty,0)
                   * IFNULL(soi.conversion_factor,1), 0)) AS total
        FROM `tabSales Order Item` soi
        JOIN `tabSales Order` so ON so.name = soi.parent
        WHERE {so_clause}
          AND soi.item_code IN {codes_sql}
          {wh_clause}
        GROUP BY soi.item_code
    """, as_dict=True)
    return {r.item_code: flt(r.total) for r in rows}


def _get_has_open_so_map(item_codes, so_statuses, warehouses=None):
    """Returns {item_code: True} for items with at least one pending SO line."""
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    so_clause = _so_status_clause(so_statuses)
    wh_clause = ""
    if warehouses:
        wh_in = _sql_in(warehouses)
        wh_clause = f" AND (COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) IN {wh_in})"
    rows = frappe.db.sql(f"""
        SELECT DISTINCT soi.item_code
        FROM `tabSales Order Item` soi
        JOIN `tabSales Order` so ON so.name = soi.parent
        WHERE {so_clause}
          AND soi.item_code IN {codes_sql}
          AND (soi.stock_qty - IFNULL(soi.delivered_qty,0) * IFNULL(soi.conversion_factor,1)) > 0
          {wh_clause}
    """)
    return {r[0]: True for r in rows}


def _get_dispatch_qty(item_codes, month, year, warehouses=None):
    """Sum of stock_qty from submitted Delivery Notes in given month
    (warehouse-scoped via Delivery Note Item.warehouse)."""
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wh_clause = _wh_clause(warehouses, "dni.warehouse")
    rows = frappe.db.sql(f"""
        SELECT dni.item_code, SUM(dni.stock_qty) AS total
        FROM `tabDelivery Note Item` dni
        JOIN `tabDelivery Note` dn ON dn.name = dni.parent
        WHERE dn.docstatus = 1
          AND MONTH(dn.posting_date) = {cint(month)}
          AND YEAR(dn.posting_date) = {cint(year)}
          AND dni.item_code IN {codes_sql}
          {wh_clause}
        GROUP BY dni.item_code
    """, as_dict=True)
    return {r.item_code: flt(r.total) for r in rows}


def _get_projection_qty(item_codes, month_name, year, warehouses=None):
    """Sum of qty_in_stock_uom from current Sales Projection
    (Sales Projection is warehouse-scoped via `source_warehouse`)."""
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wh_clause = _wh_clause(warehouses, "sp.source_warehouse")
    rows = frappe.db.sql(f"""
        SELECT spi.item, SUM(spi.qty_in_stock_uom) AS total
        FROM `tabSales Projected Items` spi
        JOIN `tabSales Projection` sp ON sp.name = spi.parent
        WHERE sp.projection_month = {frappe.db.escape(month_name)}
          AND sp.projection_year = {cint(year)}
          AND sp.docstatus IN (0, 1)
          AND spi.item IN {codes_sql}
          {wh_clause}
        GROUP BY spi.item
    """, as_dict=True)
    return {r.item: flt(r.total) for r in rows}


def _get_total_sales_qty(item_codes, month, year, warehouses=None):
    """
    Total sales qty (dispatched + pending) booked for the given month.
    Submitted SOs only (docstatus = 1).
    """
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wh_clause = ""
    if warehouses:
        wh_in = _sql_in(warehouses)
        wh_clause = f" AND (COALESCE(NULLIF(soi.warehouse,''), so.set_warehouse) IN {wh_in})"
    rows = frappe.db.sql(f"""
        SELECT soi.item_code, SUM(soi.stock_qty) AS total
        FROM `tabSales Order Item` soi
        JOIN `tabSales Order` so ON so.name = soi.parent
        WHERE so.docstatus = 1
          AND MONTH(soi.delivery_date) = {cint(month)}
          AND YEAR(soi.delivery_date) = {cint(year)}
          AND soi.item_code IN {codes_sql}
          {wh_clause}
        GROUP BY soi.item_code
    """, as_dict=True)
    return {r.item_code: flt(r.total) for r in rows}


def _get_active_pp_count(item_codes, pp_statuses=None):
    """
    Count of active Production Plans per item — joins via
    `tabProduction Plan Item.item_code` (ERPNext native child table).
    """
    if not item_codes:
        return {}
    pp_statuses = pp_statuses or _DEFAULT_PP_STATUSES
    codes_sql = _sql_in(item_codes)
    pp_sql    = _sql_in(pp_statuses)
    rows = frappe.db.sql(f"""
        SELECT ppi.item_code, COUNT(DISTINCT pp.name) AS cnt
        FROM `tabProduction Plan Item` ppi
        JOIN `tabProduction Plan` pp ON pp.name = ppi.parent
        WHERE pp.docstatus = 1
          AND pp.status IN {pp_sql}
          AND ppi.item_code IN {codes_sql}
        GROUP BY ppi.item_code
    """, as_dict=True)
    return {r.item_code: cint(r.cnt) for r in rows}


def _get_wo_pp_groups(item_codes, wo_statuses, warehouses=None):
    """
    Smart Production Plan grouping for the Overview "Open WOs" cell.

    POR-007 (2026-05-04): The "Production Plans" modal makes it obvious that
    multiple Work Orders share a parent PP — but on the main grid the user
    can't see that signal until they open the modal. This helper returns,
    per FG item:
      - `wo_pp_count`     : number of distinct PPs that THIS item's open WOs
                            are part of
      - `wo_pp_names`     : list of PP names (for tooltip)
      - `wo_pp_siblings`  : count of OTHER items that also have open WOs
                            in those same PPs (i.e. how many products are
                            being co-planned with this one)

    The frontend renders a small chip beside the WO count whenever
    `wo_pp_count > 0` so users know at a glance which items are "co-planned".
    """
    if not item_codes:
        return {}
    wo_statuses = wo_statuses or _DEFAULT_WO_STATUSES
    codes_sql = _sql_in(item_codes)
    wo_sql    = _sql_in(wo_statuses)
    wh_clause = _wh_clause(warehouses, "fg_warehouse")

    # Step 1: collect (production_item, production_plan) pairs for open WOs.
    rows = frappe.db.sql(f"""
        SELECT production_item, production_plan
        FROM `tabWork Order`
        WHERE docstatus = 1 AND status IN {wo_sql}
          AND production_plan IS NOT NULL AND production_plan != ''
          AND production_item IN {codes_sql}
          {wh_clause}
    """, as_dict=True)

    # item -> set(pp_name); pp -> set(item_code)
    item_to_pps    = {}
    pp_to_items    = {}
    for r in rows:
        item_to_pps.setdefault(r.production_item, set()).add(r.production_plan)
        pp_to_items.setdefault(r.production_plan,  set()).add(r.production_item)

    out = {}
    for code, pps in item_to_pps.items():
        siblings = set()
        for pp in pps:
            siblings |= pp_to_items.get(pp, set())
        siblings.discard(code)
        out[code] = {
            "wo_pp_count":    len(pps),
            "wo_pp_names":    sorted(pps),
            "wo_pp_siblings": len(siblings),
        }
    return out


def _get_stock(item_codes, warehouses, stock_mode="physical"):
    """
    Returns {item_code: qty} for current stock.
    stock_mode='physical'  → actual_qty only
    stock_mode='expected'  → actual_qty + ordered_qty + planned_qty
    """
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    wh_clause = ""
    if warehouses:
        wh_sql    = _sql_in(warehouses)
        wh_clause = f"AND warehouse IN {wh_sql}"

    if stock_mode == "expected":
        qty_expr = "SUM(actual_qty + ordered_qty + planned_qty)"
    else:
        qty_expr = "SUM(actual_qty)"

    rows = frappe.db.sql(f"""
        SELECT item_code, {qty_expr} AS stock
        FROM `tabBin`
        WHERE item_code IN {codes_sql}
          {wh_clause}
        GROUP BY item_code
    """, as_dict=True)
    return {r.item_code: flt(r.stock) for r in rows}


def _get_active_bom(item_codes):
    """Returns {item_code: bom_name} for the active default BOM."""
    if not item_codes:
        return {}
    codes_sql = _sql_in(item_codes)
    rows = frappe.db.sql(f"""
        SELECT item, name AS bom_name
        FROM `tabBOM`
        WHERE item IN {codes_sql}
          AND is_active = 1 AND is_default = 1 AND docstatus = 1
        ORDER BY modified DESC
    """, as_dict=True)
    # one BOM per item (first match wins)
    result = {}
    for r in rows:
        if r.item not in result:
            result[r.item] = r.bom_name
    return result


def _get_shortage_summary(item_codes, bom_map, stock_map, planned_qty_map=None,
                          warehouses=None, stock_mode="physical"):
    """
    Shallow BOM walk — check if any component is short.
    Returns {item_code: {has_shortage: bool, short_components: [...]}}

    POR-006 (2026-05-04): Two bugs fixed.
      1. Component stocks were missing. The `stock_map` argument only contains
         stocks for FG/qualifying items, NOT for their components. The old
         code did `stock_map.get(comp_code, 0)` → always 0 for components →
         every BOM line marked short (false positive). The modal
         (`get_shortage_detail`) didn't have this bug because it re-fetches
         component stock via `_get_stock(comp_codes, ...)`. Fix: build a
         component-level stock map here and merge with the FG stock_map.
      2. Required qty did not multiply by Work Order qty. Old code compared
         per-unit BOM qty (e.g. 5g of sugar per 1 unit FG) against full stock,
         so 100g of sugar would never look short for ANY FG. The modal
         multiplies by the WO `qty` (e.g. 5g × 200 units = 1000g). Fix: use
         `pending_wo_qty` from `planned_qty_map`. If no open WOs, fall back
         to 1 unit (treat the BOM batch_qty as the comparison baseline).

    Each short_component carries `item_name` and `in_stock` so the frontend
    can render a complete shortage line everywhere it appears.
    """
    result = {}
    planned_qty_map = planned_qty_map or {}

    # ── Build a stock map that covers component items, not just FG ──
    component_codes = set()
    for code in item_codes:
        bom = bom_map.get(code, "")
        if not bom:
            continue
        for c in _bom_one_level(bom):
            if c.get("item_code"):
                component_codes.add(c["item_code"])
    component_stock = (
        _get_stock(list(component_codes), warehouses or [], stock_mode)
        if component_codes else {}
    )
    # FG stock_map gets overridden by component stock for component codes.
    # (Some FGs are also components of other FGs — SFGs — and that's fine,
    #  the actual on-hand stock is the same number either way.)
    combined_stock = {**(stock_map or {}), **component_stock}

    for code in item_codes:
        bom = bom_map.get(code, "")
        if not bom:
            result[code] = {"has_shortage": False, "short_components": []}
            continue
        components = _bom_one_level(bom)
        # Use total pending WO qty (across all open WOs of this FG). When
        # there are no WOs, compare against 1-unit BOM qty so the column
        # still answers "would a 1-unit run be possible right now?"
        pending_wo_qty = flt(planned_qty_map.get(code, {}).get("pending", 0))
        qty_to_produce = pending_wo_qty if pending_wo_qty > 0 else 1
        short = []
        for comp in components:
            req_qty = flt(comp["qty"]) * qty_to_produce
            in_stk  = flt(combined_stock.get(comp["item_code"], 0))
            if in_stk < req_qty:
                short.append({
                    "item_code": comp["item_code"],
                    "item_name": comp.get("item_name", ""),
                    "required":  round(req_qty, 3),
                    "in_stock":  round(in_stk, 3),
                    "stock":     round(in_stk, 3),   # alias — JS uses both names
                    "shortage":  round(req_qty - in_stk, 3),
                    "uom":       comp["uom"],
                })
        result[code] = {
            "has_shortage":     len(short) > 0,
            "short_components": short[:5],  # top 5 for overview
            "qty_to_produce":   qty_to_produce,  # for transparency / debug
        }
    return result


def _get_possible_qty(item_codes, bom_map, stock_map):
    """
    For each item with a BOM, compute max producible qty from current stock
    (shallow BOM walk — immediate components only).
    """
    result = {}
    for code in item_codes:
        bom = bom_map.get(code, "")
        if not bom:
            result[code] = 0
            continue
        components = _bom_one_level(bom)
        if not components:
            result[code] = 0
            continue
        max_units = float("inf")
        for comp in components:
            req_per_unit = flt(comp["qty"])
            if req_per_unit <= 0:
                continue
            in_stk = flt(stock_map.get(comp["item_code"], 0))
            supportable = in_stk / req_per_unit
            max_units = min(max_units, supportable)
        result[code] = round(max_units if max_units != float("inf") else 0, 2)
    return result


def _get_cost_summary(item_codes, curr_month, curr_year, bom_map):
    """Lightweight cost summary for display in main table."""
    result = {}
    for code in item_codes:
        bom = bom_map.get(code, "")
        bom_total = 0.0
        if bom:
            row = frappe.db.sql("""
                SELECT SUM(amount) FROM `tabBOM Item` WHERE parent = %s
            """, bom)
            bom_total = flt(row[0][0]) if row else 0

        act_row = frappe.db.sql("""
            SELECT SUM(sed.basic_amount)
            FROM `tabStock Entry Detail` sed
            JOIN `tabStock Entry` se ON se.name = sed.parent
            WHERE se.docstatus = 1
              AND se.stock_entry_type = 'Manufacture'
              AND sed.is_finished_item = 0
              AND MONTH(se.posting_date) = %s
              AND YEAR(se.posting_date) = %s
              AND se.work_order IN (
                  SELECT name FROM `tabWork Order`
                  WHERE production_item = %s AND docstatus = 1
              )
        """, (curr_month, curr_year, code))
        act_total = flt(act_row[0][0]) if act_row else 0

        var_pct = 0
        if bom_total > 0:
            var_pct = round((act_total - bom_total) / bom_total * 100, 1)

        result[code] = {
            "bom_total":    round(bom_total, 2),
            "actual_total": round(act_total, 2),
            "variance_pct": var_pct,
        }
    return result


def _bom_one_level(bom_name):
    """Fetch immediate (1-level) BOM components."""
    if not bom_name:
        return []
    rows = frappe.db.sql("""
        SELECT bi.item_code, bi.item_name, bi.stock_qty AS qty, bi.stock_uom AS uom
        FROM `tabBOM Item` bi
        WHERE bi.parent = %s
          AND bi.parenttype = 'BOM'
        ORDER BY bi.idx
    """, bom_name, as_dict=True)
    return [{
        "item_code": r.item_code,
        "item_name": r.item_name,
        "qty":       flt(r.qty),
        "uom":       r.uom or "Nos",
    } for r in rows]


def _walk_bom_shallow(bom_name, qty_to_produce, warehouses, stock_mode, max_depth=3, depth=0):
    """Recursive BOM walk returning flat component list with shortage info."""
    if not bom_name or depth > max_depth:
        return []
    components = _bom_one_level(bom_name)
    result = []
    all_codes = [c["item_code"] for c in components]
    stock_map  = _get_stock(all_codes, warehouses, stock_mode) if all_codes else {}

    for comp in components:
        required = flt(comp["qty"]) * flt(qty_to_produce)
        in_stk   = flt(stock_map.get(comp["item_code"], 0))
        shortage = max(0, required - in_stk)
        sub_bom  = _get_active_bom([comp["item_code"]]).get(comp["item_code"], "")

        # Determine if this is itself an SFG with sub-components
        sub_comps = []
        if sub_bom and depth < max_depth:
            sub_comps = _walk_bom_shallow(sub_bom, required, warehouses, stock_mode,
                                          max_depth, depth + 1)

        open_docs = _get_open_docs_for_item(comp["item_code"])

        result.append({
            "item_code":  comp["item_code"],
            "item_name":  comp["item_name"],
            "uom":        comp["uom"],
            "required":   round(required, 3),
            "in_stock":   round(in_stk, 3),
            "shortage":   round(shortage, 3),
            "open_docs":  open_docs,
            "sub_comps":  sub_comps,
            "depth":      depth,
        })
    return result


def _walk_bom_deep(bom_name, qty_to_produce, warehouses, stock_mode, depth=0, max_depth=4):
    """Full BOM walk for shortage detail modal."""
    return _walk_bom_shallow(bom_name, qty_to_produce, warehouses, stock_mode,
                              max_depth=max_depth, depth=depth)


def _get_open_docs_for_item(item_code, po_statuses=None):
    """Returns open POs and MRs for a component item (for stage badge).

    `po_statuses` defaults to the user-configurable pending list — if not
    supplied, falls back to the module-level default. We never hardcode the
    universe of statuses anywhere; this matches the user-supplied filter.
    """
    docs = []
    po_statuses = po_statuses or _DEFAULT_PO_STATUSES
    po_sql = _sql_in(po_statuses)
    po_rows = frappe.db.sql(f"""
        SELECT poi.parent AS doc, 'Purchase Order' AS type,
               poi.qty, poi.received_qty, poi.uom
        FROM `tabPurchase Order Item` poi
        JOIN `tabPurchase Order` po ON po.name = poi.parent
        WHERE po.docstatus = 1
          AND po.status IN {po_sql}
          AND poi.item_code = %s
        LIMIT 5
    """, item_code, as_dict=True)
    docs.extend([{
        "type": "Purchase Order",
        "name": r.doc,
        "qty":  flt(r.qty),
        "received": flt(r.received_qty),
        "uom":  r.uom,
    } for r in po_rows])

    # Open MRs — keep the canonical "open" definition (Submitted +
    # Partially Ordered). If the user later wants this configurable too,
    # accept an `mr_statuses` arg here.
    mr_rows = frappe.db.sql("""
        SELECT mri.parent AS doc, mri.qty, mri.uom
        FROM `tabMaterial Request Item` mri
        JOIN `tabMaterial Request` mr ON mr.name = mri.parent
        WHERE mr.docstatus = 1
          AND mr.status IN ('Submitted', 'Partially Ordered', 'Pending')
          AND mri.item_code = %s
        LIMIT 5
    """, item_code, as_dict=True)
    docs.extend([{
        "type": "Material Request",
        "name": r.doc,
        "qty":  flt(r.qty),
        "uom":  r.uom,
    } for r in mr_rows])

    return docs


def _get_batch_consumption(wo_name):
    """Return batch-level consumption detail for a Work Order."""
    rows = frappe.db.sql("""
        SELECT sed.item_code, sed.item_name, sed.batch_no,
               SUM(sed.qty) AS qty, sed.uom, AVG(sed.valuation_rate) AS rate
        FROM `tabStock Entry Detail` sed
        JOIN `tabStock Entry` se ON se.name = sed.parent
        WHERE se.docstatus = 1
          AND se.stock_entry_type IN ('Manufacture', 'Material Transfer for Manufacture')
          AND se.work_order = %s
          AND sed.is_finished_item = 0
        GROUP BY sed.item_code, sed.batch_no, sed.uom
        ORDER BY sed.item_code, sed.batch_no
    """, wo_name, as_dict=True)
    return [{
        "item_code": r.item_code,
        "item_name": r.item_name,
        "batch_no":  r.batch_no or "—",
        "qty":       flt(r.qty),
        "uom":       r.uom,
        "rate":      flt(r.rate),
        "amount":    round(flt(r.qty) * flt(r.rate), 2),
    } for r in rows]


def _get_sub_assembly_wos(production_plan_name):
    """Return child Work Orders linked via Production Plan."""
    if not production_plan_name:
        return []
    # NOTE: Work Order column is `qty`; aliased to `qty_to_manufacture` for clarity.
    rows = frappe.db.sql("""
        SELECT name, production_item, qty AS qty_to_manufacture, status, produced_qty
        FROM `tabWork Order`
        WHERE production_plan = %s AND docstatus = 1
        LIMIT 10
    """, production_plan_name, as_dict=True)
    return [{
        "wo_name":           r.name,
        "production_item":   r.production_item,
        "qty_to_manufacture":flt(r.qty_to_manufacture),
        "status":            r.status,
        "produced_qty":      flt(r.produced_qty),
    } for r in rows]


def _get_valuation_rate(item_code):
    """Get current valuation rate for an item (for sorting shortages by value)."""
    return flt(frappe.db.get_value("Item", item_code, "standard_rate") or 0)


def _classify_item_type(item_row, is_sub_assembly):
    """Classify item as FG, SFG, RM, PM, or Other."""
    if cint(item_row.auto_manufacture):
        return "SFG" if is_sub_assembly else "FG"
    if cint(item_row.is_purchase_item):
        if "packaging" in (item_row.item_group or "").lower():
            return "PM"
        return "RM"
    return "Other"
